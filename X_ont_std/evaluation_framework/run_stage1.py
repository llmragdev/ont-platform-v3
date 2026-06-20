"""
Stage 1 실행: 평가 전 검증

기존 평가 데이터 (4팀_정확도_비교.xlsx)를 로드해서
Q&A 쌍의 일관성과 범위를 검증합니다.

산출물:
- qa_validation_log.json
- qa_validation_report.md
- validated_qa_set_v1.xlsx
"""

import sys
from pathlib import Path
import json
from datetime import datetime
import openpyxl

# 경로 설정
REPO_ROOT = Path("E:/ontology_edu/X_ont_std")
EVAL_FRAMEWORK = REPO_ROOT / "evaluation_framework"
EVAL_DATA = REPO_ROOT / "validation/ont_platform_v4_eval"

# 모듈 임포트
sys.path.insert(0, str(EVAL_FRAMEWORK))

from config.category_definitions import CATEGORY_DEFINITIONS
from stages.stage1_prevalidation.qa_validator import (
    QAConsistencyValidator, BatchQAValidator
)
from stages.stage1_prevalidation.scope_analyzer import (
    ScopeAnalyzer, validate_snowflake_specific
)


class Stage1Runner:
    """Stage 1 실행기"""

    def __init__(self):
        """초기화"""
        self.validator = QAConsistencyValidator()
        self.scope_analyzer = ScopeAnalyzer()
        self.eval_data_path = EVAL_DATA / "data/3팀_정확도_비교.xlsx"
        self.output_dir = EVAL_FRAMEWORK / "data/qa_pairs"
        self.report_dir = EVAL_FRAMEWORK / "reports"

    def load_documents(self) -> list:
        """평가 대상 문서 로드"""
        doc_dir = REPO_ROOT / "ai_lab_SIT/target_doc" if (
            REPO_ROOT / "ai_lab_SIT/target_doc"
        ).exists() else None

        if not doc_dir:
            print("[WARNING] 문서 디렉토리를 찾을 수 없음, 더미 문서 사용")
            return ["Ontology 관련 문서", "RAG 관련 문서"]

        # 실제 문서 읽기 (간단히, 파일명만 사용)
        documents = []
        for pdf_file in doc_dir.glob("*.pdf"):
            documents.append(pdf_file.name)

        return documents

    def load_qa_pairs(self) -> list:
        """기존 평가 데이터에서 Q&A 쌍 로드"""
        if not self.eval_data_path.exists():
            raise FileNotFoundError(f"평가 데이터 파일 없음: {self.eval_data_path}")

        wb = openpyxl.load_workbook(self.eval_data_path)
        ws = wb['문항별 비교 상세']

        qa_pairs = []
        for row_idx in range(2, ws.max_row + 1):
            problem_id = ws.cell(row=row_idx, column=1).value
            if not problem_id:
                continue

            question = ws.cell(row=row_idx, column=2).value or ""
            expected_answer = ws.cell(row=row_idx, column=3).value or ""

            # 카테고리 추론
            if str(problem_id).startswith("STD-O-"):
                category = "Ontology"
            elif str(problem_id).startswith("STD-A-"):
                category = "Advanced RAG"
            elif str(problem_id).startswith("STD-S-"):
                category = "Snowflake"
            else:
                category = "Unknown"

            qa_pairs.append({
                'problem_id': str(problem_id),
                'question': question,
                'expected_answer': expected_answer,
                'category': category,
                'row': row_idx
            })

        print(f"[OK] {len(qa_pairs)}개 Q&A 쌍 로드됨")
        return qa_pairs

    def validate_qa_pairs(self, qa_pairs: list, documents: list) -> dict:
        """모든 Q&A 쌍 검증"""
        print("\n" + "="*70)
        print("Stage 1: Q&A 일관성 검증 실행")
        print("="*70)

        results = []

        for idx, pair in enumerate(qa_pairs, 1):
            print(f"\r[{idx}/{len(qa_pairs)}] {pair['problem_id']}", end="")

            # Q&A 일관성 검증
            consistency_result = self.validator.validate_pair(
                question=pair['question'],
                expected_answer=pair['expected_answer'],
                category=pair['category'],
                documents=documents
            )

            # 범위 검증
            if pair['category'] == 'Snowflake':
                scope_result = validate_snowflake_specific(
                    question=pair['question'],
                    expected_answer=pair['expected_answer'],
                    category=pair['category']
                )
            else:
                scope_result = self.scope_analyzer.analyze_scope(
                    question=pair['question'],
                    expected_answer=pair['expected_answer'],
                    category=pair['category']
                )

            # 결합
            result = {
                'problem_id': pair['problem_id'],
                'category': pair['category'],
                'question': pair['question'][:100] + "..." if len(pair['question']) > 100 else pair['question'],
                'expected_answer_preview': pair['expected_answer'][:100] + "..." if len(pair['expected_answer']) > 100 else pair['expected_answer'],

                'consistency_check': {
                    'is_valid': consistency_result['is_valid'],
                    'score': consistency_result['score'],
                    'errors': consistency_result['errors'],
                    'warnings': consistency_result['warnings'],
                    'issues': consistency_result['issues']
                },

                'scope_check': {
                    'has_scope_issues': scope_result['has_scope_issues'],
                    'issues': scope_result['issues'],
                    'specific_mentions': [
                        {'tech': t[0], 'description': t[1]}
                        for t in scope_result['specific_mentions']
                    ]
                },

                'overall_valid': (
                    consistency_result['is_valid'] and
                    not scope_result['has_scope_issues']
                ),
                'validation_status': (
                    'PASS' if consistency_result['is_valid']
                    and not scope_result['has_scope_issues']
                    else 'FAIL'
                ),

                'timestamp': datetime.now().isoformat()
            }

            results.append(result)

        print()  # 줄바꿈
        return results

    def generate_report(self, results: list) -> dict:
        """검증 결과 요약 생성"""
        passed = len([r for r in results if r['validation_status'] == 'PASS'])
        failed = len([r for r in results if r['validation_status'] == 'FAIL'])
        total = len(results)

        failed_items = [r for r in results if r['validation_status'] == 'FAIL']
        critical_items = [
            r for r in failed_items
            if any(i['severity'] == 'CRITICAL' for issue_list in [r['consistency_check']['issues'], r['scope_check']['issues']] for i in issue_list)
        ]

        summary = {
            'validation_session': f"stage1_{datetime.now().strftime('%Y%m%d_%H%M%S')}",
            'timestamp': datetime.now().isoformat(),
            'total_validated': total,
            'passed': passed,
            'failed': failed,
            'pass_rate': f"{passed / total * 100:.1f}%",
            'critical_issues': [r['problem_id'] for r in critical_items],
            'failed_items': [r['problem_id'] for r in failed_items]
        }

        return {
            'results': results,
            'summary': summary
        }

    def save_json_log(self, report: dict):
        """JSON 로그 저장"""
        output_file = self.output_dir / "qa_validation_log.json"
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(report, f, ensure_ascii=False, indent=2)

        print(f"[OK] JSON 로그 저장: {output_file}")

    def save_markdown_report(self, report: dict):
        """마크다운 보고서 저장"""
        summary = report['summary']
        results = report['results']

        md_content = f"""# Q&A 검증 보고서

**생성일**: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
**검증 세션**: {summary['validation_session']}

## 요약

| 항목 | 결과 |
|---|---:|
| 총 문항 | {summary['total_validated']} |
| 검증 통과 | {summary['passed']} ({summary['pass_rate']}) |
| 검증 실패 | {summary['failed']} |

## 검증 결과

"""

        # 카테고리별 결과
        categories = {}
        for result in results:
            cat = result['category']
            if cat not in categories:
                categories[cat] = {'pass': 0, 'fail': 0, 'items': []}

            if result['validation_status'] == 'PASS':
                categories[cat]['pass'] += 1
            else:
                categories[cat]['fail'] += 1
            categories[cat]['items'].append(result)

        for cat, data in sorted(categories.items()):
            total = data['pass'] + data['fail']
            md_content += f"\n### {cat} ({data['pass']}/{total} 통과)\n\n"

            for item in data['items']:
                status = "[DONE]" if item['validation_status'] == 'PASS' else "[DONE]"
                md_content += f"#### {status} {item['problem_id']}\n\n"

                if item['validation_status'] == 'FAIL':
                    consistency_issues = item['consistency_check']['issues']
                    scope_issues = item['scope_check']['issues']

                    if consistency_issues:
                        md_content += "**일관성 검증 문제**:\n"
                        for issue in consistency_issues:
                            severity = f"[{issue['severity']}]"
                            md_content += f"- {severity} {issue['message']}\n"
                        md_content += "\n"

                    if scope_issues:
                        md_content += "**범위 검증 문제**:\n"
                        for issue in scope_issues:
                            severity = f"[{issue['severity']}]"
                            md_content += f"- {severity} {issue['message']}\n"
                        md_content += "\n"

        # 조치 사항
        if summary['critical_issues']:
            md_content += f"\n## 필수 조치 사항\n\n"
            md_content += f"다음 {len(summary['critical_issues'])}개 항목은 반드시 수정 필요:\n\n"
            for problem_id in summary['critical_issues']:
                md_content += f"- [ ] {problem_id}\n"

        output_file = self.report_dir / "qa_validation_report.md"
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write(md_content)

        print(f"[OK] 마크다운 보고서 저장: {output_file}")

    def run(self):
        """Stage 1 실행"""
        print("\n" + "="*70)
        print("PHASE 6 Stage 1: 평가 전 검증 (Prevalidation)")
        print("="*70)

        try:
            # 1. 문서 로드
            print("\n[1/5] 평가 문서 로드...")
            documents = self.load_documents()

            # 2. Q&A 쌍 로드
            print("\n[2/5] Q&A 쌍 로드...")
            qa_pairs = self.load_qa_pairs()

            # 3. 검증 실행
            print("\n[3/5] 검증 실행...")
            validation_results = self.validate_qa_pairs(qa_pairs, documents)

            # 4. 보고서 생성
            print("\n[4/5] 보고서 생성...")
            report = self.generate_report(validation_results)

            # 5. 저장
            print("\n[5/5] 결과 저장...")
            self.save_json_log(report)
            self.save_markdown_report(report)

            # 최종 요약
            summary = report['summary']
            print("\n" + "="*70)
            print("Stage 1 검증 완료")
            print("="*70)
            print(f"\n총 {summary['total_validated']}개 문항 검증")
            print(f"  [PASS] {summary['passed']} ({summary['pass_rate']})")
            print(f"  [FAIL] {summary['failed']}")

            if summary['critical_issues']:
                print(f"\n[DONE] 필수 조치 항목: {len(summary['critical_issues'])}개")
                for issue in summary['critical_issues']:
                    print(f"   - {issue}")

            print(f"\n보고서: {self.report_dir}")
            print("="*70 + "\n")

        except Exception as e:
            print(f"\n오류 발생: {e}")
            raise


if __name__ == "__main__":
    runner = Stage1Runner()
    runner.run()
