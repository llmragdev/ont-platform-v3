"""
Stage 2 실행: 평가 중 검증

Stage 1에서 검증된 Q&A를 기반으로
평가 중에 3개 체크포인트를 적용합니다:

1. Checkpoint 1: 답변 생성 직후
   - 답변 범위, 근거, 길이 검증

2. Checkpoint 2: 정확도 채점 전
   - 기대값/실제값 비교 가능성 검증
   - 채점 기준 적절성 검증

3. Checkpoint 3: 평가 결과 검증
   - Q&A 쌍 유효성 검증
   - 점수 타당성 검증

산출물:
- checkpoint_log.json
- evaluation_checkpoint_report.md
- constraint_applied_results.xlsx
"""

import sys
from pathlib import Path
import json
from datetime import datetime
import openpyxl
from typing import Dict, List

# 경로 설정
REPO_ROOT = Path("E:/ontology_edu/X_ont_std")
EVAL_FRAMEWORK = REPO_ROOT / "evaluation_framework"
EVAL_DATA = REPO_ROOT / "validation/ont_platform_v4_eval"

# 모듈 임포트
sys.path.insert(0, str(EVAL_FRAMEWORK))

from config.category_definitions import CATEGORY_DEFINITIONS
from stages.stage2_evaluation.constraint_enforcer import (
    OntologyConstraintEnforcer, ConstraintEvaluationMetrics
)
from stages.stage2_evaluation.checkpoints import EvaluationCheckpoints


class Stage2Runner:
    """Stage 2 실행기"""

    def __init__(self):
        """초기화"""
        # 온톨로지 개념 로드 (더미)
        self.ontology_concepts = [
            'ontology', 'concept', 'relationship', 'knowledge_graph',
            'rag', 'retrieval', 'embedding', 'vector', 'metadata',
            'semantic', 'rdf', 'owl', 'entity', 'class', 'property'
        ]

        self.constraint_enforcer = OntologyConstraintEnforcer(
            self.ontology_concepts, CATEGORY_DEFINITIONS
        )
        self.checkpoints = EvaluationCheckpoints(CATEGORY_DEFINITIONS)

        self.eval_data_path = EVAL_DATA / "reports/4팀_정확도_비교.xlsx"
        self.output_dir = EVAL_FRAMEWORK / "data"
        self.report_dir = EVAL_FRAMEWORK / "reports"

    def load_evaluation_data(self) -> list:
        """기존 평가 데이터 로드"""
        if not self.eval_data_path.exists():
            raise FileNotFoundError(f"평가 데이터 파일 없음: {self.eval_data_path}")

        wb = openpyxl.load_workbook(self.eval_data_path)
        ws = wb['문항별 비교 상세']

        qa_data = []
        for row_idx in range(2, ws.max_row + 1):
            problem_id = ws.cell(row=row_idx, column=1).value
            if not problem_id:
                continue

            question = ws.cell(row=row_idx, column=2).value or ""
            expected_answer = ws.cell(row=row_idx, column=3).value or ""
            team4_answer = ws.cell(row=row_idx, column=10).value or ""
            team4_accuracy = ws.cell(row=row_idx, column=11).value

            # 카테고리 추론
            if str(problem_id).startswith("STD-O-"):
                category = "Ontology"
            elif str(problem_id).startswith("STD-A-"):
                category = "Advanced RAG"
            elif str(problem_id).startswith("STD-S-"):
                category = "Snowflake"
            else:
                category = "Unknown"

            qa_data.append({
                'problem_id': str(problem_id),
                'question': question,
                'expected_answer': expected_answer,
                'team4_answer': team4_answer,
                'team4_original_accuracy': team4_accuracy,
                'category': category,
                'row': row_idx
            })

        print(f"[OK] {len(qa_data)}개 평가 데이터 로드됨")
        return qa_data

    def simulate_team4_retrieval(self, question: str,
                                category: str) -> Dict:
        """Team4 (ont_platform v4)의 검색 결과 시뮬레이션"""
        # 실제로는 ont_platform v4 API 호출
        # 여기서는 더미 데이터 반환

        if category == "Snowflake":
            # Snowflake는 문서가 없으므로 검색 결과 없음
            sources = []
            concepts = []
        else:
            # 다른 카테고리는 관련 문서 반환
            sources = [
                f"Document 1 for {category}",
                f"Document 2 for {category}"
            ]
            concepts = [
                'concept1', 'concept2', 'knowledge', 'relation'
            ]

        return {
            'sources': sources,
            'concepts': concepts,
            'confidence': len(sources) / 2 if sources else 0.0
        }

    def evaluate_qa_with_checkpoints(self, qa_item: Dict) -> Dict:
        """체크포인트를 적용하여 Q&A 평가"""
        problem_id = qa_item['problem_id']
        question = qa_item['question']
        expected_answer = qa_item['expected_answer']
        actual_answer = qa_item['team4_answer']
        category = qa_item['category']

        # 1. 검색 시뮬레이션
        retrieval = self.simulate_team4_retrieval(question, category)

        # 2. 범위 제약 검증
        constraint_result = self.constraint_enforcer.validate_answer_scope(
            question, category, retrieval['sources'], retrieval['concepts']
        )

        # 3. Checkpoint 1: 답변 생성 직후
        cp1_result = self.checkpoints.checkpoint_1_answer_generation(
            question=question,
            category=category,
            answer=actual_answer,
            sources=retrieval['sources'],
            constraint_applied=not constraint_result['should_answer']
        )

        # 4. 제약 적용 (필요시 기본 답변으로 변경)
        if not constraint_result['should_answer']:
            constrained_answer = (
                constraint_result['fallback_answer'] or
                "해당 질문은 평가 범위를 벗어났습니다."
            )
        else:
            constrained_answer = actual_answer

        # 5. Checkpoint 2: 채점 전
        cp2_result = self.checkpoints.checkpoint_2_accuracy_scoring(
            expected_answer=expected_answer,
            actual_answer=constrained_answer,
            category=category,
            sources=retrieval['sources']
        )

        # 6. Checkpoint 3: 결과 검증
        cp3_result = self.checkpoints.checkpoint_3_qa_validity(
            problem_id=problem_id,
            expected_answer=expected_answer,
            actual_answer=constrained_answer,
            accuracy_score=cp2_result.get('estimated_accuracy', 0),
            category=category
        )

        # 7. 제약 영향도
        constraint_impact = ConstraintEvaluationMetrics.calculate_constraint_impact(
            actual_answer, constrained_answer,
            not constraint_result['should_answer']
        )

        return {
            'problem_id': problem_id,
            'category': category,
            'question': question[:100],
            'original_accuracy': qa_item['team4_original_accuracy'],
            'constrained': not constraint_result['should_answer'],
            'constraint_reason': constraint_result['reason'],
            'constraint_confidence': constraint_result['confidence'],
            'constraints': constraint_result['constraints'],
            'checkpoints': {
                'cp1': cp1_result,
                'cp2': cp2_result,
                'cp3': cp3_result
            },
            'constraint_impact': constraint_impact,
            'final_accuracy': cp2_result.get('estimated_accuracy',
                                            qa_item['team4_original_accuracy']),
            'needs_review': cp3_result.get('needs_review', False),
            'timestamp': datetime.now().isoformat()
        }

    def generate_report(self, results: list) -> Dict:
        """평가 결과 요약"""
        constrained_items = [r for r in results if r['constrained']]
        needs_review = [r for r in results if r['needs_review']]

        accuracy_changes = [
            r['final_accuracy'] - r['original_accuracy']
            for r in results
            if r['original_accuracy']
        ]

        summary = {
            'total_evaluated': len(results),
            'constrained_count': len(constrained_items),
            'constrain_rate': f"{len(constrained_items) / len(results) * 100:.1f}%",
            'needs_review_count': len(needs_review),
            'average_accuracy_change': (
                sum(accuracy_changes) / len(accuracy_changes)
                if accuracy_changes else 0
            ),
            'checkpoint_summary': self.checkpoints.get_checkpoint_summary()
        }

        return {
            'results': results,
            'summary': summary,
            'timestamp': datetime.now().isoformat()
        }

    def save_checkpoint_log(self, report: Dict):
        """체크포인트 로그 저장"""
        output_file = self.output_dir / "checkpoint_log.json"
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(report, f, ensure_ascii=False, indent=2)

        print(f"[OK] 체크포인트 로그 저장: {output_file}")

    def save_checkpoint_report(self, report: Dict):
        """마크다운 보고서 저장"""
        summary = report['summary']
        results = report['results']

        md_content = f"""# Stage 2 평가 중 검증 보고서

**생성일**: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

## 요약

| 항목 | 결과 |
|---|---:|
| 총 평가 문항 | {summary['total_evaluated']} |
| 제약 적용 항목 | {summary['constrained_count']} ({summary['constrain_rate']}) |
| 재검토 필요 | {summary['needs_review_count']} |
| 평균 정확도 변화 | {summary['average_accuracy_change']:+.1f}%p |

## 체크포인트 결과

### Checkpoint 1 (답변 생성 직후)
- 검증 항목: 답변 범위, 근거, 길이
- 상태: 실행됨

### Checkpoint 2 (채점 전)
- 검증 항목: 기대값/실제값 비교, 채점 기준
- 상태: 실행됨

### Checkpoint 3 (결과 검증)
- 검증 항목: Q&A 유효성, 점수 타당성
- 재검토 필요: {summary['needs_review_count']}개

## 제약 적용 현황

### 제약이 적용된 항목 ({summary['constrained_count']}개)

"""

        constrained = [r for r in results if r['constrained']]
        for item in constrained:
            md_content += f"\n#### {item['problem_id']}\n"
            md_content += f"- 사유: {item['constraint_reason'][:100]}\n"
            md_content += f"- 신뢰도: {item['constraint_confidence']:.2f}\n"
            md_content += f"- 원래 정확도: {item['original_accuracy']}%\n"
            md_content += f"- 수정 후 정확도: {item['final_accuracy']}%\n"

        # 재검토 필요 항목
        review_items = [r for r in results if r['needs_review']]
        if review_items:
            md_content += f"\n## 재검토 필요 항목 ({len(review_items)}개)\n\n"
            for item in review_items:
                md_content += f"- {item['problem_id']}: {item['category']}\n"

        output_file = self.report_dir / "evaluation_checkpoint_report.md"
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write(md_content)

        print(f"[OK] 체크포인트 보고서 저장: {output_file}")

    def run(self):
        """Stage 2 실행"""
        print("\n" + "="*70)
        print("PHASE 6 Stage 2: 평가 중 검증 (Evaluation Checkpoints)")
        print("="*70)

        try:
            # 1. 평가 데이터 로드
            print("\n[1/4] 평가 데이터 로드...")
            qa_data = self.load_evaluation_data()

            # 2. 체크포인트 적용
            print("\n[2/4] 체크포인트 적용 중...")
            results = []
            for idx, qa_item in enumerate(qa_data, 1):
                print(f"\r[{idx}/{len(qa_data)}] {qa_item['problem_id']}", end="")
                result = self.evaluate_qa_with_checkpoints(qa_item)
                results.append(result)

            print()  # 줄바꿈

            # 3. 보고서 생성
            print("\n[3/4] 보고서 생성...")
            report = self.generate_report(results)

            # 4. 저장
            print("\n[4/4] 결과 저장...")
            self.save_checkpoint_log(report)
            self.save_checkpoint_report(report)

            # 최종 요약
            summary = report['summary']
            print("\n" + "="*70)
            print("[DONE] Stage 2 평가 중 검증 완료")
            print("="*70)
            print(f"\n총 {summary['total_evaluated']}개 문항 평가")
            print(f"  제약 적용: {summary['constrained_count']} ({summary['constrain_rate']})")
            print(f"  재검토 필요: {summary['needs_review_count']}개")
            print(f"  정확도 변화: {summary['average_accuracy_change']:+.1f}%p")

            print(f"\n보고서: {self.report_dir}")
            print("="*70 + "\n")

        except Exception as e:
            print(f"\n오류 발생: {e}")
            raise


if __name__ == "__main__":
    runner = Stage2Runner()
    runner.run()
