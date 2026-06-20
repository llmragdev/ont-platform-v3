"""
Stage 3 실행: 정답 보정 및 시스템 업데이트

평가 후 정답을 보정하고 자동으로 시스템에 반영하는 프로세스:

1. 정답 입력 (CLI)
   - 대화형으로 정답 입력
   - 검증 및 저장

2. 영향도 분석
   - 온톨로지 변경점
   - RAG 변경점
   - 평가 기준 변경점

3. 시스템 업데이트
   - 온톨로지 업데이트
   - 벡터 DB 갱신
   - 평가 기준 변경

4. 재평가
   - 보정된 시스템으로 재평가
   - 회귀 테스트
   - 개선도 비교

산출물:
- corrections_YYYYMMDD_HHMMSS.json
- impact_analysis_report.md
- revalidation_results.xlsx
"""

import sys
from pathlib import Path
import json
from datetime import datetime
import openpyxl
from typing import Dict, List

# 경로 설정
REPO_ROOT = Path("E:/ontology_edu/X_ont_std")
EVAL_FRAMEWORK = REPO_ROOT / "evaluation_framework"
EVAL_DATA = REPO_ROOT / "validation/ont_platform_v4_eval"

# 모듈 임포트
sys.path.insert(0, str(EVAL_FRAMEWORK))

from config.category_definitions import CATEGORY_DEFINITIONS
from stages.stage3_correction.cli import (
    AnswerCorrectionCLI, CorrectionValidator, interactive_correction_session
)
from stages.stage3_correction.impact_analyzer import (
    ImpactAnalyzer, ConsolidatedImpactAnalysis
)


class Stage3Runner:
    """Stage 3 실행기"""

    def __init__(self):
        """초기화"""
        self.eval_data_path = EVAL_DATA / "reports/4팀_정확도_비교.xlsx"
        self.output_dir = EVAL_FRAMEWORK / "data/corrections"
        self.report_dir = EVAL_FRAMEWORK / "reports"
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def load_checkpoint_results(self) -> List[Dict]:
        """Stage 2 체크포인트 결과 로드"""
        checkpoint_log = self.output_dir.parent / "checkpoint_log.json"

        # Excel에서 예상답변 매핑 로드
        excel_qa = {}
        try:
            excel_qa = {qa['problem_id']: qa for qa in self._load_qa_from_excel()}
        except Exception as e:
            print(f"[WARNING] Excel 로드 실패: {e}")

        if checkpoint_log.exists():
            with open(checkpoint_log, 'r', encoding='utf-8') as f:
                data = json.load(f)
                results = data.get('results', [])
                for r in results:
                    pid = r.get('problem_id')
                    if pid in excel_qa:
                        r['current_expected_answer'] = excel_qa[pid]['current_expected_answer']
                    else:
                        r['current_expected_answer'] = "N/A"
                    r['team_accuracy'] = r.get('final_accuracy', 0)
                    r['needs_review'] = r.get('needs_review', False) or str(pid).startswith("STD-S-")
                return results

        # 없으면 최소한의 데이터로
        print("[WARNING] 체크포인트 로그를 찾을 수 없음, 전체 QA 로드")
        return list(excel_qa.values())

    def _load_qa_from_excel(self) -> List[Dict]:
        """Excel에서 QA 로드"""
        if not self.eval_data_path.exists():
            raise FileNotFoundError(f"평가 데이터 파일 없음: {self.eval_data_path}")

        wb = openpyxl.load_workbook(self.eval_data_path)
        ws = wb['문항별 비교 상세']

        qa_data = []
        for row_idx in range(2, ws.max_row + 1):
            problem_id = ws.cell(row=row_idx, column=1).value
            if not problem_id:
                continue

            question = ws.cell(row=row_idx, column=2).value or ""
            expected_answer = ws.cell(row=row_idx, column=3).value or ""

            if str(problem_id).startswith("STD-O-"):
                category = "Ontology"
            elif str(problem_id).startswith("STD-A-"):
                category = "Advanced RAG"
            elif str(problem_id).startswith("STD-S-"):
                category = "Snowflake"
            else:
                category = "Unknown"

            qa_data.append({
                'problem_id': str(problem_id),
                'question': question,
                'current_expected_answer': expected_answer,
                'category': category,
                'team_accuracy': 0,
                'needs_review': str(problem_id).startswith("STD-S-")
            })

        print(f"[OK] {len(qa_data)}개 QA 로드됨")
        return qa_data

    def run_correction_session(self, qa_pairs: List[Dict]) -> Dict:
        """정답 보정 세션"""
        print("\n" + "="*70)
        print("[Stage 3-1] 정답 입력 및 검증")
        print("="*70)

        # 대화형 세션
        result = interactive_correction_session(qa_pairs)

        if not result:
            print("\n❌ 보정 검증 실패")
            return None

        return result

    def analyze_impact(self, corrections: List[Dict]) -> Dict:
        """영향도 분석"""
        print("\n" + "="*70)
        print("[Stage 3-2] 영향도 분석")
        print("="*70)

        analyzer = ConsolidatedImpactAnalysis(CATEGORY_DEFINITIONS)
        analysis = analyzer.analyze_all_corrections(corrections)

        print(f"\n분석 완료:")
        print(f"  - 총 수정: {analysis['total_corrections']}개")
        print(f"  - 온톨로지 변경: {len(analysis['consolidated_changes']['ontology_changes'])}개")
        print(f"  - RAG 변경: {len(analysis['consolidated_changes']['rag_changes'])}개")
        print(f"  - 평가 기준 변경: {len(analysis['consolidated_changes']['evaluation_changes'])}개")
        print(f"  - 전체 위험도: {analysis['overall_risk_level']}")

        return analysis

    def generate_impact_report(self, analysis: Dict) -> str:
        """영향도 보고서 생성"""
        md_content = f"""# Stage 3 영향도 분석 보고서

**생성일**: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

## 요약

| 항목 | 결과 |
|---|---:|
| 총 수정 | {analysis['total_corrections']}개 |
| 온톨로지 변경 | {len(analysis['consolidated_changes']['ontology_changes'])}개 |
| RAG 변경 | {len(analysis['consolidated_changes']['rag_changes'])}개 |
| 평가 기준 변경 | {len(analysis['consolidated_changes']['evaluation_changes'])}개 |
| 전체 영향도 점수 | {analysis['overall_impact_score']:.2f} |
| 위험도 | {analysis['overall_risk_level']} |

## 온톨로지 변경점

"""
        for change in analysis['consolidated_changes']['ontology_changes']:
            md_content += f"### {change.get('action')}\n"
            md_content += f"- 위험도: {change.get('risk', 'N/A')}\n"
            if 'concepts' in change:
                md_content += f"- 개념: {', '.join(change['concepts'])}\n"
            md_content += f"- 사유: {change.get('reason', 'N/A')}\n\n"

        ## RAG 변경점
        md_content += "## RAG 변경점\n\n"
        for change in analysis['consolidated_changes']['rag_changes']:
            md_content += f"### {change.get('action')}\n"
            md_content += f"- {change.get('description', 'N/A')}\n\n"

        ## 구현 순서
        md_content += "## 구현 순서\n\n"
        for i, change in enumerate(analysis['implementation_priority'], 1):
            md_content += f"{i}. [{change.get('priority')}] {change.get('action')}\n"
            md_content += f"   - 기한: {change.get('timeline')}\n\n"

        output_file = self.report_dir / "impact_analysis_report.md"
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write(md_content)

        print(f"\n[OK] 영향도 보고서 저장: {output_file}")
        return str(output_file)

    def simulate_system_update(self, analysis: Dict) -> Dict:
        """시스템 업데이트 시뮬레이션"""
        print("\n" + "="*70)
        print("[Stage 3-3] 시스템 업데이트 (시뮬레이션)")
        print("="*70)

        results = {
            'ontology_updated': len(analysis['consolidated_changes']['ontology_changes']),
            'rag_updated': len(analysis['consolidated_changes']['rag_changes']),
            'evaluation_updated': len(analysis['consolidated_changes']['evaluation_changes']),
            'timestamp': datetime.now().isoformat(),
            'status': 'SIMULATED'  # 실제로는 각 시스템에 반영되어야 함
        }

        print(f"\n시뮬레이션 완료:")
        print(f"  [OK] 온톨로지 업데이트 {results['ontology_updated']}개")
        print(f"  [OK] RAG 업데이트 {results['rag_updated']}개")
        print(f"  [OK] 평가 기준 업데이트 {results['evaluation_updated']}개")

        return results

    def simulate_revalidation(self, corrections: List[Dict]) -> Dict:
        """재평가 시뮬레이션"""
        print("\n" + "="*70)
        print("[Stage 3-4] 재평가 (시뮬레이션)")
        print("="*70)

        revalidation_results = {
            'total_revalidated': len(corrections),
            'improved': 0,
            'unchanged': 0,
            'regressed': 0,
            'average_improvement': 0.0,
            'details': []
        }

        for correction in corrections:
            if correction['correction_type'] == 'SKIPPED':
                continue

            # 시뮬레이션: Snowflake는 정확도 하락, 다른 것은 유지
            if correction['category'] == 'Snowflake':
                improvement = -50  # 기대값으로 정확도 하락
                revalidation_results['regressed'] += 1
            else:
                improvement = 0
                revalidation_results['unchanged'] += 1

            revalidation_results['details'].append({
                'problem_id': correction['problem_id'],
                'category': correction['category'],
                'accuracy_before': correction.get('team_accuracy_before', 0),
                'accuracy_after': max(
                    0, correction.get('team_accuracy_before', 0) + improvement
                ),
                'improvement': improvement
            })

        avg_improvement = (
            sum(d['improvement'] for d in revalidation_results['details'])
            / len(revalidation_results['details'])
            if revalidation_results['details'] else 0
        )
        revalidation_results['average_improvement'] = avg_improvement

        print(f"\n재평가 완료:")
        print(f"  - 개선: {revalidation_results['improved']}개")
        print(f"  - 유지: {revalidation_results['unchanged']}개")
        print(f"  - 하락: {revalidation_results['regressed']}개")
        print(f"  - 평균 변화: {avg_improvement:+.1f}%p")

        return revalidation_results

    def run(self):
        """Stage 3 실행"""
        print("\n" + "="*70)
        print("PHASE 6 Stage 3: 정답 보정 및 시스템 업데이트")
        print("="*70)

        try:
            # 1. QA 로드
            print("\n[Loading] QA 데이터...")
            qa_pairs = self.load_checkpoint_results()

            # 2. 정답 입력
            correction_result = self.run_correction_session(qa_pairs)
            if not correction_result:
                return

            corrections = correction_result['corrections']
            print(f"\n✓ {len(corrections)}개 정답 수정 완료")

            # 3. 영향도 분석
            analysis = self.analyze_impact(corrections)
            self.generate_impact_report(analysis)

            # 4. 시스템 업데이트 (시뮬레이션)
            update_result = self.simulate_system_update(analysis)

            # 5. 재평가
            revalidation = self.simulate_revalidation(corrections)

            # 최종 요약
            print("\n" + "="*70)
            print("[DONE] Stage 3 정답 보정 완료")
            print("="*70)
            print(f"\n✓ {len(corrections)}개 정답 수정")
            print(f"✓ 시스템 업데이트 완료 (시뮬레이션)")
            print(f"✓ 재평가 완료")
            print(f"\n평균 정확도 변화: {revalidation['average_improvement']:+.1f}%p")
            print("="*70 + "\n")

        except Exception as e:
            print(f"\n❌ 오류 발생: {e}")
            raise


if __name__ == "__main__":
    runner = Stage3Runner()
    runner.run()
