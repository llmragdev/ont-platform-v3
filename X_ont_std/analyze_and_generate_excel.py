# -*- coding: utf-8 -*-
"""
새 기준에 따른 정확도 평가 엑셀 생성
중간점검_통합테스트_예상 질문_답변-v1.2.xlsx 기준으로
4팀_정확도_비교_STD-S카테고리무관_추가평가.xlsx 보완
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
from collections import defaultdict

# ======================== Step 1: 기존 파일 분석 ========================

print("=" * 60)
print("Step 1: 파일 구조 분석")
print("=" * 60)

# 새 기준 파일 읽기
new_file = r"E:\ai_lab_SIT\qa_expected\중간점검_통합테스트_예상 질문_답변-v1.2.xlsx"
old_file = r"E:\ontology_edu\X_ont_std\validation\ont_platform_v4_eval\reports\4팀_정확도_비교_STD-S카테고리무관_추가평가.xlsx"

try:
    wb_new = openpyxl.load_workbook(new_file)
    ws_new = wb_new.active

    print(f"\n[새 파일] {os.path.basename(new_file)}")
    print(f"  시트명: {ws_new.title}")
    print(f"  행 수: {ws_new.max_row}")
    print(f"  열 수: {ws_new.max_column}")

    # 헤더 확인
    headers_new = []
    for col in range(1, ws_new.max_column + 1):
        cell = ws_new.cell(row=1, column=col)
        headers_new.append(cell.value)

    print(f"  헤더: {headers_new[:5]}...")

    # 샘플 데이터
    print(f"\n  샘플 데이터 (처음 3행):")
    for row in range(2, min(5, ws_new.max_row + 1)):
        row_data = []
        for col in range(1, 5):
            cell = ws_new.cell(row=row, column=col)
            row_data.append(cell.value)
        print(f"    Row {row}: {row_data}")

    # 질문 목록 추출
    questions = []
    categories = defaultdict(int)

    for row in range(2, ws_new.max_row + 1):
        q_id = ws_new.cell(row=row, column=1).value
        category = ws_new.cell(row=row, column=2).value
        question = ws_new.cell(row=row, column=3).value

        if q_id:
            questions.append({
                'id': q_id,
                'category': category,
                'question': question
            })
            if category:
                categories[category] += 1

    print(f"\n  추출된 질문 수: {len(questions)}")
    print(f"  카테고리별 분포:")
    for cat, count in sorted(categories.items()):
        print(f"    {cat}: {count}개")

except Exception as e:
    print(f"오류: {e}")
    import traceback
    traceback.print_exc()

print("\n" + "=" * 60)
print("Step 2: 기존 평가 파일 분석")
print("=" * 60)

try:
    wb_old = openpyxl.load_workbook(old_file)
    ws_old = wb_old.active

    print(f"\n[기존 파일] {os.path.basename(old_file)}")
    print(f"  시트명: {ws_old.title}")
    print(f"  행 수: {ws_old.max_row}")
    print(f"  열 수: {ws_old.max_column}")

    # 헤더 확인
    headers_old = []
    for col in range(1, ws_old.max_column + 1):
        cell = ws_old.cell(row=1, column=col)
        headers_old.append(cell.value)

    print(f"  헤더: {headers_old}")

    # 평가 데이터
    print(f"\n  샘플 평가 데이터 (처음 3행):")
    for row in range(2, min(5, ws_old.max_row + 1)):
        row_data = []
        for col in range(1, min(8, ws_old.max_column + 1)):
            cell = ws_old.cell(row=row, column=col)
            row_data.append(cell.value)
        print(f"    Row {row}: {row_data}")

except Exception as e:
    print(f"오류: {e}")
    import traceback
    traceback.print_exc()

print("\n" + "=" * 60)
print("Step 3: 새로운 엑셀 생성")
print("=" * 60)

try:
    # 새 워크북 생성
    wb_new_report = openpyxl.Workbook()
    ws = wb_new_report.active
    ws.title = "정확도 평가"

    # 스타일 정의
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    category_fill = PatternFill(start_color="D9E8F5", end_color="D9E8F5", fill_type="solid")
    category_font = Font(bold=True, size=10)

    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # 헤더 작성
    headers = [
        '질문ID',
        '카테고리',
        '질문',
        'v4 응답',
        'v4 정답 여부',
        'v5 응답',
        'v5 정답 여부',
        '신뢰도',
        '개선여부',
        '비고'
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border

    # 컬럼 너비 설정
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 25
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 10
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 15

    # 데이터 입력 (샘플)
    row = 2
    current_category = None

    for q in questions:
        if q['category'] != current_category:
            current_category = q['category']

        # 질문ID
        cell = ws.cell(row=row, column=1)
        cell.value = q['id']
        cell.alignment = center_align
        cell.border = border

        # 카테고리
        cell = ws.cell(row=row, column=2)
        cell.value = q['category']
        cell.alignment = center_align
        cell.border = border

        # 질문
        cell = ws.cell(row=row, column=3)
        cell.value = q['question']
        cell.alignment = left_align
        cell.border = border

        # v4 응답 (입력 필요)
        for col in range(4, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.alignment = left_align if col in [4, 6] else center_align
            cell.border = border

        row += 1

    # 요약 시트 추가
    ws_summary = wb_new_report.create_sheet("요약")

    summary_headers = ['항목', '값']
    for col, header in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border

    ws_summary.column_dimensions['A'].width = 20
    ws_summary.column_dimensions['B'].width = 15

    # 요약 데이터
    summary_items = [
        ['전체 질문 수', len(questions)],
        ['v4 정답 수', ''],
        ['v4 정확도', ''],
        ['v5 정답 수', ''],
        ['v5 정확도', ''],
        ['개선된 질문 수', ''],
        ['악화된 질문 수', ''],
        ['카테고리별 정확도', '별도 시트 참고'],
    ]

    for i, item in enumerate(summary_items, 2):
        ws_summary.cell(row=i, column=1).value = item[0]
        ws_summary.cell(row=i, column=2).value = item[1]
        ws_summary.cell(row=i, column=2).alignment = center_align
        ws_summary.cell(row=i, column=1).border = border
        ws_summary.cell(row=i, column=2).border = border

    # 카테고리별 시트 추가
    ws_category = wb_new_report.create_sheet("카테고리별")

    category_headers = ['카테고리', '전체', 'v4정답', 'v4정확도', 'v5정답', 'v5정확도']
    for col, header in enumerate(category_headers, 1):
        cell = ws_category.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border

    ws_category.column_dimensions['A'].width = 15
    ws_category.column_dimensions['B'].width = 10
    ws_category.column_dimensions['C'].width = 12
    ws_category.column_dimensions['D'].width = 12
    ws_category.column_dimensions['E'].width = 12
    ws_category.column_dimensions['F'].width = 12

    row = 2
    for cat, count in sorted(categories.items()):
        ws_category.cell(row=row, column=1).value = cat
        ws_category.cell(row=row, column=2).value = count
        for col in range(1, 7):
            ws_category.cell(row=row, column=col).border = border
            ws_category.cell(row=row, column=col).alignment = center_align
        row += 1

    # 파일 저장
    output_path = r"E:\ontology_edu\X_ont_std\validation\ont_platform_v4_eval\reports\4팀_정확도_비교_통합검증_v2.xlsx"
    wb_new_report.save(output_path)

    print(f"\n✓ 새 엑셀 생성 완료")
    print(f"  경로: {output_path}")
    print(f"  시트 수: {len(wb_new_report.sheetnames)}")
    print(f"  시트명: {wb_new_report.sheetnames}")
    print(f"  질문 데이터 행: {len(questions)}")

except Exception as e:
    print(f"오류: {e}")
    import traceback
    traceback.print_exc()

print("\n" + "=" * 60)
print("완료!")
print("=" * 60)
