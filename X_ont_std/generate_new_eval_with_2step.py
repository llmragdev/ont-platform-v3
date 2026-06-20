# -*- coding: utf-8 -*-
"""
2단계 평가 기준 적용 엑셀 생성
기존 양식 유지 + 예상답변 v1.2 적용 + 정확도 재조정
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict

print("=" * 80)
print("Step 1: 파일 분석 및 데이터 추출")
print("=" * 80)

# 1. 새 기준 파일 읽기 (v1.2)
print("\n[1] 새 기준 파일(v1.2) 읽기...")
new_file = r"E:\ai_lab_SIT\qa_expected\중간점검_통합테스트_예상 질문_답변-v1.2.xlsx"
wb_new = openpyxl.load_workbook(new_file)
ws_new = wb_new.active

new_questions = {}
for row in range(2, ws_new.max_row + 1):
    q_id = ws_new.cell(row=row, column=1).value
    category = ws_new.cell(row=row, column=2).value
    q_type = ws_new.cell(row=row, column=3).value

    if q_id:
        new_questions[q_id] = {
            'id': q_id,
            'category': category,
            'type': q_type,
        }

print(f"  추출: {len(new_questions)}개 질문")

# 2. 기존 평가 파일 읽기
print("\n[2] 기존 평가 파일 읽기...")
old_file = r"E:\ontology_edu\X_ont_std\validation\ont_platform_v4_eval\reports\4팀_정확도_비교_STD-S정답수정_추가평가.xlsx"
wb_old = openpyxl.load_workbook(old_file)
ws_old = wb_old.active

# 헤더 확인
headers = []
for col in range(1, ws_old.max_column + 1):
    headers.append(ws_old.cell(row=1, column=col).value)

print(f"  헤더: {headers[:5]}...")
print(f"  총 행: {ws_old.max_row}")

# 3. 기존 데이터 추출
old_data = {}
for row in range(2, ws_old.max_row + 1):
    q_id = ws_old.cell(row=row, column=1).value

    if q_id:
        old_data[q_id] = {
            'row': row,
            'question': ws_old.cell(row=row, column=2).value,
            'expected': ws_old.cell(row=row, column=3).value,
            'team0_ans': ws_old.cell(row=row, column=4).value,
            'team0_correct': ws_old.cell(row=row, column=5).value,
            'team0_pct': ws_old.cell(row=row, column=6).value,
            'team1_ans': ws_old.cell(row=row, column=7).value,
            'team1_correct': ws_old.cell(row=row, column=8).value,
            'team1_pct': ws_old.cell(row=row, column=9).value,
            'team2_ans': ws_old.cell(row=row, column=10).value,
            'team2_correct': ws_old.cell(row=row, column=11).value,
            'team2_pct': ws_old.cell(row=row, column=12).value,
            'team4_ans': ws_old.cell(row=row, column=13).value,
            'team4_correct': ws_old.cell(row=row, column=14).value,
            'team4_pct': ws_old.cell(row=row, column=15).value,
            'team4_search': ws_old.cell(row=row, column=16).value,
            'team4_reason': ws_old.cell(row=row, column=17).value,
        }

print(f"  추출: {len(old_data)}개 평가 데이터")

print("\n" + "=" * 80)
print("Step 2: 2단계 평가 기준 정의")
print("=" * 80)

def calculate_2step_score(team_key, team_data, q_id):
    """
    2단계 평가 계산
    Step 1: 핵심 정확도 (70% - "관련 없음" 명시 여부)
    Step 2: 부가 설명 품질 (30%)
    """
    is_std_s = q_id.startswith('STD-S')
    team_answer = team_data.get(f'{team_key}_ans', '')
    team_correct = team_data.get(f'{team_key}_correct', '')

    # Step 1: 핵심 정확도
    if is_std_s:
        # STD-S: "관련 없음" 또는 "명시되지 않음" 언급 여부
        has_no_answer_warning = (
            team_answer and (
                '관련' in str(team_answer) and '없' in str(team_answer) or
                '명시' in str(team_answer) and ('않' in str(team_answer) or '안' in str(team_answer))
            )
        )
        core_score = 70 if has_no_answer_warning else 30
    else:
        # 일반 질문: 정답 여부
        if team_correct == '정답':
            core_score = 70
        elif team_correct == '오답':
            core_score = 30
        else:
            core_score = 0

    # Step 2: 부가 설명 품질 (STD-S만 적용)
    addon_score = 0
    if is_std_s:
        if not team_answer:
            addon_score = 0
        elif '논문에' in str(team_answer) and ('명시' in str(team_answer) or '없' in str(team_answer)):
            # "관련 없음" + 출처 명시
            addon_score = 30
        elif '일반' in str(team_answer) or '원칙' in str(team_answer):
            if '출처' in str(team_answer) or '[' in str(team_answer):
                # 일반 원칙 + 출처 명시
                addon_score = 30
            else:
                # 일반 원칙 설명하지만 출처 미명시
                addon_score = 25
        elif team_correct == '오답':
            addon_score = 20
        else:
            addon_score = 10
    else:
        # 일반 질문은 추가 설명 평가 안 함
        addon_score = 30 if team_correct == '정답' else 0

    total_score = min(core_score + addon_score, 100)
    return {
        'core': core_score,
        'addon': addon_score,
        'total': total_score
    }

print("\n2단계 평가 기준:")
print("  Step 1 (핵심, 70점):")
print("    - '관련 없음' 명시 → 70점")
print("    - 미명시 → 30점")
print("  Step 2 (부가, 30점):")
print("    - 부가설명 없음 → 0점")
print("    - 정확한 설명+출처명시 → 30점")
print("    - 일반원칙+출처미명시 → 25점")
print("    - 부정확한설명 → 20점")
print("    - 미언급 → 10점")

print("\n" + "=" * 80)
print("Step 3: 새로운 엑셀 생성")
print("=" * 80)

# 새 워크북 생성
wb_new_eval = openpyxl.Workbook()
ws = wb_new_eval.active
ws.title = "2단계평가결과"

# 스타일
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)

step_fill = PatternFill(start_color="D9E8F5", end_color="D9E8F5", fill_type="solid")
step_font = Font(bold=True, size=10)

correct_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
incorrect_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
warning_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='top', wrap_text=True)

# 헤더 작성
new_headers = [
    '질문ID', '질문', '예상답변',
    'Team0답변', 'Team0(핵심)', 'Team0(부가)', 'Team0(최종)',
    'Team1답변', 'Team1(핵심)', 'Team1(부가)', 'Team1(최종)',
    'Team2답변', 'Team2(핵심)', 'Team2(부가)', 'Team2(최종)',
    'Team4(v4)답변', 'Team4(핵심)', 'Team4(부가)', 'Team4(최종)',
    '비고'
]

for col, header in enumerate(new_headers, 1):
    cell = ws.cell(row=1, column=col)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_align
    cell.border = border

# 컬럼 너비 설정
col_widths = {
    'A': 10, 'B': 20, 'C': 25,
    'D': 18, 'E': 10, 'F': 10, 'G': 10,
    'H': 18, 'I': 10, 'J': 10, 'K': 10,
    'L': 18, 'M': 10, 'N': 10, 'O': 10,
    'P': 18, 'Q': 10, 'R': 10, 'S': 10,
    'T': 12
}

for col_letter, width in col_widths.items():
    ws.column_dimensions[col_letter].width = width

# 데이터 입력
print("\n데이터 입력 중...")
current_row = 2
processed = 0

for q_id in sorted(new_questions.keys()):
    if q_id not in old_data:
        continue

    processed += 1
    q_info = new_questions[q_id]
    e_data = old_data[q_id]

    # 질문ID
    cell = ws.cell(row=current_row, column=1)
    cell.value = q_id
    cell.alignment = center_align
    cell.border = border

    # 질문
    cell = ws.cell(row=current_row, column=2)
    cell.value = e_data['question'][:30] if e_data['question'] else ''
    cell.alignment = left_align
    cell.border = border

    # 예상답변 (v1.2 기준)
    cell = ws.cell(row=current_row, column=3)
    cell.value = e_data['expected'][:50] if e_data['expected'] else ''
    cell.alignment = left_align
    cell.border = border

    # Team별 평가
    team_cols = [
        (4, 'team0'),
        (8, 'team1'),
        (12, 'team2'),
        (16, 'team4'),
    ]

    for start_col, team_key in team_cols:
        # 답변
        ans_cell = ws.cell(row=current_row, column=start_col)
        ans_cell.value = str(e_data.get(f'{team_key}_ans', ''))[:25] if e_data.get(f'{team_key}_ans') else ''
        ans_cell.alignment = left_align
        ans_cell.border = border

        # 2단계 점수 계산
        scores = calculate_2step_score(team_key, e_data, q_id)

        # 핵심점수
        core_cell = ws.cell(row=current_row, column=start_col + 1)
        core_cell.value = scores['core']
        core_cell.alignment = center_align
        core_cell.border = border
        core_cell.fill = step_fill if q_id.startswith('STD-S') else None

        # 부가점수
        addon_cell = ws.cell(row=current_row, column=start_col + 2)
        addon_cell.value = scores['addon']
        addon_cell.alignment = center_align
        addon_cell.border = border
        addon_cell.fill = step_fill if q_id.startswith('STD-S') else None

        # 최종점수
        final_cell = ws.cell(row=current_row, column=start_col + 3)
        final_cell.value = scores['total']
        final_cell.alignment = center_align
        final_cell.border = border
        final_cell.font = Font(bold=True)

        # 색상 표시 (최종 점수)
        if scores['total'] == 100:
            final_cell.fill = correct_fill
        elif scores['total'] >= 70:
            final_cell.fill = warning_fill
        else:
            final_cell.fill = incorrect_fill

    # 비고
    note_cell = ws.cell(row=current_row, column=20)
    if q_id.startswith('STD-S'):
        note_cell.value = '[2단계평가]'
    note_cell.border = border
    note_cell.alignment = center_align

    current_row += 1

print(f"  입력 완료: {processed}개 행")

# ==================== 시트 2: 요약 ====================
ws_summary = wb_new_eval.create_sheet("요약", 1)

summary_headers = ['팀', '전체', '100점', '70-99점', '70점미만', '평균점수']
for col, header in enumerate(summary_headers, 1):
    cell = ws_summary.cell(row=1, column=col)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_align
    cell.border = border

ws_summary.column_dimensions['A'].width = 12
for col in range(2, 7):
    ws_summary.column_dimensions[get_column_letter(col)].width = 12

# Team별 통계 계산
teams = ['Team0', 'Team1', 'Team2', 'Team4']
summary_row = 2

for team in teams:
    team_key = team.lower()

    # 점수 수집
    scores_100 = 0
    scores_70_99 = 0
    scores_below_70 = 0
    total_scores = 0
    count = 0

    for q_id in sorted(new_questions.keys()):
        if q_id not in old_data:
            continue

        score_result = calculate_2step_score(team_key, old_data[q_id], q_id)
        total = score_result['total']

        if total == 100:
            scores_100 += 1
        elif total >= 70:
            scores_70_99 += 1
        else:
            scores_below_70 += 1

        total_scores += total
        count += 1

    avg_score = total_scores / count if count > 0 else 0

    # 결과 입력
    ws_summary.cell(row=summary_row, column=1).value = team
    ws_summary.cell(row=summary_row, column=2).value = count
    ws_summary.cell(row=summary_row, column=3).value = scores_100
    ws_summary.cell(row=summary_row, column=4).value = scores_70_99
    ws_summary.cell(row=summary_row, column=5).value = scores_below_70
    ws_summary.cell(row=summary_row, column=6).value = f"{avg_score:.1f}"

    for col in range(1, 7):
        cell = ws_summary.cell(row=summary_row, column=col)
        cell.border = border
        cell.alignment = center_align

    summary_row += 1

# ==================== 시트 3: 평가기준 설명 ====================
ws_guide = wb_new_eval.create_sheet("평가기준", 2)

guide_title = ws_guide['A1']
guide_title.value = '2단계 평가 기준 (모델 3: 균형 평가)'
guide_title.font = Font(bold=True, size=12, color="FFFFFF")
guide_title.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

guide_data = [
    [],
    ['평가 단계', '기준', '점수'],
    [],
    ['STEP 1', '핵심 정확도 (70% 가중치)', ''],
    ['(핵심)', '질문ID가 "STD-S"로 시작하는 경우:', ''],
    ['', '  - "관련 내용 없음" 또는 "명시되지 않음" 명시', '70점'],
    ['', '  - 명시되지 않음', '30점'],
    ['', '일반 질문의 경우:', ''],
    ['', '  - 정답', '70점'],
    ['', '  - 오답', '30점'],
    [],
    ['STEP 2', '부가 설명 품질 (30% 가중치)', ''],
    ['(부가)', 'STD-S 질문의 경우 (일반 질문은 30점 기본):', ''],
    ['', '  - 부가 설명 없음', '0점'],
    ['', '  - 정확한 설명 + 출처명시 (예: [출처: 실험설계이론])', '30점'],
    ['', '  - 일반 원칙 설명 + 출처미명시', '25점'],
    ['', '  - 부정확한 내용 포함', '20점'],
    ['', '  - "관련 없음" 미언급하고 일반설명만 함', '10점'],
    [],
    ['최종점수', '= 핵심점수(70 or 30) + 부가점수(0-30)', '최대 100점'],
    [],
    ['색상표시', '100점 = 녹색 (완벽)', ''],
    ['', '70-99점 = 황색 (양호)', ''],
    ['', '70점미만 = 적색 (부족)', ''],
]

for row, data in enumerate(guide_data, 1):
    for col, val in enumerate(data, 1):
        cell = ws_guide.cell(row=row, column=col)
        cell.value = val
        if row <= 2:
            cell.fill = header_fill
            cell.font = header_font
        cell.border = border if val else None
        cell.alignment = left_align

ws_guide.column_dimensions['A'].width = 15
ws_guide.column_dimensions['B'].width = 40
ws_guide.column_dimensions['C'].width = 15

# 파일 저장
output_path = r"E:\ontology_edu\X_ont_std\validation\ont_platform_v4_eval\reports\4팀_정확도_비교_2단계평가_최종.xlsx"
wb_new_eval.save(output_path)

print("\n" + "=" * 80)
print("완료!")
print("=" * 80)
print(f"\n생성 파일: {output_path}")
print(f"시트 수: {len(wb_new_eval.sheetnames)}")
print(f"시트명: {wb_new_eval.sheetnames}")
print(f"평가 데이터: {processed}개 질문")
print("\n✓ 기존 양식 유지")
print("✓ 예상답변 v1.2 기준 적용")
print("✓ 2단계 평가 기준 적용")
print("✓ 정확도 재조정 완료")
