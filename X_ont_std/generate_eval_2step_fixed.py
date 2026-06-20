# -*- coding: utf-8 -*-
"""
2단계 평가 기준 적용 엑셀 생성 (수정 버전)
기존 양식 유지 + 예상답변 v1.2 적용 + 정확도 재조정
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict

print("=" * 80)
print("Step 1: 파일 분석")
print("=" * 80)

# 1. 새 기준 파일 읽기 (v1.2)
print("\n[1] 새 기준 파일(v1.2) 분석...")
new_file = r"E:\ai_lab_SIT\qa_expected\중간점검_통합테스트_예상 질문_답변-v1.2.xlsx"
wb_new = openpyxl.load_workbook(new_file)

# 시트 확인
print(f"  시트: {wb_new.sheetnames}")

ws_new = wb_new.active
print(f"  활성 시트: {ws_new.title}")
print(f"  행 수: {ws_new.max_row}, 열 수: {ws_new.max_column}")

# 헤더 확인
headers_new = []
for col in range(1, ws_new.max_column + 1):
    headers_new.append(ws_new.cell(row=1, column=col).value)
print(f"  헤더: {headers_new}")

# 질문 추출
new_questions = {}
for row in range(2, ws_new.max_row + 1):
    q_id = ws_new.cell(row=row, column=1).value

    if q_id:
        q_data = {}
        for col, header in enumerate(headers_new, 1):
            q_data[header] = ws_new.cell(row=row, column=col).value

        new_questions[q_id] = q_data

print(f"  추출된 질문: {len(new_questions)}개")
if new_questions:
    sample_id = list(new_questions.keys())[0]
    print(f"  샘플 (ID: {sample_id}):")
    for key, val in list(new_questions[sample_id].items())[:5]:
        print(f"    {key}: {str(val)[:40]}")

# 2. 기존 평가 파일 읽기
print("\n[2] 기존 평가 파일 분석...")
old_file = r"E:\ontology_edu\X_ont_std\validation\ont_platform_v4_eval\reports\4팀_정확도_비교_STD-S정답수정_추가평가.xlsx"
wb_old = openpyxl.load_workbook(old_file)
ws_old = wb_old.active

# 헤더
headers_old = []
for col in range(1, 18):  # 최대 17개 컬럼
    cell_val = ws_old.cell(row=1, column=col).value
    if cell_val:
        headers_old.append(cell_val)

print(f"  헤더: {headers_old}")
print(f"  행 수: {ws_old.max_row}")

# 기존 데이터 추출
old_data = {}
for row in range(2, ws_old.max_row + 1):
    q_id = ws_old.cell(row=row, column=1).value

    if q_id:
        row_data = {}
        for col, header in enumerate(headers_old, 1):
            row_data[header] = ws_old.cell(row=row, column=col).value

        old_data[q_id] = {
            'row': row,
            'data': row_data
        }

print(f"  추출된 평가: {len(old_data)}개")

print("\n" + "=" * 80)
print("Step 2: 2단계 평가 점수 계산")
print("=" * 80)

def calculate_2step_score(team_key, row_data, q_id):
    """
    2단계 평가 계산
    Step 1: 핵심 정확도 (70% - "관련 없음" 명시 여부)
    Step 2: 부가 설명 품질 (30%)
    """
    is_std_s = q_id.startswith('STD-S')

    # 팀 데이터 컬럼 찾기
    team_lower = team_key.lower()

    # 컬럼명 패턴
    answer_col = None
    correct_col = None

    for header in row_data.keys():
        if team_key in header and '답변' in header:
            answer_col = header
        if team_key in header and '정답' in header and '%' not in header:
            correct_col = header

    team_answer = str(row_data.get(answer_col, '')) if answer_col else ''
    team_correct = str(row_data.get(correct_col, '')) if correct_col else ''

    # Step 1: 핵심 정확도
    if is_std_s:
        # STD-S: "관련 없음" 또는 "명시되지 않음" 언급 여부
        has_no_answer_warning = (
            team_answer and (
                ('관련' in team_answer and '없' in team_answer) or
                ('명시' in team_answer and ('않' in team_answer or '안' in team_answer))
            )
        )
        core_score = 70 if has_no_answer_warning else 30
    else:
        # 일반 질문: 정답 여부
        if '정답' in team_correct:
            core_score = 70
        elif '오답' in team_correct:
            core_score = 30
        else:
            core_score = 0

    # Step 2: 부가 설명 품질 (STD-S만 적용)
    addon_score = 0
    if is_std_s:
        if not team_answer or team_answer == 'None':
            addon_score = 0
        elif '논문에' in team_answer and ('명시' in team_answer or '없' in team_answer):
            addon_score = 30
        elif '일반' in team_answer or '원칙' in team_answer:
            if '출처' in team_answer or '[' in team_answer:
                addon_score = 30
            else:
                addon_score = 25
        elif '오답' in team_correct:
            addon_score = 20
        else:
            addon_score = 10
    else:
        addon_score = 30 if '정답' in team_correct else 0

    total_score = min(core_score + addon_score, 100)

    return {
        'core': core_score,
        'addon': addon_score,
        'total': total_score,
        'answer': team_answer[:30] if team_answer else '',
        'correct': team_correct
    }

print("점수 계산 로직 준비 완료")

print("\n" + "=" * 80)
print("Step 3: 새로운 엑셀 생성")
print("=" * 80)

# 새 워크북
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "2단계평가결과"

# 스타일
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)

step_fill = PatternFill(start_color="D9E8F5", end_color="D9E8F5", fill_type="solid")
correct_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
incorrect_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
warning_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='top', wrap_text=True)

# 헤더
new_headers = [
    '질문ID', '기본질문', '새기준답변',
    'Team0답변', 'Team0핵심', 'Team0부가', 'Team0최종',
    'Team1답변', 'Team1핵심', 'Team1부가', 'Team1최종',
    'Team2답변', 'Team2핵심', 'Team2부가', 'Team2최종',
    'Team4(v4)답변', 'Team4핵심', 'Team4부가', 'Team4최종',
    '평가결과'
]

for col, header in enumerate(new_headers, 1):
    cell = ws.cell(row=1, column=col)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_align
    cell.border = border

# 컬럼 너비
for col in range(1, 21):
    ws.column_dimensions[get_column_letter(col)].width = 12

ws.column_dimensions['A'].width = 10
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 20

# 데이터 입력
print("\n데이터 입력...")
current_row = 2
processed = 0

for q_id in sorted(new_questions.keys()):
    if q_id not in old_data:
        continue

    processed += 1
    new_q = new_questions[q_id]
    old_row_data = old_data[q_id]['data']

    # 질문ID
    cell = ws.cell(row=current_row, column=1)
    cell.value = q_id
    cell.alignment = center_align
    cell.border = border

    # 기본 질문 (기존)
    cell = ws.cell(row=current_row, column=2)
    q_text = old_row_data.get('질문', '') if old_row_data.get('질문') else ''
    cell.value = str(q_text)[:25] if q_text else ''
    cell.alignment = left_align
    cell.border = border

    # 새 기준 답변 (v1.2)
    cell = ws.cell(row=current_row, column=3)
    expected = new_q.get('예상 답변', '') if new_q.get('예상 답변') else ''
    cell.value = str(expected)[:40] if expected else ''
    cell.alignment = left_align
    cell.border = border

    # Team별 평가
    teams = ['Team0', 'Team1', 'Team2', 'Team4']
    all_scores = {}

    for team_idx, team in enumerate(teams):
        col_start = 4 + (team_idx * 4)

        scores = calculate_2step_score(team, old_row_data, q_id)
        all_scores[team] = scores

        # 답변
        cell = ws.cell(row=current_row, column=col_start)
        cell.value = scores['answer']
        cell.alignment = left_align
        cell.border = border

        # 핵심 점수
        cell = ws.cell(row=current_row, column=col_start + 1)
        cell.value = scores['core']
        cell.alignment = center_align
        cell.border = border
        if q_id.startswith('STD-S'):
            cell.fill = step_fill

        # 부가 점수
        cell = ws.cell(row=current_row, column=col_start + 2)
        cell.value = scores['addon']
        cell.alignment = center_align
        cell.border = border
        if q_id.startswith('STD-S'):
            cell.fill = step_fill

        # 최종 점수
        cell = ws.cell(row=current_row, column=col_start + 3)
        cell.value = scores['total']
        cell.alignment = center_align
        cell.border = border
        cell.font = Font(bold=True)

        # 색상
        if scores['total'] == 100:
            cell.fill = correct_fill
        elif scores['total'] >= 70:
            cell.fill = warning_fill
        else:
            cell.fill = incorrect_fill

    # 평가 결과 (최종 점수 평균)
    avg_score = sum(s['total'] for s in all_scores.values()) / len(all_scores)
    cell = ws.cell(row=current_row, column=20)
    cell.value = round(avg_score, 1)
    cell.alignment = center_align
    cell.border = border
    cell.font = Font(bold=True)

    current_row += 1

print(f"입력 완료: {processed}개 질문")

# ==================== 시트 2: 요약 ====================
ws_summary = wb.create_sheet("요약", 1)

summary_headers = ['팀', '전체', '100점', '70-99점', '70미만', '평균점수']
for col, header in enumerate(summary_headers, 1):
    cell = ws_summary.cell(row=1, column=col)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_align
    cell.border = border

for col in range(1, 7):
    ws_summary.column_dimensions[get_column_letter(col)].width = 12

summary_row = 2
for team in ['Team0', 'Team1', 'Team2', 'Team4']:
    scores_100 = 0
    scores_70_99 = 0
    scores_below_70 = 0
    total_scores = 0
    count = 0

    for q_id in sorted(new_questions.keys()):
        if q_id not in old_data:
            continue

        score_result = calculate_2step_score(team, old_data[q_id]['data'], q_id)
        total = score_result['total']

        if total == 100:
            scores_100 += 1
        elif total >= 70:
            scores_70_99 += 1
        else:
            scores_below_70 += 1

        total_scores += total
        count += 1

    avg_score = total_scores / count if count > 0 else 0

    ws_summary.cell(row=summary_row, column=1).value = team
    ws_summary.cell(row=summary_row, column=2).value = count
    ws_summary.cell(row=summary_row, column=3).value = scores_100
    ws_summary.cell(row=summary_row, column=4).value = scores_70_99
    ws_summary.cell(row=summary_row, column=5).value = scores_below_70
    ws_summary.cell(row=summary_row, column=6).value = round(avg_score, 1)

    for col in range(1, 7):
        cell = ws_summary.cell(row=summary_row, column=col)
        cell.border = border
        cell.alignment = center_align

    summary_row += 1

# ==================== 시트 3: 평가기준 ====================
ws_guide = wb.create_sheet("평가기준", 2)

guide_data = [
    ['2단계 평가 기준 (모델 3: 균형 평가)'],
    [],
    ['STEP 1: 핵심 정확도 (70% 가중치)', ''],
    ['STD-S 질문', '- 관련 없음/명시안됨 명시 = 70점'],
    ['', '- 명시 안 됨 = 30점'],
    ['일반 질문', '- 정답 = 70점 / 오답 = 30점'],
    [],
    ['STEP 2: 부가 설명 (30% 가중치)', ''],
    ['STD-S 질문', '- 정확+출처명시 = 30점'],
    ['', '- 일반원칙+출처미명시 = 25점'],
    ['', '- 부정확한설명 = 20점'],
    ['', '- 미언급 = 10점'],
    ['일반 질문', '- 정답 = 30점 / 오답 = 0점'],
    [],
    ['최종점수', '= 핵심 + 부가 (최대 100점)'],
    [],
    ['색상표시', '100점 = 녹색 | 70-99점 = 황색 | 미만 = 적색'],
]

for row, data in enumerate(guide_data, 1):
    for col, val in enumerate(data, 1):
        cell = ws_guide.cell(row=row, column=col)
        cell.value = val
        if row <= 1:
            cell.fill = header_fill
            cell.font = header_font
        cell.border = border if val else None
        cell.alignment = left_align

ws_guide.column_dimensions['A'].width = 25
ws_guide.column_dimensions['B'].width = 30

# 저장
output_path = r"E:\ontology_edu\X_ont_std\validation\ont_platform_v4_eval\reports\4팀_정확도_비교_2단계평가_최종.xlsx"
wb.save(output_path)

print("\n완료!")
print(f"파일: {output_path}")
print(f"시트: {wb.sheetnames}")
print(f"데이터: {processed}개 질문")
