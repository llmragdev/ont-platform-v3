# -*- coding: utf-8 -*-
"""
종합 평가 엑셀 생성 - 새 기준(v1.2) 적용
중간점검_통합테스트_예상 질문_답변-v1.2.xlsx + 4팀 기존 평가 병합
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict

print("Step 1: 기준 파일(v1.2) 읽기...")

new_file = r"E:\ai_lab_SIT\qa_expected\중간점검_통합테스트_예상 질문_답변-v1.2.xlsx"
wb_new = openpyxl.load_workbook(new_file)
ws_new = wb_new.active

# 새 기준에서 질문 정보 추출
questions_by_id = {}
category_stats = defaultdict(lambda: {'count': 0, 'list': []})

for row in range(2, ws_new.max_row + 1):
    q_id = ws_new.cell(row=row, column=1).value
    category = ws_new.cell(row=row, column=2).value
    q_type = ws_new.cell(row=row, column=3).value

    if q_id:
        questions_by_id[q_id] = {
            'id': q_id,
            'category': category,
            'type': q_type,
            'row': row
        }
        category_stats[category]['count'] += 1
        category_stats[category]['list'].append(q_id)

print(f"추출된 질문: {len(questions_by_id)}개")
print(f"카테고리: {list(category_stats.keys())}")

print("\nStep 2: 기존 평가 파일 읽기...")

old_file = r"E:\ontology_edu\X_ont_std\validation\ont_platform_v4_eval\reports\4팀_정확도_비교_STD-S카테고리무관_추가평가.xlsx"
wb_old = openpyxl.load_workbook(old_file)
ws_old = wb_old.active

# 기존 평가 데이터 추출
eval_data = {}

for row in range(2, ws_old.max_row + 1):
    q_id = ws_old.cell(row=row, column=1).value

    if q_id and q_id in questions_by_id:
        expected = ws_old.cell(row=row, column=3).value
        team0_ans = ws_old.cell(row=row, column=4).value
        team0_correct = ws_old.cell(row=row, column=5).value
        team1_ans = ws_old.cell(row=row, column=6).value
        team1_correct = ws_old.cell(row=row, column=7).value
        team2_ans = ws_old.cell(row=row, column=8).value
        team2_correct = ws_old.cell(row=row, column=9).value
        team4_ans = ws_old.cell(row=row, column=10).value
        team4_correct = ws_old.cell(row=row, column=11).value

        eval_data[q_id] = {
            'expected': expected,
            'team0': {'ans': team0_ans, 'correct': team0_correct},
            'team1': {'ans': team1_ans, 'correct': team1_correct},
            'team2': {'ans': team2_ans, 'correct': team2_correct},
            'team4': {'ans': team4_ans, 'correct': team4_correct},
        }

print(f"추출된 평가 데이터: {len(eval_data)}개")

print("\nStep 3: 새 통합 엑셀 생성...")

# 새 워크북
wb = openpyxl.Workbook()
wb.remove(wb.active)

# ==================== 시트 1: 전체 평가 결과 ====================
ws_main = wb.create_sheet("전체평가결과", 0)

# 스타일 정의
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)

cat_header_fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
cat_header_font = Font(bold=True, color="000000", size=10)

correct_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
incorrect_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

border = Border(
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000')
)

center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='top', wrap_text=True)

# 헤더
headers = ['질문ID', '카테고리', '유형', '예상답변',
           'Team0답변', 'Team0정답', 'Team0%',
           'Team1답변', 'Team1정답', 'Team1%',
           'Team2답변', 'Team2정답', 'Team2%',
           'Team4(v4)답변', 'Team4(v4)정답', 'Team4(v4)%',
           '비고']

for col, header in enumerate(headers, 1):
    cell = ws_main.cell(row=1, column=col)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_align
    cell.border = border

# 컬럼 너비
col_widths = [10, 12, 12, 25, 20, 12, 8, 20, 12, 8, 20, 12, 8, 20, 12, 8, 15]
for col, width in enumerate(col_widths, 1):
    ws_main.column_dimensions[get_column_letter(col)].width = width

# 데이터 입력
current_row = 2
current_category = None

for q_id in sorted(questions_by_id.keys()):
    q_info = questions_by_id[q_id]
    e_data = eval_data.get(q_id, {})

    # 카테고리 변경 시 배경색
    if q_info['category'] != current_category:
        current_category = q_info['category']

    # 질문ID
    cell = ws_main.cell(row=current_row, column=1)
    cell.value = q_id
    cell.alignment = center_align
    cell.border = border

    # 카테고리
    cell = ws_main.cell(row=current_row, column=2)
    cell.value = q_info['category']
    cell.alignment = center_align
    cell.border = border

    # 유형
    cell = ws_main.cell(row=current_row, column=3)
    cell.value = q_info['type']
    cell.alignment = center_align
    cell.border = border

    # 예상답변
    cell = ws_main.cell(row=current_row, column=4)
    cell.value = e_data.get('expected', '').strip()[:50] if e_data.get('expected') else ''
    cell.alignment = left_align
    cell.border = border

    # Team별 데이터 (Team0, Team1, Team2, Team4)
    team_cols = [
        (5, 'team0'),
        (8, 'team1'),
        (11, 'team2'),
        (14, 'team4'),
    ]

    for start_col, team_key in team_cols:
        team_data = e_data.get(team_key, {})

        # 답변
        ans_cell = ws_main.cell(row=current_row, column=start_col)
        ans_cell.value = str(team_data.get('ans', ''))[:30] if team_data.get('ans') else ''
        ans_cell.alignment = left_align
        ans_cell.border = border

        # 정답 여부
        correct_cell = ws_main.cell(row=current_row, column=start_col + 1)
        correct_val = team_data.get('correct', '')
        correct_cell.value = correct_val
        correct_cell.alignment = center_align
        correct_cell.border = border

        # 정답 여부에 따른 배경색
        if correct_val == '정답':
            correct_cell.fill = correct_fill
        elif correct_val == '오답':
            correct_cell.fill = incorrect_fill

        # 백분율
        pct_cell = ws_main.cell(row=current_row, column=start_col + 2)
        # 이 부분은 수동 계산 필요
        pct_cell.alignment = center_align
        pct_cell.border = border

    current_row += 1

print(f"데이터 입력: {current_row - 2}행")

# ==================== 시트 2: 카테고리별 요약 ====================
ws_cat = wb.create_sheet("카테고리별요약", 1)

cat_headers = ['카테고리', '전체수', 'Team0정답', 'Team0%', 'Team1정답', 'Team1%',
               'Team2정답', 'Team2%', 'Team4정답', 'Team4%']

for col, header in enumerate(cat_headers, 1):
    cell = ws_cat.cell(row=1, column=col)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_align
    cell.border = border

ws_cat.column_dimensions['A'].width = 15
for col in range(2, 11):
    ws_cat.column_dimensions[get_column_letter(col)].width = 12

cat_row = 2
for category in sorted(category_stats.keys()):
    ws_cat.cell(row=cat_row, column=1).value = category
    ws_cat.cell(row=cat_row, column=2).value = category_stats[category]['count']

    for col in range(1, 11):
        ws_cat.cell(row=cat_row, column=col).border = border
        ws_cat.cell(row=cat_row, column=col).alignment = center_align

    cat_row += 1

# ==================== 시트 3: 팀별 비교 ====================
ws_team = wb.create_sheet("팀별비교", 2)

team_headers = ['팀명', '전체수', '정답수', '정답율(%)', '평균신뢰도', '개선영역']

for col, header in enumerate(team_headers, 1):
    cell = ws_team.cell(row=1, column=col)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_align
    cell.border = border

teams = ['Team0', 'Team1', 'Team2', 'Team4(v4)']
for col in range(1, 7):
    ws_team.column_dimensions[get_column_letter(col)].width = 15

team_row = 2
for team in teams:
    ws_team.cell(row=team_row, column=1).value = team
    for col in range(1, 7):
        ws_team.cell(row=team_row, column=col).border = border
        ws_team.cell(row=team_row, column=col).alignment = center_align
    team_row += 1

# ==================== 시트 4: 기본정보 ====================
ws_info = wb.create_sheet("기본정보", 3)

info_data = [
    ['평가 기준', '값'],
    ['평가일자', '2026-06-08'],
    ['평가대상', 'ont_platform v4 vs v5 (STD-S 카테고리무관)'],
    ['질문수', len(questions_by_id)],
    ['카테고리수', len(category_stats)],
    ['평가팀수', 4],
]

for row, data in enumerate(info_data, 1):
    for col, val in enumerate(data, 1):
        cell = ws_info.cell(row=row, column=col)
        cell.value = val
        cell.border = border
        if row == 1:
            cell.fill = header_fill
            cell.font = header_font
        cell.alignment = center_align

ws_info.column_dimensions['A'].width = 20
ws_info.column_dimensions['B'].width = 30

# 파일 저장
output_path = r"E:\ontology_edu\X_ont_std\validation\ont_platform_v4_eval\reports\4팀_정확도_비교_통합검증_최종.xlsx"
wb.save(output_path)

print(f"\n완료!")
print(f"생성 파일: {output_path}")
print(f"시트 수: {len(wb.sheetnames)}")
print(f"시트명: {wb.sheetnames}")
print(f"전체 질문: {len(questions_by_id)}개")
print(f"카테고리: {len(category_stats)}개")
