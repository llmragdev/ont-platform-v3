# -*- coding: utf-8 -*-
"""
2단계 평가 기준 적용 엑셀 생성 (최종 버전)
각 질문별 시트에서 예상 답변 추출
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

print("=" * 80)
print("Step 1: 새 기준 파일 분석")
print("=" * 80)

# 새 기준 파일
new_file = r"E:\ai_lab_SIT\qa_expected\중간점검_통합테스트_예상 질문_답변-v1.2.xlsx"
wb_new = openpyxl.load_workbook(new_file)

print(f"\n총 시트 수: {len(wb_new.sheetnames)}")

# 각 질문별 시트에서 예상 답변 추출
new_expected = {}

for sheet_name in wb_new.sheetnames:
    if sheet_name.startswith('STD-'):
        ws = wb_new[sheet_name]
        q_id = sheet_name

        # 답변 추출 (최대 1000자)
        answer_cell = ws['A1']
        if answer_cell and answer_cell.value:
            expected = str(answer_cell.value)
            new_expected[q_id] = expected[:200]  # 처음 200자
            print(f"  {q_id}: {expected[:50]}...")

print(f"\n추출된 예상 답변: {len(new_expected)}개")

print("\n" + "=" * 80)
print("Step 2: 기존 평가 파일 분석")
print("=" * 80)

old_file = r"E:\ontology_edu\X_ont_std\validation\ont_platform_v4_eval\reports\4팀_정확도_비교_STD-S정답수정_추가평가.xlsx"
wb_old = openpyxl.load_workbook(old_file)
ws_old = wb_old.active

# 헤더 확인
headers = []
for col in range(1, 18):
    val = ws_old.cell(row=1, column=col).value
    if val:
        headers.append((col, val))

print(f"헤더: {[h[1] for h in headers]}")

# 데이터 추출
old_data = {}
for row in range(2, ws_old.max_row + 1):
    q_id = ws_old.cell(row=row, column=1).value

    if q_id:
        old_data[q_id] = {
            'row': row,
            'cols': {}
        }

        for col, header in headers:
            cell_val = ws_old.cell(row=row, column=col).value
            old_data[q_id]['cols'][header] = cell_val

print(f"추출된 평가: {len(old_data)}개")

print("\n" + "=" * 80)
print("Step 3: 2단계 평가 점수 계산")
print("=" * 80)

def calculate_2step_score(team_name, row_cols, q_id):
    """2단계 평가 점수 계산"""
    is_std_s = q_id.startswith('STD-S')

    # 컬럼 찾기
    answer_col = f'{team_name} 답변'
    correct_col = f'{team_name} 정답 여부'

    answer = str(row_cols.get(answer_col, '')) if answer_col in row_cols else ''
    correct = str(row_cols.get(correct_col, '')) if correct_col in row_cols else ''

    # Step 1: 핵심 정확도 (70점 또는 30점)
    if is_std_s:
        has_warning = (
            ('관련' in answer and '없' in answer) or
            ('명시' in answer and ('않' in answer or '안' in answer))
        )
        core = 70 if has_warning else 30
    else:
        core = 70 if '정답' in correct else 30

    # Step 2: 부가 설명 (30점, 25점, 20점, 10점, 또는 0점)
    addon = 0
    if is_std_s:
        if '논문' in answer and ('명시' in answer or '없' in answer):
            addon = 30
        elif '일반' in answer or '원칙' in answer:
            if '출처' in answer or '[' in answer:
                addon = 30
            else:
                addon = 25
        elif '오답' in correct:
            addon = 20
        elif answer and answer != 'None':
            addon = 10
        else:
            addon = 0
    else:
        addon = 30 if '정답' in correct else 0

    total = min(core + addon, 100)

    return {
        'core': core,
        'addon': addon,
        'total': total,
        'answer': answer[:25] if answer else '',
    }

print("점수 계산 준비 완료")

print("\n" + "=" * 80)
print("Step 4: 새 엑셀 생성")
print("=" * 80)

# 새 워크북
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "2단계평가결과"

# 스타일
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=10)

step_fill = PatternFill(start_color="D9E8F5", end_color="D9E8F5", fill_type="solid")
correct_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
incorrect_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
warning_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='top', wrap_text=True)

# 헤더 작성
new_headers = [
    '질문ID', '질문내용', '[v1.2]예상답변',
    'Team0답변', 'Team0_핵심', 'Team0_부가', 'Team0_최종',
    'Team1답변', 'Team1_핵심', 'Team1_부가', 'Team1_최종',
    'Team2답변', 'Team2_핵심', 'Team2_부가', 'Team2_최종',
    'Team4(v4)답변', 'Team4_핵심', 'Team4_부가', 'Team4_최종',
    '평가'
]

for col, header in enumerate(new_headers, 1):
    cell = ws.cell(row=1, column=col)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_align
    cell.border = border

# 컬럼 너비
widths = [9, 12, 18, 15, 9, 9, 9, 15, 9, 9, 9, 15, 9, 9, 9, 15, 9, 9, 9, 7]
for col, width in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(col)].width = width

# 데이터 입력
print("\n데이터 입력...")
current_row = 2
count = 0

for q_id in sorted(old_data.keys()):
    if q_id not in new_expected:
        continue

    count += 1
    row_cols = old_data[q_id]['cols']

    # 질문 ID
    cell = ws.cell(row=current_row, column=1)
    cell.value = q_id
    cell.alignment = center_align
    cell.border = border

    # 질문 내용
    cell = ws.cell(row=current_row, column=2)
    q_text = row_cols.get('질문', '')
    cell.value = str(q_text)[:20] if q_text else ''
    cell.alignment = left_align
    cell.border = border

    # v1.2 예상 답변
    cell = ws.cell(row=current_row, column=3)
    cell.value = new_expected.get(q_id, '')[:40]
    cell.alignment = left_align
    cell.border = border

    # Team별 평가
    teams = ['Team0', 'Team1', 'Team2', 'Team4']
    avg_total = 0

    for team_idx, team in enumerate(teams):
        col_start = 4 + (team_idx * 4)

        # 실제 Team 이름 매핑
        if team == 'Team4':
            team_display = 'Team4 답변 (ont_platform v4)'
        else:
            team_display = f'{team} 답변'

        # 답변 조회 (정확한 컬럼명)
        answer_key = None
        for key in row_cols.keys():
            if team in key and '답변' in key:
                answer_key = key
                break

        scores = calculate_2step_score(team_display if team_display in row_cols else team, row_cols, q_id)

        # 답변
        cell = ws.cell(row=current_row, column=col_start)
        cell.value = scores['answer']
        cell.alignment = left_align
        cell.border = border

        # 핵심
        cell = ws.cell(row=current_row, column=col_start + 1)
        cell.value = scores['core']
        cell.alignment = center_align
        cell.border = border
        if q_id.startswith('STD-S'):
            cell.fill = step_fill

        # 부가
        cell = ws.cell(row=current_row, column=col_start + 2)
        cell.value = scores['addon']
        cell.alignment = center_align
        cell.border = border
        if q_id.startswith('STD-S'):
            cell.fill = step_fill

        # 최종
        cell = ws.cell(row=current_row, column=col_start + 3)
        cell.value = scores['total']
        cell.alignment = center_align
        cell.border = border
        cell.font = Font(bold=True)

        if scores['total'] == 100:
            cell.fill = correct_fill
        elif scores['total'] >= 70:
            cell.fill = warning_fill
        else:
            cell.fill = incorrect_fill

        avg_total += scores['total']

    # 평가
    cell = ws.cell(row=current_row, column=20)
    cell.value = round(avg_total / 4, 1)
    cell.alignment = center_align
    cell.border = border
    cell.font = Font(bold=True)

    current_row += 1

print(f"입력 완료: {count}개 질문\n")

# ==================== 시트 2: 요약 ====================
ws_summary = wb.create_sheet("요약", 1)

s_headers = ['팀', '전체', '만점(100)', '양호(70-99)', '부족(<70)', '평균']
for col, header in enumerate(s_headers, 1):
    cell = ws_summary.cell(row=1, column=col)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_align
    cell.border = border

for col in range(1, 7):
    ws_summary.column_dimensions[get_column_letter(col)].width = 12

s_row = 2
for team in ['Team0', 'Team1', 'Team2', 'Team4']:
    cnt_100 = 0
    cnt_70_99 = 0
    cnt_below = 0
    sum_scores = 0
    total_cnt = 0

    for q_id in old_data.keys():
        if q_id not in new_expected:
            continue

        s = calculate_2step_score(team, old_data[q_id]['cols'], q_id)
        score = s['total']

        if score == 100:
            cnt_100 += 1
        elif score >= 70:
            cnt_70_99 += 1
        else:
            cnt_below += 1

        sum_scores += score
        total_cnt += 1

    avg = sum_scores / total_cnt if total_cnt > 0 else 0

    ws_summary.cell(row=s_row, column=1).value = team
    ws_summary.cell(row=s_row, column=2).value = total_cnt
    ws_summary.cell(row=s_row, column=3).value = cnt_100
    ws_summary.cell(row=s_row, column=4).value = cnt_70_99
    ws_summary.cell(row=s_row, column=5).value = cnt_below
    ws_summary.cell(row=s_row, column=6).value = round(avg, 1)

    for col in range(1, 7):
        cell = ws_summary.cell(row=s_row, column=col)
        cell.border = border
        cell.alignment = center_align

    s_row += 1

# ==================== 시트 3: 평가기준 ====================
ws_guide = wb.create_sheet("평가기준", 2)

guide = [
    ['2단계 평가 기준 (모델 3: 균형 평가)'],
    [],
    ['STEP 1: 핵심 정확도 (70점 가중치)'],
    ['STD-S 질문', '"관련없음" 명시 = 70점 / 미명시 = 30점'],
    ['일반 질문', '정답 = 70점 / 오답 = 30점'],
    [],
    ['STEP 2: 부가 설명 (30점 가중치)'],
    ['STD-S 질문', '정확+출처명시=30점, 일반원칙+출처미명시=25점'],
    ['', '부정확=20점, 미언급=10점, 없음=0점'],
    ['일반 질문', '정답=30점, 오답=0점'],
    [],
    ['최종점수', '= 핵심 + 부가 (최대 100점)'],
    ['색상표시', '100점=녹색 | 70-99점=황색 | 70미만=적색'],
]

for row, data in enumerate(guide, 1):
    for col, val in enumerate(data, 1):
        cell = ws_guide.cell(row=row, column=col)
        cell.value = val
        if row <= 1:
            cell.fill = header_fill
            cell.font = header_font
        cell.border = border if val else None
        cell.alignment = left_align

ws_guide.column_dimensions['A'].width = 20
ws_guide.column_dimensions['B'].width = 40

# 저장
output = r"E:\ontology_edu\X_ont_std\validation\ont_platform_v4_eval\reports\4팀_정확도_비교_2단계평가_최종.xlsx"
wb.save(output)

print("=" * 80)
print("완료!")
print("=" * 80)
print(f"\n파일: 4팀_정확도_비교_2단계평가_최종.xlsx")
print(f"시트: {wb.sheetnames}")
print(f"데이터: {count}개 질문")
print(f"위치: {output}")
