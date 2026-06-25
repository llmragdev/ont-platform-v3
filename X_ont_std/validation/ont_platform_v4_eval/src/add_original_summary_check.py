from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = next(p for p in Path(__file__).resolve().parents if (p / "ont_platform").exists())
SOURCE = ROOT / "validation" / "ont_platform_v4_eval" / "data" / "3팀_정확도_비교.xlsx"
OUTPUT = ROOT / "validation" / "ont_platform_v4_eval" / "reports" / "4팀_정확도_비교.xlsx"


def style_table(ws, rows: int, cols: int, name: str) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=1, max_row=rows, min_col=1, max_col=cols):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if cell.row == 1:
                cell.fill = fill
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ref = f"A1:{get_column_letter(cols)}{rows}"
    tab = Table(displayName=name, ref=ref)
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(tab)


def main() -> None:
    src = load_workbook(SOURCE, data_only=True)
    out = load_workbook(OUTPUT)

    if "원본 요약 대조" in out.sheetnames:
        del out["원본 요약 대조"]
    ws = out.create_sheet("원본 요약 대조")
    ws.sheet_view.showGridLines = False

    original_team = {
        str(row[0]): float(row[1])
        for row in src["팀별 요약"].iter_rows(min_row=2, values_only=True)
        if row[0]
    }
    new_team = {
        str(row[0]).replace(" (ont_platform v4)", ""): float(row[1])
        for row in out["팀별 요약"].iter_rows(min_row=2, values_only=True)
        if row[0]
    }

    rows = [["항목", "원본 요약 탭 (%)", "상세 탭 기반 새 보고서 (%)", "차이", "비고"]]
    for team in ("Team0", "Team1", "Team2"):
        old = original_team.get(team)
        new = new_team.get(team)
        rows.append([team, old, new, round(new - old, 2), "원본 엑셀 내부 요약값과 상세 점수 평균 차이"])
    rows.append(["Team4", None, new_team.get("Team4"), None, "ont_platform v4 동일 24문항 신규 평가"])
    rows.append([])
    rows.append(["판정", None, None, None, "이 파일은 사용자가 핵심이라고 지정한 '문항별 비교 상세' 점수를 기준으로 평균을 재계산함"])

    for row in rows:
        ws.append(row)
    style_table(ws, len(rows), 5, "OriginalSummaryCheck")
    widths = [18, 20, 26, 12, 70]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, float):
                cell.number_format = "0.00"

    out.save(OUTPUT)
    print(f"UPDATED {OUTPUT}")


if __name__ == "__main__":
    main()
