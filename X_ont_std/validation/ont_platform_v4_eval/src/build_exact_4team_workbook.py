from __future__ import annotations

import json
import os
import re
import shutil
import sys
import time
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = next(p for p in Path(__file__).resolve().parents if (p / "ont_platform").exists())
EVAL_DIR = ROOT / "validation" / "ont_platform_v4_eval"
V4_BACKEND = ROOT / "ont_platform" / "v4" / "backend"
SOURCE_XLSX = EVAL_DIR / "data" / "3팀_정확도_비교.xlsx"
OUTPUT_XLSX = EVAL_DIR / "reports" / "4팀_정확도_비교.xlsx"
RESULTS_DIR = EVAL_DIR / "results" / "same24"
TARGET_DOC_DIR = Path(r"E:\ai_lab_SIT\target_doc")
V3_ENV = ROOT / "ont_platform" / "v3" / ".env"

COMPANY_ID = "codex_eval"
PROJECT_ID = "ont_platform_v4_same24"
HEADERS = {
    "X-User-ID": "codex-evaluator",
    "X-Company-ID": COMPANY_ID,
    "X-Project-ID": PROJECT_ID,
    "X-Role": "Admin",
}


def load_env_file(path: Path) -> None:
    if not path.exists():
        return
    for raw in path.read_text(encoding="utf-8").splitlines():
        line = raw.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        if " #" in value:
            value = value.split(" #", 1)[0].strip()
        if "  " in value:
            value = value.split("  ", 1)[0].strip()
        if key and value and key not in os.environ:
            os.environ[key] = value

    default_key = os.environ.get("GEMINI_API_KEY", "")
    if len(default_key) > 45:
        for name in ("GEMINI_API_KEY1", "GEMINI_API_KEY2", "GEMINI_API_KEY3", "GEMINI_API_KEY4"):
            candidate = os.environ.get(name, "")
            if candidate.startswith("AIza") and len(candidate) <= 45:
                os.environ["GEMINI_API_KEY"] = candidate
                break


def load_same24_cases() -> list[dict[str, Any]]:
    wb = load_workbook(SOURCE_XLSX, data_only=True)
    item_ws = wb["문항별 비교"]
    detail_ws = wb["문항별 비교 상세"]
    category_by_id = {
        str(row[0]).strip(): str(row[1]).strip()
        for row in item_ws.iter_rows(min_row=2, values_only=True)
        if row[0]
    }

    cases = []
    for row in detail_ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        cases.append(
            {
                "problem_id": str(row[0]).strip(),
                "category": category_by_id.get(str(row[0]).strip(), ""),
                "question": str(row[1] or "").strip(),
                "expected_answer": str(row[2] or "").strip(),
                "team0_answer": str(row[3] or "").strip(),
                "team0_accuracy": float(row[4] or 0),
                "team1_answer": str(row[5] or "").strip(),
                "team1_accuracy": float(row[6] or 0),
                "team2_answer": str(row[7] or "").strip(),
                "team2_accuracy": float(row[8] or 0),
            }
        )
    return cases


def clean_eval_storage() -> None:
    sys.path.insert(0, str(V4_BACKEND))
    from storage_config import get_project_root

    project_root = get_project_root(COMPANY_ID, PROJECT_ID)
    if project_root.exists():
        shutil.rmtree(project_root)


def create_client():
    sys.path.insert(0, str(V4_BACKEND))
    from fastapi.testclient import TestClient
    from app.main import app

    return TestClient(app)


def upload_pdfs(client) -> list[dict[str, Any]]:
    uploads = []
    pdfs = sorted(TARGET_DOC_DIR.glob("*.pdf"))
    for idx, pdf in enumerate(pdfs, start=1):
        started = time.perf_counter()
        with pdf.open("rb") as fh:
            response = client.post(
                "/api/documents/upload",
                headers=HEADERS,
                files={"file": (pdf.name, fh, "application/pdf")},
                data={"shard_id": "default"},
                timeout=180,
            )
        elapsed_ms = round((time.perf_counter() - started) * 1000, 2)
        item = {
            "filename": pdf.name,
            "status_code": response.status_code,
            "elapsed_ms": elapsed_ms,
        }
        try:
            item["response"] = response.json()
        except Exception:
            item["response"] = {"text": response.text[:500]}
        uploads.append(item)
        print(f"UPLOAD {idx}/{len(pdfs)} {pdf.name} status={response.status_code} elapsed_ms={elapsed_ms}")
    return uploads


def extract_json(text: str) -> dict[str, Any] | None:
    cleaned = text.strip()
    if cleaned.startswith("```"):
        parts = cleaned.split("```")
        cleaned = parts[1] if len(parts) > 1 else cleaned
        if cleaned.lower().startswith("json"):
            cleaned = cleaned[4:]
    match = re.search(r"\{.*\}", cleaned, flags=re.S)
    if match:
        cleaned = match.group(0)
    try:
        return json.loads(cleaned)
    except Exception:
        return None


def lexical_score(expected: str, answer: str) -> tuple[float, str]:
    stopwords = {
        "그리고", "그러나", "따라서", "위해", "대한", "있는", "한다", "된다", "한다는",
        "통해", "기반", "경우", "같은", "서로", "이를", "등을", "에서", "으로",
    }
    tokens = re.findall(r"[가-힣A-Za-z0-9]{2,}", expected)
    keywords = []
    for token in tokens:
        if token in stopwords:
            continue
        if token not in keywords:
            keywords.append(token)
    keywords = keywords[:18]
    if not keywords:
        return 0.0, "키워드 추출 실패"
    hit = [kw for kw in keywords if kw in answer]
    score = round((len(hit) / len(keywords)) * 100, 1)
    return score, f"LLM 채점 실패 fallback: 예상답변 핵심어 {len(keywords)}개 중 {len(hit)}개 포함"


def judge_answer(llm, question: str, expected: str, answer: str) -> dict[str, Any]:
    prompt = f"""
당신은 RAG 질의응답 평가자입니다. 아래 질문의 예상 답변과 시스템 답변을 비교해 정확도를 0~100점으로 채점하세요.

채점 기준:
- 예상 답변의 핵심 개념, 관계, 절차, 조건을 얼마나 충족하는지 평가합니다.
- 표현이 달라도 의미가 같으면 인정합니다.
- 근거 없는 일반론, 누락, 반대 의미, 환각은 감점합니다.
- 너무 후한 점수를 주지 말고, 기존 평가표의 50/60/70/75/80/87.5 같은 실무 채점 감각에 맞춥니다.

반드시 JSON만 반환하세요.
{{
  "score": 0에서 100 사이 숫자,
  "rationale": "짧은 채점 근거",
  "missing": "누락되거나 약한 핵심"
}}

[질문]
{question}

[예상 답변]
{expected}

[시스템 답변]
{answer}
""".strip()
    text = llm.generate(prompt, temperature=0.0, max_tokens=512) if llm.enabled else None
    if text:
        parsed = extract_json(text)
        if parsed and "score" in parsed:
            score = max(0.0, min(100.0, float(parsed["score"])))
            return {
                "score": round(score, 1),
                "rationale": str(parsed.get("rationale", "")).strip(),
                "missing": str(parsed.get("missing", "")).strip(),
                "judge_raw": text,
                "judge_method": "Gemini LLM judge",
            }
    score, rationale = lexical_score(expected, answer)
    return {
        "score": score,
        "rationale": rationale,
        "missing": "LLM 채점 응답 파싱 실패",
        "judge_raw": text or "",
        "judge_method": "lexical fallback",
    }


def run_same24_eval() -> dict[str, Any]:
    RESULTS_DIR.mkdir(parents=True, exist_ok=True)
    load_env_file(V3_ENV)
    clean_eval_storage()
    client = create_client()
    uploads = upload_pdfs(client)

    sys.path.insert(0, str(V4_BACKEND))
    from app.services.llm_client import LlmClient

    judge_llm = LlmClient(model=os.getenv("LLM_MODEL_NAME") or "gemini-2.5-flash-lite")
    cases = load_same24_cases()
    results = []
    for idx, case in enumerate(cases, start=1):
        started = time.perf_counter()
        response = client.post(
            "/api/hybrid/ask",
            headers=HEADERS,
            json={"question": case["question"]},
            timeout=180,
        )
        elapsed_ms = round((time.perf_counter() - started) * 1000, 2)
        try:
            payload = response.json()
        except Exception:
            payload = {"answer": response.text[:1000]}
        answer = str(payload.get("answer", "") or "")
        vector_items = payload.get("structured_data", {}).get("vector", {}).get("items", [])
        judge = judge_answer(judge_llm, case["question"], case["expected_answer"], answer)
        result = {
            **case,
            "team4_answer": answer,
            "team4_accuracy": judge["score"],
            "team4_rationale": judge["rationale"],
            "team4_missing": judge["missing"],
            "team4_judge_method": judge["judge_method"],
            "team4_status_code": response.status_code,
            "team4_elapsed_ms": elapsed_ms,
            "team4_vector_hits": len(vector_items) if isinstance(vector_items, list) else 0,
            "team4_sources": [
                {
                    "filename": item.get("filename", ""),
                    "page": item.get("page", ""),
                    "score": item.get("score", ""),
                    "text": str(item.get("text", ""))[:350],
                }
                for item in vector_items[:5]
                if isinstance(item, dict)
            ],
            "team4_payload": payload,
        }
        results.append(result)
        print(
            f"QUERY {idx}/{len(cases)} {case['problem_id']} "
            f"score={judge['score']} status={response.status_code} elapsed_ms={elapsed_ms}"
        )

    summary = build_summary(results)
    out = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "source_xlsx": str(SOURCE_XLSX),
        "target_docs": [str(p) for p in sorted(TARGET_DOC_DIR.glob("*.pdf"))],
        "uploads": uploads,
        "results": results,
        "summary": summary,
    }
    (RESULTS_DIR / "same24_team4_results.json").write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
    return out


def build_summary(results: list[dict[str, Any]]) -> dict[str, Any]:
    teams = ["team0", "team1", "team2", "team4"]
    overall = {
        team: round(sum(r[f"{team}_accuracy"] for r in results) / len(results), 2)
        for team in teams
    }
    categories = sorted({r["category"] for r in results})
    by_category = {}
    for category in categories:
        rows = [r for r in results if r["category"] == category]
        by_category[category] = {
            team: round(sum(r[f"{team}_accuracy"] for r in rows) / len(rows), 2)
            for team in teams
        }
    ranks = sorted(overall.items(), key=lambda x: x[1], reverse=True)
    return {"overall": overall, "by_category": by_category, "ranks": ranks}


def apply_table_style(ws, start_row: int, start_col: int, end_row: int, end_col: int, name: str | None = None) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if cell.row == start_row:
                cell.fill = header_fill
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    if name:
        ref = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"
        tab = Table(displayName=name, ref=ref)
        tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        ws.add_table(tab)


def write_rows(ws, headers: list[str], rows: list[list[Any]], table_name: str) -> None:
    ws.append(headers)
    for row in rows:
        ws.append(row)
    apply_table_style(ws, 1, 1, len(rows) + 1, len(headers), table_name)
    ws.freeze_panes = "A2"
    for col in range(1, len(headers) + 1):
        width = 14
        if col in (2, 3, 4, 6, 8, 10, 12, 13):
            width = 42
        if col == 1:
            width = 13
        ws.column_dimensions[get_column_letter(col)].width = width


def build_workbook(eval_data: dict[str, Any]) -> None:
    results = eval_data["results"]
    summary = eval_data["summary"]

    wb = Workbook()
    ws = wb.active
    ws.title = "문항별 비교 상세"
    detail_headers = [
        "문제ID",
        "질문",
        "예상 답변",
        "Team0 답변",
        "Team0 정확도 (%)",
        "Team1 답변",
        "Team1 정확도 (%)",
        "Team2 답변",
        "Team2 정확도 (%)",
        "Team4 답변 (ont_platform v4)",
        "Team4 정확도 (%)",
        "Team4 채점 근거",
        "Team4 검색 근거",
    ]
    detail_rows = []
    for r in results:
        sources = "\n\n".join(
            f"{idx}. {s.get('filename', '')} p.{s.get('page', '')} score={s.get('score', '')}\n{s.get('text', '')}"
            for idx, s in enumerate(r["team4_sources"], start=1)
        )
        detail_rows.append(
            [
                r["problem_id"],
                r["question"],
                r["expected_answer"],
                r["team0_answer"],
                r["team0_accuracy"],
                r["team1_answer"],
                r["team1_accuracy"],
                r["team2_answer"],
                r["team2_accuracy"],
                r["team4_answer"],
                r["team4_accuracy"],
                f"{r['team4_rationale']}\n누락/약점: {r['team4_missing']}\n채점방식: {r['team4_judge_method']}",
                sources,
            ]
        )
    write_rows(ws, detail_headers, detail_rows, "QuestionDetail")
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 48
    ws.column_dimensions["J"].width = 54
    ws.column_dimensions["L"].width = 42
    ws.column_dimensions["M"].width = 48

    ws2 = wb.create_sheet("문항별 비교")
    item_headers = ["문제ID", "카테고리", "질문", "Team0 (%)", "Team1 (%)", "Team2 (%)", "Team4 (%)", "4팀 평균 (%)"]
    item_rows = [
        [
            r["problem_id"],
            r["category"],
            r["question"],
            r["team0_accuracy"],
            r["team1_accuracy"],
            r["team2_accuracy"],
            r["team4_accuracy"],
            round((r["team0_accuracy"] + r["team1_accuracy"] + r["team2_accuracy"] + r["team4_accuracy"]) / 4, 2),
        ]
        for r in results
    ]
    write_rows(ws2, item_headers, item_rows, "QuestionSummary")

    ws3 = wb.create_sheet("카테고리별 요약")
    cat_headers = ["카테고리", "Team0 (%)", "Team1 (%)", "Team2 (%)", "Team4 (%)", "4팀 평균 (%)"]
    cat_rows = []
    for category, scores in summary["by_category"].items():
        avg4 = round((scores["team0"] + scores["team1"] + scores["team2"] + scores["team4"]) / 4, 2)
        cat_rows.append([category, scores["team0"], scores["team1"], scores["team2"], scores["team4"], avg4])
    overall = summary["overall"]
    cat_rows.append(
        [
            "전체 평균",
            overall["team0"],
            overall["team1"],
            overall["team2"],
            overall["team4"],
            round((overall["team0"] + overall["team1"] + overall["team2"] + overall["team4"]) / 4, 2),
        ]
    )
    write_rows(ws3, cat_headers, cat_rows, "CategorySummary")

    ws4 = wb.create_sheet("팀별 요약")
    rank_by_team = {team: idx for idx, (team, _) in enumerate(summary["ranks"], start=1)}
    label = {"team0": "Team0", "team1": "Team1", "team2": "Team2", "team4": "Team4 (ont_platform v4)"}
    team_rows = []
    for team in ("team0", "team1", "team2", "team4"):
        rank = rank_by_team[team]
        recommendation = "도입 우선 검토" if rank == 1 else "도입 고려" if rank == 2 else "개선 필요"
        team_rows.append([label[team], overall[team], f"{rank}위", recommendation])
    write_rows(ws4, ["팀", "전체 정확도 (%)", "순위", "권고"], team_rows, "TeamSummary")

    ws5 = wb.create_sheet("Team4 평가 원자료")
    raw_headers = [
        "문제ID",
        "상태코드",
        "응답시간(ms)",
        "Vector Hits",
        "Team4 정확도 (%)",
        "채점방식",
        "질문",
        "예상 답변",
        "Team4 답변",
        "채점 근거",
    ]
    raw_rows = [
        [
            r["problem_id"],
            r["team4_status_code"],
            r["team4_elapsed_ms"],
            r["team4_vector_hits"],
            r["team4_accuracy"],
            r["team4_judge_method"],
            r["question"],
            r["expected_answer"],
            r["team4_answer"],
            f"{r['team4_rationale']}\n누락/약점: {r['team4_missing']}",
        ]
        for r in results
    ]
    write_rows(ws5, raw_headers, raw_rows, "Team4Raw")

    ws6 = wb.create_sheet("채점 방식")
    scoring_rows = [
        ["항목", "내용"],
        ["비교 기준", "3팀_정확도_비교.xlsx의 24개 동일 문항과 동일 예상 답변을 사용"],
        ["Team0~2", "원본 엑셀의 답변과 정확도 값을 그대로 보존"],
        ["Team4", "ont_platform v4에 동일 질문을 질의하고, 생성 답변을 예상 답변과 비교 채점"],
        ["Team4 채점자", "Gemini LLM judge. JSON 파싱 실패 시 예상답변 핵심어 포함률 fallback 사용"],
        ["주의", "Team4 점수는 자동 채점값이므로 최종 승인 전 표본 수동 검토를 권장"],
        ["생성일", eval_data["generated_at"]],
        ["원본 파일", str(SOURCE_XLSX)],
    ]
    for row in scoring_rows:
        ws6.append(row)
    apply_table_style(ws6, 1, 1, len(scoring_rows), 2, "ScoringMethod")
    ws6.column_dimensions["A"].width = 18
    ws6.column_dimensions["B"].width = 90

    for sheet in wb.worksheets:
        sheet.sheet_view.showGridLines = False
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, float):
                    cell.number_format = "0.00"

    wb.save(OUTPUT_XLSX)
    print(f"WROTE {OUTPUT_XLSX}")


def main() -> None:
    eval_data = run_same24_eval()
    build_workbook(eval_data)


if __name__ == "__main__":
    main()
