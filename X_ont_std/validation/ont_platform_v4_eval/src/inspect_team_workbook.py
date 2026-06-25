from pathlib import Path

from openpyxl import load_workbook


ROOT = next(p for p in Path(__file__).resolve().parents if (p / "ont_platform").exists())
SOURCE = ROOT / "validation" / "ont_platform_v4_eval" / "data" / "3팀_정확도_비교.xlsx"


def main() -> None:
    wb = load_workbook(SOURCE, data_only=True)
    print("sheets")
    for i, ws in enumerate(wb.worksheets):
        print(i, repr(ws.title), ws.max_row, ws.max_column)

    target = next((ws for ws in wb.worksheets if "상세" in ws.title), wb.worksheets[1])
    print("target", repr(target.title))
    for row in target.iter_rows(min_row=1, max_row=min(target.max_row, 12), values_only=True):
        print([str(v)[:180] if v is not None else None for v in row])


if __name__ == "__main__":
    main()
