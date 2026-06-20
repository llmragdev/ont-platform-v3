from pathlib import Path

from openpyxl import load_workbook


ROOT = next(p for p in Path(__file__).resolve().parents if (p / "ont_platform").exists())
OUTPUT = ROOT / "validation" / "ont_platform_v4_eval" / "reports" / "4팀_정확도_비교.xlsx"


def main() -> None:
    print("exists", OUTPUT.exists(), "size", OUTPUT.stat().st_size if OUTPUT.exists() else None)
    wb = load_workbook(OUTPUT, data_only=True)
    print("sheets", wb.sheetnames)

    detail = wb["문항별 비교 상세"]
    print("detail rows/cols", detail.max_row, detail.max_column)
    print("detail headers", [detail.cell(1, c).value for c in range(1, detail.max_column + 1)])
    print("first detail", [detail.cell(2, c).value for c in (1, 2, 5, 7, 9, 11, 12)])

    team = wb["팀별 요약"]
    print("team summary")
    for row in team.iter_rows(min_row=1, max_row=team.max_row, values_only=True):
        print(row)

    category = wb["카테고리별 요약"]
    print("category summary")
    for row in category.iter_rows(min_row=1, max_row=category.max_row, values_only=True):
        print(row)


if __name__ == "__main__":
    main()
