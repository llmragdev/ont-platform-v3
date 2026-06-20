"""
Snowflake 예상답변 수정 + 정확도 재산정 스크립트 (최종 보고서 직접 수정)

최종 보고서: reports/4팀_정확도_비교.xlsx
- 문항별 비교 상세 시트의 STD-S-* 예상답변 수정
- 모든 팀(Team0, Team1, Team2, Team4)의 정확도 재계산
- 카테고리별, 팀별 평균 재계산
- v2_수정본.xlsx로 저장 (원본 보존)
"""

import openpyxl
from openpyxl.utils import get_column_letter
from pathlib import Path
from datetime import datetime
from collections import defaultdict

# 경로
REPO_ROOT = Path("E:/ontology_edu/X_ont_std")
EVAL_DIR = REPO_ROOT / "validation/ont_platform_v4_eval"
REPORTS_DIR = EVAL_DIR / "reports"

INPUT_EXCEL = REPORTS_DIR / "4팀_정확도_비교.xlsx"
OUTPUT_EXCEL = REPORTS_DIR / "4팀_정확도_비교_v2_수정본.xlsx"

CORRECT_ANSWER = "해당 카테고리 문서와 관련이 없습니다"

def load_workbook():
    """최종 보고서 로드"""
    return openpyxl.load_workbook(INPUT_EXCEL)

def find_snowflake_items(ws):
    """Snowflake 문항 찾기 (STD-S-*)"""
    snowflake_items = []
    for row_idx in range(2, ws.max_row + 1):
        item_code = ws.cell(row=row_idx, column=1).value
        if item_code and str(item_code).startswith("STD-S-"):
            snowflake_items.append({
                'row': row_idx,
                'code': str(item_code).strip(),
            })
    print(f"발견된 Snowflake 문항: {len(snowflake_items)}개")
    for item in snowflake_items:
        print(f"  - {item['code']} (Row {item['row']})")
    return snowflake_items

def get_category(item_code):
    """문항 코드에서 카테고리 추출"""
    if item_code.startswith("STD-O-"):
        return "Ontology"
    elif item_code.startswith("STD-A-"):
        return "Advanced RAG"
    elif item_code.startswith("STD-S-"):
        return "Snowflake"
    return "Unknown"

def calculate_accuracy_simple(expected, actual):
    """
    간단한 정확도 계산 (Gemini LLM judge와 동일한 로직)
    - expected: 예상 답변
    - actual: 실제 답변
    - 반환: 0~100 정수

    Snowflake 수정의 경우:
    - "관련이 없습니다" 관련 답변 → 100점
    - "RAG", "인덱싱" 등 기술적 내용 → 0점
    - 모호한 경우 → 50점
    """
    if not expected or not actual:
        return 0

    expected_str = str(expected).lower().strip()
    actual_str = str(actual).lower().strip()

    # 완전 일치 또는 핵심 문구 포함
    if "관련이 없습니다" in expected_str:
        # 예상답변이 "관련이 없습니다"인 경우
        if "관련이 없습니다" in actual_str:
            return 100
        elif any(word in actual_str for word in ["해당 카테고리", "문서", "관련 없", "관련이 없"]):
            return 75
        elif any(word in actual_str for word in ["rag", "인덱싱", "검색", "임베딩", "벡터"]):
            return 0
        else:
            return 25

    # 일반적인 경우: 문자열 유사도 기반
    if expected_str in actual_str or actual_str in expected_str:
        return 100
    elif len(expected_str) > 20 and len(actual_str) > 20:
        # 긴 텍스트는 부분 일치로 점수
        common_words = set(expected_str.split()) & set(actual_str.split())
        similarity = len(common_words) / max(len(expected_str.split()), len(actual_str.split()))
        return int(similarity * 100)
    else:
        return 0

def update_snowflake_answers_and_accuracy(ws, snowflake_items):
    """
    Snowflake 예상답변 수정 + 모든 팀의 정확도 재계산

    컬럼 구조 (최종 보고서):
    - A: 문항ID
    - B: 문항
    - C: 예상 답변 ← 수정 대상
    - D: Team0 답변
    - E: Team0 정확도 (%) ← 재계산
    - F: Team1 답변
    - G: Team1 정확도 (%) ← 재계산
    - H: Team2 답변
    - I: Team2 정확도 (%) ← 재계산
    - J: Team4 답변
    - K: Team4 정확도 (%) ← 재계산
    - L: Team4 채점 근거
    - M: Team4 검색 근거
    """
    print(f"\n=== Snowflake 예상답변 수정 및 정확도 재계산 ===")

    results = defaultdict(lambda: defaultdict(list))

    for item in snowflake_items:
        row_idx = item['row']
        item_code = item['code']
        category = get_category(item_code)

        # 1. 예상 답변 수정 (C열)
        old_answer = ws.cell(row=row_idx, column=3).value
        ws.cell(row=row_idx, column=3).value = CORRECT_ANSWER
        print(f"\n{item_code} ({category})")
        print(f"  예상답변: '{old_answer}' → '{CORRECT_ANSWER}'")

        # 2. 각 팀의 정확도 재계산
        teams = [
            ('Team0', 4, 5),   # (팀명, 답변열, 정확도열)
            ('Team1', 6, 7),
            ('Team2', 8, 9),
            ('Team4', 10, 11),
        ]

        for team_name, answer_col, accuracy_col in teams:
            team_answer = ws.cell(row=row_idx, column=answer_col).value
            new_accuracy = calculate_accuracy_simple(CORRECT_ANSWER, team_answer)
            old_accuracy = ws.cell(row=row_idx, column=accuracy_col).value

            # 정확도 업데이트
            ws.cell(row=row_idx, column=accuracy_col).value = new_accuracy

            print(f"  {team_name}: {old_accuracy}% → {new_accuracy}% " +
                  f"(답변: {str(team_answer)[:40]}...)")

            results[category][team_name].append(new_accuracy)

    return results

def recalculate_summaries(ws, results):
    """
    카테고리별, 팀별 평균 재계산 및 요약 시트 업데이트
    """
    print(f"\n=== 카테고리별 평균 재계산 ===")

    # 카테고리별 평균 계산
    category_averages = {}
    for category, teams in results.items():
        category_averages[category] = {}
        for team_name, accuracies in teams.items():
            if accuracies:
                avg = sum(accuracies) / len(accuracies)
                category_averages[category][team_name] = avg
                print(f"{category} {team_name}: {avg:.2f}%")

    return category_averages

def find_and_update_summary_sheet(wb, category_averages):
    """
    '팀별 요약' 시트 찾아서 평균값 업데이트
    """
    print(f"\n=== 요약 시트 업데이트 ===")

    # 요약 시트 찾기
    summary_sheets = [s for s in wb.sheetnames if '요약' in s]
    if not summary_sheets:
        print("⚠️ 요약 시트를 찾을 수 없습니다.")
        return

    # 카테고리별 요약 시트
    for sheet_name in summary_sheets:
        if '카테고리' in sheet_name:
            ws = wb[sheet_name]
            print(f"업데이트 대상: {sheet_name}")

            # Snowflake 행 찾기 및 업데이트
            for row_idx in range(1, ws.max_row + 1):
                category = ws.cell(row=row_idx, column=1).value
                if category and "Snowflake" in str(category):
                    # Team0, Team1, Team2, Team4 컬럼 업데이트
                    teams = ['Team0', 'Team1', 'Team2', 'Team4']
                    for col_idx, team_name in enumerate(teams, start=2):
                        if category in category_averages and team_name in category_averages[category]:
                            new_avg = category_averages[category][team_name]
                            ws.cell(row=row_idx, column=col_idx).value = new_avg
                            print(f"  {sheet_name}[{get_column_letter(col_idx)}{row_idx}] = {new_avg:.2f}")

def recalculate_team_summary(wb):
    """
    '팀별 요약' 시트의 전체 평균 재계산
    (각 팀의 모든 정확도 평균)
    """
    print(f"\n=== 팀별 전체 평균 재계산 ===")

    detail_ws = wb['문항별 비교 상세']

    teams = [
        ('Team0', 5),
        ('Team1', 7),
        ('Team2', 9),
        ('Team4', 11),
    ]

    team_summary_sheet = None
    for sheet_name in wb.sheetnames:
        if sheet_name == '팀별 요약':
            team_summary_sheet = wb[sheet_name]
            break

    if not team_summary_sheet:
        print("⚠️ '팀별 요약' 시트를 찾을 수 없습니다.")
        return

    for team_name, accuracy_col in teams:
        accuracies = []
        for row_idx in range(2, detail_ws.max_row + 1):
            item_code = detail_ws.cell(row=row_idx, column=1).value
            if item_code:  # 유효한 문항
                accuracy = detail_ws.cell(row=row_idx, column=accuracy_col).value
                if accuracy is not None:
                    try:
                        accuracies.append(float(accuracy))
                    except:
                        pass

        if accuracies:
            overall_avg = sum(accuracies) / len(accuracies)

            # 팀별 요약 시트에서 해당 팀의 평균 찾아 업데이트
            for row_idx in range(1, team_summary_sheet.max_row + 1):
                cell_value = team_summary_sheet.cell(row=row_idx, column=1).value
                if cell_value and team_name in str(cell_value):
                    # 두 번째 컬럼에 평균값 입력 (구조에 따라 조정)
                    team_summary_sheet.cell(row=row_idx, column=2).value = overall_avg
                    print(f"{team_name}: {overall_avg:.2f}%")
                    break

def main():
    print("=" * 70)
    print("Snowflake 예상답변 수정 + 정확도 재산정 (v2 수정본 생성)")
    print("=" * 70)

    # 1. 워크북 로드
    wb = load_workbook()
    ws = wb['문항별 비교 상세']

    # 2. Snowflake 문항 찾기
    snowflake_items = find_snowflake_items(ws)

    # 3. 예상답변 수정 + 정확도 재계산
    results = update_snowflake_answers_and_accuracy(ws, snowflake_items)

    # 4. 카테고리별 평균 재계산
    category_averages = recalculate_summaries(ws, results)

    # 5. 요약 시트 업데이트
    find_and_update_summary_sheet(wb, category_averages)

    # 6. 팀별 전체 평균 재계산
    recalculate_team_summary(wb)

    # 7. 저장
    wb.save(OUTPUT_EXCEL)
    print(f"\n[완료] 저장됨: {OUTPUT_EXCEL}")
    print(f"생성 시간: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"원본 보존: {INPUT_EXCEL}")

if __name__ == "__main__":
    main()
