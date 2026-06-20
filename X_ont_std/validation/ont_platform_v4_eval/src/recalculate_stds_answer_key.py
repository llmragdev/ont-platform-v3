from __future__ import annotations

import json
import os
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = next(p for p in Path(__file__).resolve().parents if (p / "ont_platform").exists())
EVAL_DIR = ROOT / "validation" / "ont_platform_v4_eval"
V4_BACKEND = ROOT / "ont_platform" / "v4" / "backend"
V3_ENV = ROOT / "ont_platform" / "v3" / ".env"

BASE_XLSX = EVAL_DIR / "reports" / "4팀_정확도_비교.xlsx"
OUT_XLSX = EVAL_DIR / "reports" / "4팀_정확도_비교_STD-S정답수정_추가평가.xlsx"
OUT_MD = EVAL_DIR / "reports" / "STD-S정답수정_추가평가_보고서.md"
OUT_JSON = EVAL_DIR / "results" / "revised_stds" / "stds_rescore_results.json"

REVISED_STDS_EXPECTED = (
    "현재 평가 문서셋에는 Snowflake, Snowflake RAG, ranking_issue, Snowflake 기반 QA 테스트에 대한 "
    "직접적인 근거가 없다. 따라서 이 질문에는 '관련 문서 없음', '제공된 문서에서 확인 불가', "
    "'근거 부족으로 답변 불가'라고 답하는 것이 정답이다. 문서 근거 없이 일반 RAG, 온톨로지, "
    "엔터프라이즈 운영 지식을 이용해 Snowflake 관련 답을 생성하면 오답으로 본다."
)


def load_env_file(path: Path) -> None:
    if not path.exists():
        return
    for raw in path.read_text(encoding="utf-8").splitlines():
        line = raw.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        if " #" in value:
            value = value.split(" #", 1)[0].strip()
        if "  " in value:
            value = value.split("  ", 1)[0].strip()
        if key and value and key not in os.environ:
            os.environ[key] = value

    default_key = os.environ.get("GEMINI_API_KEY", "")
    if len(default_key) > 45:
        for name in ("GEMINI_API_KEY1", "GEMINI_API_KEY2", "GEMINI_API_KEY3", "GEMINI_API_KEY4"):
            candidate = os.environ.get(name, "")
            if candidate.startswith("AIza") and len(candidate) <= 45:
                os.environ["GEMINI_API_KEY"] = candidate
                break


def extract_json(text: str) -> dict[str, Any] | None:
    cleaned = text.strip()
    if cleaned.startswith("```"):
        parts = cleaned.split("```")
        cleaned = parts[1] if len(parts) > 1 else cleaned
        if cleaned.lower().startswith("json"):
            cleaned = cleaned[4:]
    match = re.search(r"\{.*\}", cleaned, flags=re.S)
    if match:
        cleaned = match.group(0)
    try:
        return json.loads(cleaned)
    except Exception:
        return None


def fallback_score(answer: str) -> tuple[float, str]:
    text = answer.strip()
    absence_terms = [
        "찾지 못", "없", "관련", "확인할 수", "확인할 수 없", "정보를 찾", "직접적", "근거", "제공된 문서",
        "답변하기 어렵", "키워드를 더 구체", "사내 지식 베이스",
    ]
    snowflake_terms = ["Snowflake", "스노우플레이크", "ranking_issue", "RAG", "QA", "평가", "정형", "테이블"]
    absence_hits = sum(1 for term in absence_terms if term in text)
    generated_hits = sum(1 for term in snowflake_terms if term in text)
    if absence_hits >= 3 and len(text) < 500:
        return 95.0, "fallback: 문서 근거 없음/확인 불가 취지의 답변"
    if absence_hits >= 2:
        return 80.0, "fallback: 근거 부족 취지는 있으나 부가 설명이 있음"
    if generated_hits >= 3 and len(text) > 120:
        return 10.0, "fallback: 관련 없음 대신 Snowflake/RAG 일반론을 생성"
    return 30.0, "fallback: 관련 없음 취지가 약하거나 불명확"


def judge_answer(llm, question: str, answer: str) -> dict[str, Any]:
    prompt = f"""
당신은 RAG 평가 정답표를 수정한 뒤 답변을 재채점하는 평가자입니다.

이번 STD-S 문항의 새 정답 기준:
{REVISED_STDS_EXPECTED}

채점 기준:
- 답변이 현재 문서셋에 Snowflake 관련 직접 근거가 없다고 명확히 말하면 90~100점입니다.
- '사내 지식 베이스/제공 문서에서 직접 정보를 찾지 못했다'는 취지면 높게 평가합니다.
- 근거 없음이라고 말한 뒤 약간의 일반 설명을 덧붙이면 70~85점입니다.
- 관련 없음이라고 하지 않고 Snowflake/RAG 일반론을 정답처럼 생성하면 0~30점입니다.
- 질문의 특정 용어를 억지로 온톨로지/RAG 일반 지식에 연결하면 낮게 평가합니다.

반드시 JSON만 반환하세요.
{{
  "score": 0에서 100 사이 숫자,
  "rationale": "짧은 채점 근거"
}}

[질문]
{question}

[답변]
{answer}
""".strip()
    text = llm.generate(prompt, temperature=0.0, max_tokens=384) if llm.enabled else None
    if text:
        parsed = extract_json(text)
        if parsed and "score" in parsed:
            score = max(0.0, min(100.0, float(parsed["score"])))
            return {
                "score": round(score, 1),
                "rationale": str(parsed.get("rationale", "")).strip(),
                "method": "Gemini LLM judge",
                "raw": text,
            }
    score, rationale = fallback_score(answer)
    return {"score": score, "rationale": rationale, "method": "fallback heuristic", "raw": text or ""}


def style_sheet(ws) -> None:
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if cell.row == 1:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if isinstance(cell.value, float):
                cell.number_format = "0.00"


def recompute_summary(wb) -> dict[str, Any]:
    detail = wb["문항별 비교 상세"]
    item = wb["문항별 비교"]
    category = wb["카테고리별 요약"]
    team = wb["팀별 요약"]

    rows = []
    category_by_id = {}
    for r in item.iter_rows(min_row=2, values_only=True):
        if r[0]:
            category_by_id[str(r[0])] = str(r[1])

    for idx in range(2, detail.max_row + 1):
        pid = detail.cell(idx, 1).value
        if not pid:
            continue
        scores = {
            "Team0": float(detail.cell(idx, 5).value or 0),
            "Team1": float(detail.cell(idx, 7).value or 0),
            "Team2": float(detail.cell(idx, 9).value or 0),
            "Team4": float(detail.cell(idx, 11).value or 0),
        }
        rows.append({"id": str(pid), "category": category_by_id.get(str(pid), ""), "scores": scores})

    # Update item sheet.
    for idx in range(2, item.max_row + 1):
        pid = item.cell(idx, 1).value
        if not pid:
            continue
        match = next(r for r in rows if r["id"] == str(pid))
        scores = match["scores"]
        item.cell(idx, 4, scores["Team0"])
        item.cell(idx, 5, scores["Team1"])
        item.cell(idx, 6, scores["Team2"])
        item.cell(idx, 7, scores["Team4"])
        item.cell(idx, 8, round(sum(scores.values()) / 4, 2))

    categories = sorted({r["category"] for r in rows})
    summary_by_cat = {}
    for cat in categories:
        cat_rows = [r for r in rows if r["category"] == cat]
        summary_by_cat[cat] = {
            team_name: round(sum(r["scores"][team_name] for r in cat_rows) / len(cat_rows), 2)
            for team_name in ("Team0", "Team1", "Team2", "Team4")
        }

    overall = {
        team_name: round(sum(r["scores"][team_name] for r in rows) / len(rows), 2)
        for team_name in ("Team0", "Team1", "Team2", "Team4")
    }

    # Rewrite category summary values in existing layout.
    for row_idx in range(2, category.max_row + 1):
        cat = category.cell(row_idx, 1).value
        if not cat:
            continue
        scores = overall if cat == "전체 평균" else summary_by_cat.get(str(cat))
        if not scores:
            continue
        values = [scores["Team0"], scores["Team1"], scores["Team2"], scores["Team4"]]
        for col, value in enumerate(values, start=2):
            category.cell(row_idx, col, value)
        category.cell(row_idx, 6, round(sum(values) / 4, 2))

    ranks = sorted(overall.items(), key=lambda x: x[1], reverse=True)
    rank_by_team = {name: idx for idx, (name, _) in enumerate(ranks, start=1)}
    labels = {
        "Team0": "Team0",
        "Team1": "Team1",
        "Team2": "Team2",
        "Team4": "Team4 (ont_platform v4)",
    }
    for idx, team_name in enumerate(("Team0", "Team1", "Team2", "Team4"), start=2):
        rank = rank_by_team[team_name]
        team.cell(idx, 1, labels[team_name])
        team.cell(idx, 2, overall[team_name])
        team.cell(idx, 3, f"{rank}위")
        team.cell(idx, 4, "도입 우선 검토" if rank == 1 else "도입 고려" if rank == 2 else "개선 필요")

    for ws in (detail, item, category, team):
        style_sheet(ws)

    return {"overall": overall, "by_category": summary_by_cat, "ranks": ranks}


def main() -> None:
    OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)

    load_env_file(V3_ENV)
    sys.path.insert(0, str(V4_BACKEND))
    from app.services.llm_client import LlmClient

    llm = LlmClient(model=os.getenv("LLM_MODEL_NAME") or "gemini-2.5-flash-lite")
    wb = load_workbook(BASE_XLSX)
    detail = wb["문항별 비교 상세"]

    headers = [detail.cell(1, c).value for c in range(1, detail.max_column + 1)]
    header_to_col = {name: idx for idx, name in enumerate(headers, start=1)}
    answer_cols = {
        "Team0": header_to_col["Team0 답변"],
        "Team1": header_to_col["Team1 답변"],
        "Team2": header_to_col["Team2 답변"],
        "Team4": header_to_col["Team4 답변 (ont_platform v4)"],
    }
    score_cols = {
        "Team0": header_to_col["Team0 정확도 (%)"],
        "Team1": header_to_col["Team1 정확도 (%)"],
        "Team2": header_to_col["Team2 정확도 (%)"],
        "Team4": header_to_col["Team4 정확도 (%)"],
    }

    rescore_rows = []
    for row_idx in range(2, detail.max_row + 1):
        pid = str(detail.cell(row_idx, 1).value or "")
        if not pid.startswith("STD-S"):
            continue
        question = str(detail.cell(row_idx, 2).value or "")
        old_expected = str(detail.cell(row_idx, 3).value or "")
        detail.cell(row_idx, 3, REVISED_STDS_EXPECTED)
        for team_name in ("Team0", "Team1", "Team2", "Team4"):
            answer = str(detail.cell(row_idx, answer_cols[team_name]).value or "")
            judge = judge_answer(llm, question, answer)
            detail.cell(row_idx, score_cols[team_name], judge["score"])
            if team_name == "Team4":
                current = str(detail.cell(row_idx, header_to_col["Team4 채점 근거"]).value or "")
                detail.cell(
                    row_idx,
                    header_to_col["Team4 채점 근거"],
                    f"[STD-S 정답표 수정 재채점]\n{judge['rationale']}\n채점방식: {judge['method']}\n\n[이전 채점 근거]\n{current}",
                )
            rescore_rows.append(
                {
                    "problem_id": pid,
                    "question": question,
                    "team": team_name,
                    "old_expected": old_expected,
                    "revised_expected": REVISED_STDS_EXPECTED,
                    "answer": answer,
                    "score": judge["score"],
                    "rationale": judge["rationale"],
                    "method": judge["method"],
                }
            )
            print(f"RESCORE {pid} {team_name} score={judge['score']} method={judge['method']}")

    summary = recompute_summary(wb)

    if "정답표 수정 설명" in wb.sheetnames:
        del wb["정답표 수정 설명"]
    ws = wb.create_sheet("정답표 수정 설명")
    rows = [
        ["항목", "내용"],
        ["평가 유형", "추가 평가. 기존 4팀_정확도_비교.xlsx와 기존 results는 수정하지 않음"],
        ["수정 대상", "문항별 비교 상세 탭의 STD-S-* 예상 답변"],
        ["수정 이유", "Snowflake 관련 문항은 현재 평가 문서셋에 직접 근거가 없으므로 '관련 없음/문서 근거 없음'이 정답이어야 함"],
        ["새 정답 기준", REVISED_STDS_EXPECTED],
        ["산출 파일", str(OUT_XLSX)],
        ["원자료 JSON", str(OUT_JSON)],
        ["생성일", datetime.now().isoformat(timespec="seconds")],
    ]
    for row in rows:
        ws.append(row)
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 100
    style_sheet(ws)

    wb.save(OUT_XLSX)

    output = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "base_xlsx": str(BASE_XLSX),
        "output_xlsx": str(OUT_XLSX),
        "revised_expected": REVISED_STDS_EXPECTED,
        "summary": summary,
        "rescore_rows": rescore_rows,
    }
    OUT_JSON.write_text(json.dumps(output, ensure_ascii=False, indent=2), encoding="utf-8")

    md = f"""# STD-S 정답표 수정 추가 평가 보고서

생성일: {output['generated_at']}

## 목적

기존 `4팀_정확도_비교.xlsx`는 수정하지 않고, 동일 파일을 기준으로 `문항별 비교 상세`의 `STD-S-*` 예상 답변만 수정해 추가 재평가했다.

## 수정 기준

{REVISED_STDS_EXPECTED}

## 산출물

- 추가 평가 엑셀: `{OUT_XLSX}`
- 추가 평가 JSON: `{OUT_JSON}`

## 재산정 결과

| 팀 | 전체 정확도 |
|---|---:|
| Team0 | {summary['overall']['Team0']:.2f}% |
| Team1 | {summary['overall']['Team1']:.2f}% |
| Team2 | {summary['overall']['Team2']:.2f}% |
| Team4 (ont_platform v4) | {summary['overall']['Team4']:.2f}% |

## 카테고리별 결과

| 카테고리 | Team0 | Team1 | Team2 | Team4 |
|---|---:|---:|---:|---:|
"""
    for cat, scores in summary["by_category"].items():
        md += f"| {cat} | {scores['Team0']:.2f}% | {scores['Team1']:.2f}% | {scores['Team2']:.2f}% | {scores['Team4']:.2f}% |\n"
    md += """
## 주의

이 평가는 추가 평가본이다. 기존 `reports\\4팀_정확도_비교.xlsx`와 기존 `results\\same24\\same24_team4_results.json`은 수정하지 않았다.
"""
    OUT_MD.write_text(md, encoding="utf-8")

    print(f"WROTE {OUT_XLSX}")
    print(f"WROTE {OUT_JSON}")
    print(f"WROTE {OUT_MD}")


if __name__ == "__main__":
    main()
