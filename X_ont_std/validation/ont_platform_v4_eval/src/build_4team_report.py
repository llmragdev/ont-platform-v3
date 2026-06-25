from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = next(p for p in Path(__file__).resolve().parents if (p / "ont_platform").exists())
EVAL_DIR = ROOT / "validation" / "ont_platform_v4_eval"
REPORTS_DIR = EVAL_DIR / "reports"
SOURCE_XLSX = EVAL_DIR / "data" / "3팀_정확도_비교.xlsx"
V4_RESULTS_DIR = EVAL_DIR / "results" / "previous30"
OUTPUT_XLSX = REPORTS_DIR / "4팀_정확도_비교_ont_platform_v4.xlsx"
OUTPUT_ASCII_XLSX = REPORTS_DIR / "4team_accuracy_comparison_ont_platform_v4.xlsx"


def load_source():
    wb = load_workbook(SOURCE_XLSX, data_only=True)
    return wb.worksheets[0], wb.worksheets[2], wb.worksheets[3]


def table(ws, sr, sc, headers, rows, name=None):
    blue = "5B9BD5"
    white = "FFFFFF"
    side = Side(style="thin", color="D9D9D9")
    border = Border(left=side, right=side, top=side, bottom=side)

    for c, header in enumerate(headers, start=sc):
        cell = ws.cell(sr, c, header)
        cell.font = Font(bold=True, color=white)
        cell.fill = PatternFill("solid", fgColor=blue)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for r_idx, row in enumerate(rows, start=sr + 1):
        for c_idx, value in enumerate(row, start=sc):
            cell = ws.cell(r_idx, c_idx, value)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if isinstance(value, (int, float)):
                cell.number_format = "0.00"

    if name and rows:
        end_row = sr + len(rows)
        end_col = sc + len(headers) - 1
        ref = f"{get_column_letter(sc)}{sr}:{get_column_letter(end_col)}{end_row}"
        tab = Table(displayName=name, ref=ref)
        tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        ws.add_table(tab)


def title(ws, text, subtitle=None):
    ws.sheet_view.showGridLines = False
    ws["A1"] = text
    ws["A1"].font = Font(size=18, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    ws["A1"].alignment = Alignment(vertical="center")
    ws.merge_cells("A1:H1")
    ws.row_dimensions[1].height = 30
    if subtitle:
        ws["A2"] = subtitle
        ws["A2"].font = Font(size=10, color="666666")
        ws.merge_cells("A2:H2")


def main() -> None:
    ws_item, ws_cat, ws_team = load_source()
    v4_acc = json.loads((V4_RESULTS_DIR / "accuracy_report.json").read_text(encoding="utf-8"))
    v4_perf = json.loads((V4_RESULTS_DIR / "performance_report.json").read_text(encoding="utf-8"))
    v4_results = json.loads((V4_RESULTS_DIR / "test_results.json").read_text(encoding="utf-8"))

    team_rows = []
    for row in ws_team.iter_rows(min_row=2, values_only=True):
        if row[0]:
            team_rows.append([row[0], float(row[1]), row[2], row[3], "3팀_정확도_비교.xlsx"])
    team_rows.append([
        "Team3 (ont_platform v4)",
        round(v4_acc["overall_accuracy"] * 100, 2),
        "",
        "도입 우선 검토",
        "ont_platform v4 직접 실행",
    ])
    team_rows = sorted(team_rows, key=lambda x: x[1], reverse=True)
    for idx, row in enumerate(team_rows, start=1):
        row[2] = f"{idx}위"

    source_cat = []
    for row in ws_cat.iter_rows(min_row=2, values_only=True):
        if row[0] and row[0] != "전체 평균":
            source_cat.append([row[0], float(row[1]), float(row[2]), float(row[3]), float(row[4])])

    cat_rows = [
        [
            "Ontology",
            "Ontology",
            source_cat[0][1],
            source_cat[0][2],
            source_cat[0][3],
            source_cat[0][4],
            round(v4_acc["by_category"]["ontology"] * 100, 2),
            "문항 주제 유사. 기존은 8문항, v4는 12문항",
        ],
        [
            "Advanced RAG",
            "NLP & 생성형 AI",
            source_cat[1][1],
            source_cat[1][2],
            source_cat[1][3],
            source_cat[1][4],
            round(v4_acc["by_category"]["nlp"] * 100, 2),
            "직접 동등 비교 아님. v4 NLP 12문항과 참고 비교",
        ],
        [
            "Snowflake RAG",
            "국방 & 지식통합",
            source_cat[2][1],
            source_cat[2][2],
            source_cat[2][3],
            source_cat[2][4],
            round(v4_acc["by_category"]["defense"] * 100, 2),
            "직접 동등 비교 아님. v4 국방 6문항과 참고 비교",
        ],
    ]

    keyword = [r["evaluation"]["keyword_presence_ratio"] for r in v4_results]
    completeness = [r["evaluation"]["answer_completeness"] for r in v4_results]
    relevance = [r["evaluation"]["answer_relevance"] for r in v4_results]
    sensitivity = []
    for assumed in [1.0, 0.8, 0.6, 0.4, 0.3]:
        adjusted = [
            r["evaluation"]["keyword_presence_ratio"] * 0.4
            + r["evaluation"]["answer_completeness"] * 0.3
            + assumed * 0.3
            for r in v4_results
        ]
        sensitivity.append([assumed, round(sum(adjusted) / len(adjusted) * 100, 2)])

    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("요약")
    title(ws, "4팀 정확도 비교 결과보고서", f"생성일: {datetime.now():%Y-%m-%d %H:%M:%S} / 원본 3팀 엑셀 + ont_platform v4 실측")
    table(
        ws,
        4,
        1,
        ["항목", "내용"],
        [
            ["최종 판정", "ont_platform v4가 전체 평균 76.30%로 4개 팀 중 1위"],
            ["주의", "기존 3팀 엑셀은 24문항, ont_platform v4는 30문항이라 완전 동일 문항 비교는 아님"],
            ["높게 나온 주된 이유", "LLM 30/30 사용, 벡터 hit 150개, 자동 산식에서 검색 연관성 1.0이 전 문항에 적용"],
            ["검증 필요", "동일 24문항으로 ont_platform v4를 재평가하면 더 공정한 4팀 순위가 됨"],
        ],
        "SummaryNotes",
    )
    table(
        ws,
        10,
        1,
        ["KPI", "값"],
        [
            ["ont_platform v4 정확도", round(v4_acc["overall_accuracy"] * 100, 2)],
            ["Team1 기존 최고점", 69.17],
            ["Team1 대비 차이", round(v4_acc["overall_accuracy"] * 100 - 69.17, 2)],
            ["Team0 대비 차이", round(v4_acc["overall_accuracy"] * 100 - 58.54, 2)],
            ["API 성공률", round(v4_perf["success_rate"] * 100, 1)],
            ["평균 응답시간(ms)", v4_perf["avg_response_time_ms"]],
        ],
        "KpiTable",
    )

    ws2 = wb.create_sheet("4팀_종합비교")
    title(ws2, "4팀 종합 정확도 비교", "Team3은 ont_platform v4 직접 실행 결과입니다.")
    table(ws2, 4, 1, ["팀", "정확도 (%)", "순위", "권고", "출처"], team_rows, "OverallComparison")
    chart = BarChart()
    chart.type = "bar"
    chart.title = "팀별 전체 정확도 (%)"
    chart.y_axis.title = "팀"
    chart.x_axis.title = "정확도 (%)"
    chart.add_data(Reference(ws2, min_col=2, min_row=4, max_row=4 + len(team_rows)), titles_from_data=True)
    chart.set_categories(Reference(ws2, min_col=1, min_row=5, max_row=4 + len(team_rows)))
    chart.height = 7
    chart.width = 14
    ws2.add_chart(chart, "G4")

    ws3 = wb.create_sheet("카테고리_참고비교")
    title(ws3, "카테고리별 참고 비교", "카테고리와 문항 수가 달라 동등 비교가 아니라 방향성 참고입니다.")
    table(
        ws3,
        4,
        1,
        ["3팀 카테고리", "v4 대응 카테고리", "Team0 (%)", "Team1 (%)", "Team2 (%)", "3팀 평균 (%)", "ont_platform v4 (%)", "비교 메모"],
        cat_rows,
        "CategoryComparison",
    )
    chart2 = BarChart()
    chart2.type = "col"
    chart2.title = "카테고리별 참고 정확도"
    chart2.y_axis.title = "정확도 (%)"
    chart2.add_data(Reference(ws3, min_col=3, max_col=7, min_row=4, max_row=7), titles_from_data=True)
    chart2.set_categories(Reference(ws3, min_col=1, min_row=5, max_row=7))
    chart2.height = 8
    chart2.width = 18
    ws3.add_chart(chart2, "J4")

    ws4 = wb.create_sheet("v4_상세")
    title(ws4, "ont_platform v4 30문항 상세", "자동 평가 산식: 키워드 40% + 답변완성도 30% + 검색연관성 30%")
    detail_rows = []
    for idx, r in enumerate(v4_results, start=1):
        ev = r["evaluation"]
        detail_rows.append([
            idx,
            r["category"],
            r["query"],
            r["status"],
            ev["keyword_presence_ratio"],
            ev["answer_completeness"],
            ev["answer_relevance"],
            ev["accuracy_score"] * 100,
            r["elapsed_ms"],
            r["quality_metrics"].get("llm_used"),
            r["quality_metrics"].get("vector_hits"),
        ])
    table(
        ws4,
        4,
        1,
        ["No", "카테고리", "질문", "상태", "키워드비율", "완성도", "연관성", "정확도 (%)", "응답시간(ms)", "LLM 사용", "Vector Hits"],
        detail_rows,
        "V4Detail",
    )

    ws5 = wb.create_sheet("산식_재검토")
    title(ws5, "76.30% 산식 재검토", "높은 점수의 원인을 구성요소와 민감도로 확인합니다.")
    table(
        ws5,
        4,
        1,
        ["구성요소", "값 (%)", "해석"],
        [
            ["평균 키워드 포함율", round(sum(keyword) / len(keyword) * 100, 2), "답변 내용에 골든 키워드가 포함된 비율"],
            ["평균 답변 완성도", round(sum(completeness) / len(completeness) * 100, 2), "답변 길이 기반 점수"],
            ["평균 검색 연관성", round(sum(relevance) / len(relevance) * 100, 2), "거리 점수를 정규화했으나 전 문항 1.0으로 산출"],
            ["최종 정확도", round(v4_acc["overall_accuracy"] * 100, 2), "0.4*키워드 + 0.3*완성도 + 0.3*연관성"],
        ],
        "FormulaComponents",
    )
    table(ws5, 11, 1, ["가정 검색연관성", "재계산 정확도 (%)"], sensitivity, "Sensitivity")
    notes = [
        "연관성 1.0이면 현재 76.30%",
        "연관성 0.8이면 Team1과 비슷한 70.30%",
        "연관성 0.6이면 64.30%로 Team2 수준",
        "연관성 0.4이면 58.30%로 Team0 수준",
        "따라서 76.30%는 검색연관성 산식에 민감함",
    ]
    ws5["D11"] = "해석"
    ws5["D11"].font = Font(bold=True, color="FFFFFF")
    ws5["D11"].fill = PatternFill("solid", fgColor="5B9BD5")
    for idx, note in enumerate(notes, start=12):
        ws5.cell(idx, 4, note)

    ws6 = wb.create_sheet("3팀_원본_문항요약")
    title(ws6, "3팀 원본 문항별 비교", "원본 엑셀의 문항별 비교 시트를 값으로 복사했습니다.")
    source_rows = list(ws_item.iter_rows(min_row=2, values_only=True))
    table(ws6, 4, 1, ["문제ID", "카테고리", "질문", "Team0 (%)", "Team1 (%)", "Team2 (%)", "평균 (%)"], source_rows, "OriginalItemSummary")

    for sheet in wb.worksheets:
        sheet.freeze_panes = "A4"
        for col in range(1, min(sheet.max_column, 12) + 1):
            letter = get_column_letter(col)
            max_len = 0
            for cell in sheet[letter]:
                value = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, min(len(value), 60))
            sheet.column_dimensions[letter].width = max(10, min(max_len + 2, 45))
        for row in range(1, sheet.max_row + 1):
            sheet.row_dimensions[row].height = 22

    ws3.column_dimensions["H"].width = 45
    ws4.column_dimensions["C"].width = 45
    ws5.column_dimensions["C"].width = 45
    ws5.column_dimensions["D"].width = 45

    for row in range(5, 5 + len(team_rows)):
        if ws2.cell(row, 1).value == "Team3 (ont_platform v4)":
            for col in range(1, 6):
                ws2.cell(row, col).fill = PatternFill("solid", fgColor="E2F0D9")
                ws2.cell(row, col).font = Font(bold=True)

    wb.save(OUTPUT_XLSX)
    wb.save(OUTPUT_ASCII_XLSX)
    print(OUTPUT_XLSX)


if __name__ == "__main__":
    main()
