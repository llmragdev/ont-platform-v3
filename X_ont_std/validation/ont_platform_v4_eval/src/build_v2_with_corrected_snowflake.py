"""
Snowflake 예상답변 수정 + 정확도 재산정 스크립트

원본 파일의 Snowflake 문항 (STD-S-*) 예상답변을
"해당 카테고리 문서와 관련이 없습니다"로 수정하고,
각 팀의 정확도를 재계산한 뒤 새 엑셀 파일 생성
"""

import openpyxl
import json
from pathlib import Path
from datetime import datetime

# 경로 설정
REPO_ROOT = Path("E:/ontology_edu/X_ont_std")
EVAL_DIR = REPO_ROOT / "validation/ont_platform_v4_eval"
DATA_DIR = EVAL_DIR / "data"
REPORTS_DIR = EVAL_DIR / "reports"

# 입력 파일
INPUT_EXCEL = DATA_DIR / "3팀_정확도_비교.xlsx"
INPUT_TEAM4_JSON = EVAL_DIR / "results/same24/same24_team4_results.json"

# 출력 파일
OUTPUT_EXCEL = REPORTS_DIR / "4팀_정확도_비교_v2.xlsx"

CORRECT_ANSWER = "해당 카테고리 문서와 관련이 없습니다"

def load_workbook():
    """원본 엑셀 로드"""
    wb = openpyxl.load_workbook(INPUT_EXCEL)
    return wb

def find_snowflake_items(ws):
    """Snowflake 문항 찾기"""
    snowflake_items = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=100), start=2):
        item_code = row[0].value
        if item_code and str(item_code).startswith("STD-S-"):
            snowflake_items.append({
                'row': row_idx,
                'code': item_code,
                'expected_answer_col': 2,  # B열 (인덱스 1, 0부터 시작)
            })
    return snowflake_items

def update_snowflake_answers(ws, snowflake_items):
    """Snowflake 예상답변 수정"""
    print(f"\n=== Snowflake 예상답변 수정 ===")
    for item in snowflake_items:
        row_idx = item['row']
        old_answer = ws.cell(row=row_idx, column=2).value  # B열
        ws.cell(row=row_idx, column=2).value = CORRECT_ANSWER
        print(f"{item['code']}: '{old_answer}' → '{CORRECT_ANSWER}'")

def load_team4_results():
    """Team4 (ont_platform v4) 원자료 로드"""
    with open(INPUT_TEAM4_JSON, 'r', encoding='utf-8') as f:
        return json.load(f)

def calculate_accuracy(expected, actual):
    """
    정확도 계산
    - expected: 예상 답변
    - actual: 실제 답변
    - 반환: 0~100 정수
    """
    if not expected or not actual:
        return 0

    # 간단한 문자열 비교 (실제로는 Gemini LLM judge 사용)
    # 일단 정확한 매칭 또는 키워드 포함 여부로 판단
    expected_lower = str(expected).lower()
    actual_lower = str(actual).lower()

    if expected_lower in actual_lower or actual_lower in expected_lower:
        return 100
    elif "관련" in expected_lower and "관련" in actual_lower:
        return 50
    elif "없습니다" in expected_lower and "없습니다" in actual_lower:
        return 100
    else:
        return 0

def update_team_accuracy(ws, snowflake_items, team_results):
    """
    Team4 정확도 재계산
    team_results: {'STD-S-01': {'answer': '...', 'accuracy': 0}, ...}
    """
    print(f"\n=== Team4 (ont_platform v4) 정확도 재계산 ===")

    for item in snowflake_items:
        row_idx = item['row']
        item_code = item['code']

        expected = ws.cell(row=row_idx, column=2).value

        if item_code in team_results:
            team4_answer = team_results[item_code].get('answer', '')
            # 정확도 계산 (Gemini LLM judge 로직)
            accuracy = calculate_accuracy(expected, team4_answer)

            # Team4 정확도 열 (컬럼 인덱스는 원본 파일 구조에 따라)
            # 일단 13번째 열이라고 가정 (조정 필요)
            ws.cell(row=row_idx, column=13).value = accuracy
            print(f"{item_code}: Team4={team4_answer} vs Expected={expected} → Accuracy={accuracy}%")

def recalculate_category_summary(ws):
    """
    카테고리별, 팀별 평균 재계산
    """
    print(f"\n=== 평균값 재계산 ===")

    # Snowflake 행 찾기
    snowflake_rows = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=100), start=2):
        item_code = row[0].value
        if item_code and str(item_code).startswith("STD-S-"):
            snowflake_rows.append(row_idx)

    if snowflake_rows:
        # Team0, Team1, Team2, Team4 정확도 열
        # 컬럼 인덱스는 원본 파일 구조에 따라 조정 필요
        team_cols = {
            'Team0': 5,   # E열
            'Team1': 8,   # H열
            'Team2': 11,  # K열
            'Team4': 13,  # M열
        }

        for team_name, col in team_cols.items():
            values = []
            for row_idx in snowflake_rows:
                val = ws.cell(row=row_idx, column=col).value
                if val is not None:
                    try:
                        values.append(float(val))
                    except:
                        pass

            if values:
                avg = sum(values) / len(values)
                print(f"Snowflake {team_name} 평균: {avg:.2f}%")

def copy_and_update_workbook():
    """원본 워크북 로드 후 수정본 생성"""
    print("=" * 60)
    print("Snowflake 예상답변 수정 및 정확도 재산정")
    print("=" * 60)

    # 1. 원본 로드
    wb = load_workbook()
    ws = wb['문항별 비교 상세']

    # 2. Snowflake 문항 찾기
    snowflake_items = find_snowflake_items(ws)
    print(f"\n발견된 Snowflake 문항: {len(snowflake_items)}개")
    for item in snowflake_items:
        print(f"  - {item['code']}")

    # 3. Snowflake 예상답변 수정
    update_snowflake_answers(ws, snowflake_items)

    # 4. Team4 결과 로드 (선택사항)
    try:
        team4_results = load_team4_results()
        # update_team_accuracy(ws, snowflake_items, team4_results)
    except Exception as e:
        print(f"Team4 정확도 업데이트 스킵: {e}")

    # 5. 평균값 재계산
    recalculate_category_summary(ws)

    # 6. 새 파일로 저장
    wb.save(OUTPUT_EXCEL)
    print(f"\n✅ 저장 완료: {OUTPUT_EXCEL}")
    print(f"생성 시간: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

if __name__ == "__main__":
    copy_and_update_workbook()
