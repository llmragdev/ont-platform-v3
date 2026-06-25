from __future__ import annotations

import json
import os
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


ROOT = next(p for p in Path(__file__).resolve().parents if (p / "ont_platform").exists())
EVAL_DIR = ROOT / "validation" / "ont_platform_v4_eval"
V4_BACKEND = ROOT / "ont_platform" / "v4" / "backend"
V3_ENV = ROOT / "ont_platform" / "v3" / ".env"

BASE_XLSX = EVAL_DIR / "reports" / "4팀_정확도_비교.xlsx"
OUT_XLSX = EVAL_DIR / "reports" / "4팀_정확도_비교_STD-S카테고리무관_추가평가.xlsx"
OUT_MD = EVAL_DIR / "reports" / "STD-S카테고리무관_추가평가_보고서.md"
OUT_JSON = EVAL_DIR / "results" / "revised_stds_category_irrelevant" / "stds_category_irrelevant_rescore.json"

REVISED_EXPECTED = "질문은 해당 카테고리 문서와 관련이 없습니다."


def load_env_file(path: Path) -> None:
    if not path.exists():
        return
    for raw in path.read_text(encoding="utf-8").splitlines():
        line = raw.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        if " #" in value:
            value = value.split(" #", 1)[0].strip()
        if "  " in value:
            value = value.split("  ", 1)[0].strip()
        if key and value and key not in os.environ:
            os.environ[key] = value

    default_key = os.environ.get("GEMINI_API_KEY", "")
    if len(default_key) > 45:
        for name in ("GEMINI_API_KEY1", "GEMINI_API_KEY2", "GEMINI_API_KEY3", "GEMINI_API_KEY4"):
            candidate = os.environ.get(name, "")
            if candidate.startswith("AIza") and len(candidate) <= 45:
                os.environ["GEMINI_API_KEY"] = candidate
                break


def extract_json(text: str) -> dict[str, Any] | None:
    cleaned = text.strip()
    if cleaned.startswith("```"):
        parts = cleaned.split("```")
        cleaned = parts[1] if len(parts) > 1 else cleaned
        if cleaned.lower().startswith("json"):
            cleaned = cleaned[4:]
    match = re.search(r"\{.*\}", cleaned, flags=re.S)
    if match:
        cleaned = match.group(0)
    try:
        return json.loads(cleaned)
    except Exception:
        return None


def judge_with_fallback(answer: str) -> tuple[float, str]:
    text = answer.strip()
    says_irrelevant = any(term in text for term in ["관련이 없습니다", "관련 없음", "관련 없는", "해당 카테고리"])
    says_not_found = any(term in text for term in ["찾지 못", "확인할 수 없", "정보가 없", "직접적인 정보", "직접적", "근거가 없"])
    hallucinates = any(term in text for term in ["Snowflake", "스노우플레이크", "ranking_issue", "정형 테이블", "PDF 문서", "QA 테스트"])
    long_general = len(text) >= 180

    if says_irrelevant and not long_general:
        return 100.0, "fallback: 해당 카테고리 문서와 관련 없다는 취지가 명확함"
    if says_irrelevant:
        return 90.0, "fallback: 카테고리 무관 취지는 있으나 부가 설명이 있음"
    if says_not_found and not hallucinates:
        return 80.0, "fallback: 직접 정보 없음/확인 불가 취지"
    if says_not_found:
        return 65.0, "fallback: 정보 없음 취지는 있으나 일반론이 섞임"
    if hallucinates or long_general:
        return 10.0, "fallback: 카테고리 무관 대신 일반론을 생성"
    return 30.0, "fallback: 카테고리 무관 취지가 불명확"


def judge_answer(llm, question: str, answer: str) -> dict[str, Any]:
    prompt = f"""
당신은 STD-S 문항 정답표를 엄격히 수정해 재채점하는 평가자입니다.

새 정답:
{REVISED_EXPECTED}

평가 기준:
- 답변이 "질문은 해당 카테고리 문서와 관련이 없습니다" 또는 같은 의미를 명확히 말하면 95~100점입니다.
- "제공 문서/사내 지식 베이스에서 직접 정보를 찾지 못했다"처럼 근거 없음만 말하면 70~85점입니다.
- 근거 없음이라고 말하면서도 Snowflake/RAG 일반론을 길게 덧붙이면 50~70점입니다.
- 관련 없음이라고 하지 않고 Snowflake, RAG, ranking_issue, QA 테스트 관련 정답을 만들어내면 0~20점입니다.
- 정답처럼 보이는 일반론일수록 더 낮게 평가합니다. 이 문항들의 핵심은 답을 생성하지 않는 것입니다.

반드시 JSON만 반환하세요.
{{"score": 0에서 100 사이 숫자, "rationale": "짧은 채점 근거"}}

[질문]
{question}

[답변]
{answer}
""".strip()
    text = llm.generate(prompt, temperature=0.0, max_tokens=384) if llm.enabled else None
    if text:
        parsed = extract_json(text)
        if parsed and "score" in parsed:
            return {
                "score": round(max(0.0, min(100.0, float(parsed["score"]))), 1),
                "rationale": str(parsed.get("rationale", "")).strip(),
                "method": "Gemini LLM judge",
                "raw": text,
            }
    score, rationale = judge_with_fallback(answer)
    return {"score": score, "rationale": rationale, "method": "fallback heuristic", "raw": text or ""}


def style_sheet(ws) -> None:
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill = PatternFill("solid", fgColor="1F4E78")
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if cell.row == 1:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if isinstance(cell.value, float):
                cell.number_format = "0.00"


def recompute(wb) -> dict[str, Any]:
    detail = wb["문항별 비교 상세"]
    item = wb["문항별 비교"]
    category = wb["카테고리별 요약"]
    team = wb["팀별 요약"]

    cat_by_id = {str(r[0]): str(r[1]) for r in item.iter_rows(min_row=2, values_only=True) if r[0]}
    rows = []
    for idx in range(2, detail.max_row + 1):
        pid = str(detail.cell(idx, 1).value or "")
        if not pid:
            continue
        scores = {
            "Team0": float(detail.cell(idx, 5).value or 0),
            "Team1": float(detail.cell(idx, 7).value or 0),
            "Team2": float(detail.cell(idx, 9).value or 0),
            "Team4": float(detail.cell(idx, 11).value or 0),
        }
        rows.append({"id": pid, "category": cat_by_id.get(pid, ""), "scores": scores})

        item.cell(idx, 4, scores["Team0"])
        item.cell(idx, 5, scores["Team1"])
        item.cell(idx, 6, scores["Team2"])
        item.cell(idx, 7, scores["Team4"])
        item.cell(idx, 8, round(sum(scores.values()) / 4, 2))

    by_category = {}
    for cat in sorted({r["category"] for r in rows}):
        cat_rows = [r for r in rows if r["category"] == cat]
        by_category[cat] = {
            key: round(sum(r["scores"][key] for r in cat_rows) / len(cat_rows), 2)
            for key in ("Team0", "Team1", "Team2", "Team4")
        }
    overall = {
        key: round(sum(r["scores"][key] for r in rows) / len(rows), 2)
        for key in ("Team0", "Team1", "Team2", "Team4")
    }

    for row_idx in range(2, category.max_row + 1):
        cat = category.cell(row_idx, 1).value
        if not cat:
            continue
        scores = overall if cat == "전체 평균" else by_category.get(str(cat))
        if not scores:
            continue
        values = [scores["Team0"], scores["Team1"], scores["Team2"], scores["Team4"]]
        for col, value in enumerate(values, start=2):
            category.cell(row_idx, col, value)
        category.cell(row_idx, 6, round(sum(values) / 4, 2))

    ranks = sorted(overall.items(), key=lambda item: item[1], reverse=True)
    rank_by_team = {name: idx for idx, (name, _) in enumerate(ranks, start=1)}
    labels = {"Team0": "Team0", "Team1": "Team1", "Team2": "Team2", "Team4": "Team4 (ont_platform v4)"}
    for idx, key in enumerate(("Team0", "Team1", "Team2", "Team4"), start=2):
        rank = rank_by_team[key]
        team.cell(idx, 1, labels[key])
        team.cell(idx, 2, overall[key])
        team.cell(idx, 3, f"{rank}위")
        team.cell(idx, 4, "도입 우선 검토" if rank == 1 else "도입 고려" if rank == 2 else "개선 필요")

    for ws in wb.worksheets:
        style_sheet(ws)
    return {"overall": overall, "by_category": by_category, "ranks": ranks}


def main() -> None:
    OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    load_env_file(V3_ENV)
    sys.path.insert(0, str(V4_BACKEND))
    from app.services.llm_client import LlmClient

    llm = LlmClient(model=os.getenv("LLM_MODEL_NAME") or "gemini-2.5-flash-lite")
    wb = load_workbook(BASE_XLSX)
    detail = wb["문항별 비교 상세"]

    headers = [detail.cell(1, col).value for col in range(1, detail.max_column + 1)]
    col = {name: idx for idx, name in enumerate(headers, start=1)}
    answer_cols = {
        "Team0": col["Team0 답변"],
        "Team1": col["Team1 답변"],
        "Team2": col["Team2 답변"],
        "Team4": col["Team4 답변 (ont_platform v4)"],
    }
    score_cols = {
        "Team0": col["Team0 정확도 (%)"],
        "Team1": col["Team1 정확도 (%)"],
        "Team2": col["Team2 정확도 (%)"],
        "Team4": col["Team4 정확도 (%)"],
    }

    rows = []
    for row_idx in range(2, detail.max_row + 1):
        pid = str(detail.cell(row_idx, 1).value or "")
        if not pid.startswith("STD-S"):
            continue
        question = str(detail.cell(row_idx, 2).value or "")
        old_expected = str(detail.cell(row_idx, 3).value or "")
        detail.cell(row_idx, 3, REVISED_EXPECTED)
        for team_name in ("Team0", "Team1", "Team2", "Team4"):
            answer = str(detail.cell(row_idx, answer_cols[team_name]).value or "")
            result = judge_answer(llm, question, answer)
            detail.cell(row_idx, score_cols[team_name], result["score"])
            if team_name == "Team4":
                previous = str(detail.cell(row_idx, col["Team4 채점 근거"]).value or "")
                detail.cell(
                    row_idx,
                    col["Team4 채점 근거"],
                    f"[STD-S 카테고리 무관 정답표 재채점]\n{result['rationale']}\n채점방식: {result['method']}\n\n[이전 채점 근거]\n{previous}",
                )
            rows.append({
                "problem_id": pid,
                "team": team_name,
                "question": question,
                "old_expected": old_expected,
                "revised_expected": REVISED_EXPECTED,
                "answer": answer,
                **result,
            })
            print(f"RESCORE {pid} {team_name} score={result['score']} method={result['method']}")

    if "정답표 수정 설명" in wb.sheetnames:
        del wb["정답표 수정 설명"]
    ws = wb.create_sheet("정답표 수정 설명")
    for row in [
        ["항목", "내용"],
        ["평가 유형", "추가 평가. 기존 결과 파일은 수정하지 않음"],
        ["수정 대상", "문항별 비교 상세 탭의 STD-S-* 예상 답변"],
        ["새 정답", REVISED_EXPECTED],
        ["수정 이유", "Snowflake 문항은 해당 카테고리 문서와 관련 없는 질문이므로 답을 생성하지 않는 것이 정답"],
        ["출력 엑셀", str(OUT_XLSX)],
        ["출력 JSON", str(OUT_JSON)],
        ["생성일", datetime.now().isoformat(timespec="seconds")],
    ]:
        ws.append(row)
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 100

    summary = recompute(wb)
    wb.save(OUT_XLSX)

    output = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "base_xlsx": str(BASE_XLSX),
        "output_xlsx": str(OUT_XLSX),
        "revised_expected": REVISED_EXPECTED,
        "summary": summary,
        "rescore_rows": rows,
    }
    OUT_JSON.write_text(json.dumps(output, ensure_ascii=False, indent=2), encoding="utf-8")

    md = f"""# STD-S 카테고리 무관 정답표 추가 평가 보고서

생성일: {output['generated_at']}

## 새 정답

`{REVISED_EXPECTED}`

## 산출물

- 엑셀: `{OUT_XLSX}`
- JSON: `{OUT_JSON}`

## 전체 정확도

| 팀 | 전체 정확도 |
|---|---:|
| Team0 | {summary['overall']['Team0']:.2f}% |
| Team1 | {summary['overall']['Team1']:.2f}% |
| Team2 | {summary['overall']['Team2']:.2f}% |
| Team4 (ont_platform v4) | {summary['overall']['Team4']:.2f}% |

## 카테고리별 정확도

| 카테고리 | Team0 | Team1 | Team2 | Team4 |
|---|---:|---:|---:|---:|
"""
    for cat, scores in summary["by_category"].items():
        md += f"| {cat} | {scores['Team0']:.2f}% | {scores['Team1']:.2f}% | {scores['Team2']:.2f}% | {scores['Team4']:.2f}% |\n"
    md += "\n기존 파일은 수정하지 않은 추가 평가본이다.\n"
    OUT_MD.write_text(md, encoding="utf-8")

    print(f"WROTE {OUT_XLSX}")
    print(f"WROTE {OUT_JSON}")
    print(f"WROTE {OUT_MD}")


if __name__ == "__main__":
    main()
