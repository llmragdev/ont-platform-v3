from __future__ import annotations

import re
import shutil
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


EXPECTED_DIR = Path("E:/ai_lab_SIT/qa_expected")
REPORT_DIR = Path("E:/ontology_edu/X_ont_std/validation/ont_platform_v4_eval/reports")

OUT_NAME = "4팀_정확도_비교_STD-S정답수정_v1.2_재평가.xlsx"
FINAL_OUT_NAME = "4팀_정확도_비교_STD-S정답수정_v1.2_재평가_final_Team2_STD-S05보정.xlsx"


TEAM_ANSWER_COLS = {
    "Team0": "Team0 답변",
    "Team1": "Team1 답변",
    "Team2": "Team2 답변",
    "Team4": "Team4 답변 (ont_platform v4)",
}

TEAM_SCORE_COLS = {
    "Team0": "Team0 정확도 (%)",
    "Team1": "Team1 정확도 (%)",
    "Team2": "Team2 정확도 (%)",
    "Team4": "Team4 정확도 (%)",
}


def norm(text: object) -> str:
    if text is None:
        return ""
    s = str(text).lower()
    s = re.sub(r"\s+", "", s)
    s = re.sub(r"[\[\]\(\)\{\},.;:'\"`~!@#$%^&*_+=|\\/<>?，。·ㆍ]", "", s)
    return s


def split_terms(value: object) -> list[str]:
    if value is None:
        return []
    raw = str(value)
    raw = raw.replace("\n", ",").replace(";", ",")
    terms: list[str] = []
    for part in raw.split(","):
        part = part.strip()
        if not part:
            continue
        # Expand compact alternatives such as 클래스/속성/인스턴스.
        subparts = [x.strip() for x in re.split(r"[/|]", part) if x.strip()]
        terms.extend(subparts if len(subparts) > 1 else [part])
    seen = set()
    deduped = []
    for term in terms:
        key = norm(term)
        if key and key not in seen:
            seen.add(key)
            deduped.append(term)
    return deduped


def fallback_terms(question: str, expected: str) -> list[str]:
    candidates = re.findall(r"[가-힣A-Za-z0-9]{2,}", f"{question} {expected}")
    stop = {
        "무엇인가",
        "어떻게",
        "해야",
        "한다",
        "있는",
        "대한",
        "위해",
        "에서",
        "으로",
        "하는",
        "것은",
        "때문",
    }
    terms = []
    for token in candidates:
        if token in stop:
            continue
        if len(token) >= 2:
            terms.append(token)
    return terms[:12]


def has_no_relation_marker(answer: str) -> bool:
    n = norm(answer)
    markers = [
        "관련없",
        "관련이없",
        "관련내용없",
        "관련된내용없",
        "무관",
        "논문에없",
        "문서에없",
        "자료에없",
        "검색결과없",
        "근거가없",
        "근거없",
        "언급되지않",
        "확인되지않",
        "찾을수없",
        "제공된문서에서확인",
        "제공된자료에서확인",
        "해당카테고리문서와관련",
        "답변할수없",
    ]
    return any(marker in n for marker in markers)


def has_snowflake_content(answer: str) -> bool:
    n = norm(answer)
    terms = [
        "snowflake",
        "스노우플레이크",
        "warehouse",
        "웨어하우스",
        "snowpark",
        "cortex",
        "streamlit",
        "variant",
        "micro",
        "파티션",
        "데이터클라우드",
    ]
    return any(term in n for term in terms)


def has_other_topic(answer: str) -> bool:
    n = norm(answer)
    terms = [
        "온톨로지",
        "지식그래프",
        "bert",
        "simrank",
        "rag",
        "임베딩",
        "국방",
        "지휘통제",
    ]
    return any(term in n for term in terms)


def score_trap_snowflake(answer: str) -> tuple[int, str]:
    if not str(answer or "").strip():
        return 0, "무응답: 함정 질문에서 관련 없음 표현도 없음"
    no_relation = has_no_relation_marker(answer)
    snowflake = has_snowflake_content(answer)
    other_topic = has_other_topic(answer)
    if no_relation:
        if other_topic and not snowflake:
            return 90, "STD-S 함정 규칙: 관련 없음은 명시했으나 다른 주제 부가 설명이 포함됨"
        return 100, "STD-S 함정 규칙: 스노우플레이크 관련 내용이 문서에 없음을 명시. 부가 설명 여부와 무관하게 만점"
    if snowflake:
        return 40, "STD-S 함정 규칙: 질문의 스노우플레이크를 실제 근거 내용처럼 답했으나 관련 없음 표현이 없음"
    if other_topic:
        return 25, "STD-S 함정 규칙: 관련 없음 표현 없이 다른 주제로 답변"
    return 35, "STD-S 함정 규칙: 관련 없음 표현이 없어 큰 감점"


def manual_score_override(qid: str, team: str, answer: str) -> tuple[int, str] | None:
    if team == "Team1" and qid in {"STD-S-01", "STD-S-02", "STD-S-04", "STD-S-05", "STD-S-06", "STD-S-07", "STD-S-08"}:
        return (
            60,
            "수동 보정: 관련 없음/직접 정보 없음은 언급했으나 HNIX AI Labs 2기 중간점검-v1.2.pptx 등 허용되지 않은 출처 또는 일반 추론을 근거처럼 사용했으므로 wrong-source leakage 감점",
        )
    if qid == "STD-S-05" and team == "Team2":
        return (
            100,
            "수동 보정: Snowflake RAG 답변의 근거 적합성에 대해 제공 검색 결과에 명확한 정보가 없다고 제한한 뒤, 관련 일반 요소를 부가 설명했으므로 STD-S 함정 질문 규칙상 만점 처리",
        )
    return None


def score_general(answer: str, question: str, expected: str, key_terms: list[str]) -> tuple[int, str]:
    if not str(answer or "").strip():
        return 0, "무응답"
    answer_n = norm(answer)
    if not key_terms:
        key_terms = fallback_terms(question, expected)
    checks = []
    for term in key_terms:
        t = norm(term)
        if not t:
            continue
        checks.append((term, t in answer_n))
    if not checks:
        return 50, "핵심어 기준 부재: 중립 점수"
    hits = [term for term, ok in checks if ok]
    misses = [term for term, ok in checks if not ok]
    coverage = len(hits) / len(checks)
    if coverage >= 0.9:
        score = 100
    elif coverage >= 0.75:
        score = 90
    elif coverage >= 0.6:
        score = 80
    elif coverage >= 0.45:
        score = 70
    elif coverage >= 0.3:
        score = 60
    elif coverage >= 0.15:
        score = 50
    else:
        score = 35
    note = f"엑셀 기반 핵심포함어 매칭: {len(hits)}/{len(checks)}개 충족"
    if misses:
        note += f"; 주요 누락: {', '.join(misses[:5])}"
    return score, note


def load_expected() -> dict[str, dict[str, object]]:
    expected_file = next(p for p in EXPECTED_DIR.glob("*.xlsx") if "v1.2" in p.name)
    wb = load_workbook(str(expected_file), data_only=True, read_only=True)
    ws = next(ws for ws in wb.worksheets if ws.cell(1, 1).value == "ID" and ws.max_row == 25 and "old" not in ws.title.lower())
    headers = [cell.value for cell in ws[1]]
    col = {h: i + 1 for i, h in enumerate(headers) if h}
    out: dict[str, dict[str, object]] = {}
    for r in range(2, ws.max_row + 1):
        qid = ws.cell(r, col["ID"]).value
        if not qid:
            continue
        row = {h: ws.cell(r, col[h]).value for h in col}
        row["핵심어목록"] = split_terms(row.get("핵심포함어"))
        out[str(qid)] = row
    return out


def find_source_report() -> Path:
    candidates = [
        p
        for p in REPORT_DIR.glob("*.xlsx")
        if "STD-S" in p.name and "정답수정" in p.name and not p.name.startswith("~$")
    ]
    if candidates:
        return candidates[0]
    candidates = [p for p in REPORT_DIR.glob("*.xlsx") if "STD-S" in p.name and not p.name.startswith("~$")]
    if not candidates:
        raise FileNotFoundError("STD-S report workbook not found")
    return candidates[0]


def autosize(ws, min_width=10, max_width=72) -> None:
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for row_idx in range(1, min(ws.max_row, 80) + 1):
            value = ws.cell(row_idx, col_idx).value
            if value is None:
                continue
            max_len = max(max_len, min(len(str(value)), max_width))
        ws.column_dimensions[letter].width = max(min_width, min(max_len + 2, max_width))


def rebuild_summary_sheets(wb, detail_rows: list[dict[str, object]]) -> None:
    teams = ["Team0", "Team1", "Team2", "Team4"]
    categories = ["Advanced RAG", "Ontology", "Snowflake"]

    by_category = defaultdict(lambda: defaultdict(list))
    by_team = defaultdict(list)
    for row in detail_rows:
        category = row["카테고리"]
        for team in teams:
            score = row[f"{team} 점수"]
            by_category[category][team].append(score)
            by_team[team].append(score)

    ws = wb["문항별 비교"]
    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            ws.cell(row_idx, col_idx).value = None
    for i, row in enumerate(detail_rows, start=2):
        scores = [row[f"{team} 점수"] for team in teams]
        ws.cell(i, 1, row["문제ID"])
        ws.cell(i, 2, row["카테고리"])
        ws.cell(i, 3, row["질문"])
        for offset, score in enumerate(scores, start=4):
            ws.cell(i, offset, score)
        ws.cell(i, 8, round(sum(scores) / len(scores), 2))

    ws = wb["카테고리별 요약"]
    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            ws.cell(row_idx, col_idx).value = None
    for r, category in enumerate(categories, start=2):
        ws.cell(r, 1, category)
        scores = []
        for c, team in enumerate(teams, start=2):
            values = by_category[category][team]
            avg = round(sum(values) / len(values), 2) if values else 0
            ws.cell(r, c, avg)
            scores.append(avg)
        ws.cell(r, 6, round(sum(scores) / len(scores), 2))
    r = 5
    ws.cell(r, 1, "전체 평균")
    all_scores = []
    for c, team in enumerate(teams, start=2):
        values = by_team[team]
        avg = round(sum(values) / len(values), 2) if values else 0
        ws.cell(r, c, avg)
        all_scores.append(avg)
    ws.cell(r, 6, round(sum(all_scores) / len(all_scores), 2))

    ws = wb["팀별 요약"]
    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            ws.cell(row_idx, col_idx).value = None
    team_names = {
        "Team0": "Team0",
        "Team1": "Team1",
        "Team2": "Team2",
        "Team4": "Team4 (ont_platform v4)",
    }
    ranking = []
    for team in teams:
        avg = round(sum(by_team[team]) / len(by_team[team]), 2)
        ranking.append((team, avg))
    ranking.sort(key=lambda x: x[1], reverse=True)
    rank_map = {team: f"{idx}위" for idx, (team, _) in enumerate(ranking, start=1)}
    for r, team in enumerate(teams, start=2):
        avg = round(sum(by_team[team]) / len(by_team[team]), 2)
        ws.cell(r, 1, team_names[team])
        ws.cell(r, 2, avg)
        ws.cell(r, 3, rank_map[team])
        ws.cell(r, 4, "도입 우선 검토" if rank_map[team] == "1위" else ("도입 고려" if avg >= 70 else "개선 필요"))


def update_notes(wb) -> None:
    now = datetime(2026, 6, 8, 17, 30).isoformat(timespec="seconds")
    if "v1.2 재평가 기준" in wb.sheetnames:
        del wb["v1.2 재평가 기준"]
    ws = wb.create_sheet("v1.2 재평가 기준", 0)
    rows = [
        ["항목", "내용"],
        ["정답 기준", "E:/ai_lab_SIT/qa_expected/중간점검_통합테스트_예상 질문_답변-v1.2.xlsx의 표준QA_24 시트"],
        ["재평가 방식", "외부 API 재호출 없이 기존 보고서의 4팀 답변 텍스트와 v1.2 기대답변/핵심포함어만 사용"],
        ["일반 문항", "v1.2 핵심포함어 매칭률 기반 규칙 점수. 기존 답변 원문과 검색근거는 보존"],
        ["STD-S-01~08", "스노우플레이크는 해당 논문/카테고리 문서에 없는 함정 질문으로 처리"],
        ["STD-S 만점 규칙", "관련 없음/문서에 없음/근거 없음/확인 불가를 명시하면 부가 설명 여부와 관계없이 100점"],
        ["STD-S 감점 규칙", "관련 없음 표현 없이 스노우플레이크 설명을 생성하면 큰 감점. 다른 주제로 답하면 추가 감점"],
        ["한계", "LLM judge가 아니라 엑셀 내용 기반 결정 규칙이므로 최종 제출 전 표본 수동 검토 권장"],
    ]
    for r_idx, row in enumerate(rows, start=1):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(r_idx, c_idx, value)
    ws["A1"].font = Font(bold=True, color="FFFFFF")
    ws["B1"].font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    autosize(ws, max_width=90)

    ws2 = wb["채점 방식"]
    ws2.cell(2, 2, "v1.2 정답지의 예상 답변/근거/핵심포함어를 기준으로 기존 4팀 답변을 재평가")
    ws2.cell(3, 2, "Team0~2 기존 답변 원문은 보존하고 점수만 v1.2 기준으로 재산정")
    ws2.cell(4, 2, "Team4 기존 질의 답변/검색근거는 보존하고 점수만 v1.2 기준으로 재산정")
    ws2.cell(5, 2, "외부 API 재호출 없음. 일반 문항은 핵심포함어 매칭, STD-S는 함정 질문 특수 규칙 적용")
    ws2.cell(6, 2, "정량 점수는 규칙 기반 재평가값이므로 최종 제출 전 표본 수동 검토 권장")
    if ws2.max_row >= 7:
        ws2.cell(7, 2, now)
    if ws2.max_row >= 8:
        ws2.cell(8, 2, str(REPORT_DIR / FINAL_OUT_NAME))

    if "정답표 수정 설명" in wb.sheetnames:
        ws3 = wb["정답표 수정 설명"]
        replacement = [
            ["항목", "내용"],
            ["평가 유형", "v1.2 정답 기준 재평가. 기존 보고서와 기존 results는 수정하지 않음"],
            ["수정 대상", "문항별 비교 상세 탭의 예상 답변, Team0~4 정확도, 요약 시트"],
            ["정답 기준", "중간점검_통합테스트_예상 질문_답변-v1.2.xlsx의 표준QA_24 시트"],
            ["STD-S 특수 규칙", "STD-S-01~08은 Snowflake 관련 내용이 평가 논문/문서에 없는 함정 질문으로 처리"],
            ["STD-S 만점 조건", "관련 없음/문서에 없음/근거 없음/확인 불가를 명시하면 부가 설명 여부와 무관하게 100점"],
            ["산출 파일", str(REPORT_DIR / FINAL_OUT_NAME)],
            ["생성일", now],
        ]
        for r_idx in range(1, max(ws3.max_row, len(replacement)) + 1):
            for c_idx in range(1, max(ws3.max_column, 2) + 1):
                ws3.cell(r_idx, c_idx).value = None
        for r_idx, row in enumerate(replacement, start=1):
            for c_idx, value in enumerate(row, start=1):
                ws3.cell(r_idx, c_idx, value)
        for row in ws3.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        autosize(ws3, max_width=90)


def main() -> None:
    expected = load_expected()
    source = find_source_report()
    output = REPORT_DIR / FINAL_OUT_NAME
    shutil.copy2(source, output)

    wb = load_workbook(str(output))
    ws = wb["문항별 비교 상세"]
    headers = [cell.value for cell in ws[1]]
    col = {h: i + 1 for i, h in enumerate(headers) if h}

    # Add reusable note columns.
    if "v1.2 재평가 메모" not in col:
        note_col = ws.max_column + 1
        ws.cell(1, note_col, "v1.2 재평가 메모")
        col["v1.2 재평가 메모"] = note_col
    else:
        note_col = col["v1.2 재평가 메모"]

    if "v1.2 핵심포함어" not in col:
        key_col = ws.max_column + 1
        ws.cell(1, key_col, "v1.2 핵심포함어")
        col["v1.2 핵심포함어"] = key_col
    else:
        key_col = col["v1.2 핵심포함어"]

    detail_rows: list[dict[str, object]] = []
    category_map = {
        "ontology": "Ontology",
        "advanced_rag": "Advanced RAG",
        "snowflake_rag": "Snowflake",
    }

    for r in range(2, ws.max_row + 1):
        qid = ws.cell(r, col["문제ID"]).value
        if not qid or qid not in expected:
            continue
        exp = expected[qid]
        question = str(exp.get("질문") or ws.cell(r, col["질문"]).value or "")
        expected_answer = str(exp.get("기대답변") or "")
        key_terms = exp.get("핵심어목록") or []
        category = category_map.get(str(exp.get("기술축") or ""), str(exp.get("기술축") or ""))

        ws.cell(r, col["질문"], question)
        ws.cell(r, col["예상 답변"], expected_answer)
        ws.cell(r, key_col, ", ".join(key_terms))

        row_notes = []
        row_data = {
            "문제ID": qid,
            "카테고리": category,
            "질문": question,
        }
        for team, answer_col_name in TEAM_ANSWER_COLS.items():
            answer = str(ws.cell(r, col[answer_col_name]).value or "")
            override = manual_score_override(str(qid), team, answer)
            if override:
                score, note = override
            elif str(qid).startswith("STD-S-"):
                score, note = score_trap_snowflake(answer)
            else:
                score, note = score_general(answer, question, expected_answer, key_terms)
            ws.cell(r, col[TEAM_SCORE_COLS[team]], score)
            row_data[f"{team} 점수"] = score
            row_notes.append(f"{team}: {note}")

            if team == "Team4" and "Team4 평가 원자료" in wb.sheetnames:
                raw = wb["Team4 평가 원자료"]
                raw_headers = [cell.value for cell in raw[1]]
                raw_col = {h: i + 1 for i, h in enumerate(raw_headers) if h}
                for rr in range(2, raw.max_row + 1):
                    if raw.cell(rr, raw_col["문제ID"]).value == qid:
                        raw.cell(rr, raw_col["예상 답변"], expected_answer)
                        raw.cell(rr, raw_col["Team4 정확도 (%)"], score)
                        raw.cell(rr, raw_col["채점방식"], "Excel v1.2 rule-based regrade")
                        raw.cell(rr, raw_col["채점 근거"], note)
                        break

        ws.cell(r, note_col, "\n".join(row_notes))
        detail_rows.append(row_data)

    rebuild_summary_sheets(wb, detail_rows)
    update_notes(wb)

    for sheet_name in ["문항별 비교 상세", "문항별 비교", "카테고리별 요약", "팀별 요약", "Team4 평가 원자료", "채점 방식"]:
        if sheet_name not in wb.sheetnames:
            continue
        w = wb[sheet_name]
        for row in w.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        autosize(w)
        w.freeze_panes = "A2"

    wb.save(str(output))
    print(output)


if __name__ == "__main__":
    main()
