from __future__ import annotations

import json
import os
import shutil
import sys
import time
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = next(p for p in Path(__file__).resolve().parents if (p / "ont_platform").exists())
EVAL_DIR = ROOT / "validation" / "ont_platform_v5_eval"
V5_BACKEND = ROOT / "ont_platform" / "v5" / "backend"
SOURCE_XLSX = EVAL_DIR / "data" / "3팀_정확도_비교.xlsx"
RESULTS_DIR = EVAL_DIR / "results" / "same24_auto"
TARGET_DOC_DIR = Path(r"E:\ai_lab_SIT\target_doc")
V3_ENV = ROOT / "ont_platform" / "v3" / ".env"

COMPANY_ID = "codex_eval"
PROJECT_ID = "ont_platform_v5_same24_auto"
HEADERS = {
    "X-User-ID": "codex-evaluator",
    "X-Company-ID": COMPANY_ID,
    "X-Project-ID": PROJECT_ID,
    "X-Role": "Admin",
}


def load_env_file(path: Path) -> None:
    if not path.exists():
        return
    for raw in path.read_text(encoding="utf-8").splitlines():
        line = raw.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        if " #" in value:
            value = value.split(" #", 1)[0].strip()
        if "  " in value:
            value = value.split("  ", 1)[0].strip()
        if key and value and key not in os.environ:
            os.environ[key] = value

    default_key = os.environ.get("GEMINI_API_KEY", "")
    if len(default_key) > 45:
        for name in ("GEMINI_API_KEY1", "GEMINI_API_KEY2", "GEMINI_API_KEY3", "GEMINI_API_KEY4"):
            candidate = os.environ.get(name, "")
            if candidate.startswith("AIza") and len(candidate) <= 45:
                os.environ["GEMINI_API_KEY"] = candidate
                break


def load_same24_cases() -> list[dict[str, Any]]:
    wb = load_workbook(SOURCE_XLSX, data_only=True)
    item_ws = wb["문항별 비교"]
    detail_ws = wb["문항별 비교 상세"]
    category_by_id = {
        str(row[0]).strip(): str(row[1]).strip()
        for row in item_ws.iter_rows(min_row=2, values_only=True)
        if row[0]
    }
    cases = []
    for row in detail_ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        cases.append({
            "problem_id": str(row[0]).strip(),
            "category": category_by_id.get(str(row[0]).strip(), ""),
            "question": str(row[1] or "").strip(),
            "expected_answer": str(row[2] or "").strip(),
            "team0_answer": str(row[3] or "").strip(),
            "team0_accuracy": float(row[4] or 0),
            "team1_answer": str(row[5] or "").strip(),
            "team1_accuracy": float(row[6] or 0),
            "team2_answer": str(row[7] or "").strip(),
            "team2_accuracy": float(row[8] or 0),
        })
    return cases


def clean_eval_storage() -> None:
    sys.path.insert(0, str(V5_BACKEND))
    from storage_config import get_project_root

    project_root = get_project_root(COMPANY_ID, PROJECT_ID)
    if project_root.exists():
        shutil.rmtree(project_root)


def create_client():
    sys.path.insert(0, str(V5_BACKEND))
    from fastapi.testclient import TestClient
    from app.main import app

    return TestClient(app)


def upload_pdfs(client) -> list[dict[str, Any]]:
    uploads = []
    pdfs = sorted(TARGET_DOC_DIR.glob("*.pdf"))
    for idx, pdf in enumerate(pdfs, start=1):
        started = time.perf_counter()
        with pdf.open("rb") as fh:
            response = client.post(
                "/api/documents/upload",
                headers=HEADERS,
                files={"file": (pdf.name, fh, "application/pdf")},
                data={"shard_id": "default"},
                timeout=180,
            )
        elapsed_ms = round((time.perf_counter() - started) * 1000, 2)
        item = {"filename": pdf.name, "status_code": response.status_code, "elapsed_ms": elapsed_ms}
        try:
            item["response"] = response.json()
        except Exception:
            item["response"] = {"text": response.text[:500]}
        uploads.append(item)
        print(f"UPLOAD {idx}/{len(pdfs)} {pdf.name} status={response.status_code} elapsed_ms={elapsed_ms}")
    return uploads


def extract_json(text: str) -> dict[str, Any] | None:
    cleaned = text.strip()
    if cleaned.startswith("```"):
        parts = cleaned.split("```")
        cleaned = parts[1] if len(parts) > 1 else cleaned
        if cleaned.lower().startswith("json"):
            cleaned = cleaned[4:]
    start = cleaned.find("{")
    end = cleaned.rfind("}")
    if start >= 0 and end > start:
        cleaned = cleaned[start:end + 1]
    try:
        return json.loads(cleaned)
    except Exception:
        return None


def judge_answer(llm, question: str, expected: str, answer: str) -> dict[str, Any]:
    prompt = f"""
당신은 RAG 질의응답 평가자입니다. 아래 질문의 예상 답변과 시스템 답변을 비교해 정확도를 0~100점으로 채점하세요.

채점 기준:
- 예상 답변의 핵심 개념, 관계, 절차, 조건을 얼마나 충족하는지 평가합니다.
- 표현이 달라도 의미가 같으면 인정합니다.
- 근거 없는 일반론, 누락, 반대 의미, 환각은 감점합니다.
- 기존 평가표의 50/60/70/75/80/87.5 같은 실무 채점 감각에 맞춥니다.

반드시 JSON만 반환하세요.
{{"score": 0에서 100 사이 숫자, "rationale": "짧은 채점 근거", "missing": "누락되거나 약한 핵심"}}

[질문]
{question}

[예상 답변]
{expected}

[시스템 답변]
{answer}
""".strip()
    text = llm.generate(prompt, temperature=0.0, max_tokens=512) if llm.enabled else None
    parsed = extract_json(text or "")
    if parsed and "score" in parsed:
        return {
            "score": round(max(0.0, min(100.0, float(parsed["score"]))), 1),
            "rationale": str(parsed.get("rationale", "")).strip(),
            "missing": str(parsed.get("missing", "")).strip(),
            "method": "Gemini LLM judge",
            "raw": text,
        }
    return {"score": 0.0, "rationale": "LLM judge unavailable", "missing": "", "method": "fallback_zero", "raw": text or ""}


def summarize(results: list[dict[str, Any]]) -> dict[str, Any]:
    teams = ["team0", "team1", "team2", "team5"]
    overall = {
        team: round(sum(float(r[f"{team}_accuracy"]) for r in results) / len(results), 2)
        for team in teams
    }
    categories = sorted({r["category"] for r in results})
    by_category = {}
    for category in categories:
        rows = [r for r in results if r["category"] == category]
        by_category[category] = {
            team: round(sum(float(r[f"{team}_accuracy"]) for r in rows) / len(rows), 2)
            for team in teams
        }
    return {"overall": overall, "by_category": by_category, "ranks": sorted(overall.items(), key=lambda x: x[1], reverse=True)}


def main() -> None:
    RESULTS_DIR.mkdir(parents=True, exist_ok=True)
    load_env_file(V3_ENV)
    clean_eval_storage()
    client = create_client()
    uploads = upload_pdfs(client)

    sys.path.insert(0, str(V5_BACKEND))
    from app.services.llm_client import LlmClient

    judge_llm = LlmClient(model=os.getenv("LLM_MODEL_NAME") or "gemini-2.5-flash-lite")
    cases = load_same24_cases()
    results = []
    for idx, case in enumerate(cases, start=1):
        started = time.perf_counter()
        response = client.post(
            "/api/v5/hybrid/ask",
            headers=HEADERS,
            json={"question": case["question"], "search_mode": "auto"},
            timeout=180,
        )
        elapsed_ms = round((time.perf_counter() - started) * 1000, 2)
        try:
            payload = response.json()
        except Exception:
            payload = {"answer": response.text[:1000]}
        answer = str(payload.get("answer", "") or "")
        judge = judge_answer(judge_llm, case["question"], case["expected_answer"], answer)
        quality = payload.get("quality_metrics", {}) if isinstance(payload, dict) else {}
        vector_items = payload.get("structured_data", {}).get("vector", {}).get("items", []) if isinstance(payload, dict) else []
        result = {
            **case,
            "team5_answer": answer,
            "team5_accuracy": judge["score"],
            "team5_rationale": judge["rationale"],
            "team5_missing": judge["missing"],
            "team5_judge_method": judge["method"],
            "team5_status_code": response.status_code,
            "team5_elapsed_ms": elapsed_ms,
            "team5_search_mode": "auto",
            "team5_llm_used": quality.get("llm_used"),
            "team5_no_answer": quality.get("no_answer", False),
            "team5_evidence_gate": quality.get("evidence_gate", {}),
            "team5_vector_hits": quality.get("vector_hits", len(vector_items) if isinstance(vector_items, list) else 0),
            "team5_ontology_hits": quality.get("ontology_hits", 0),
            "team5_trace": payload.get("trace", []) if isinstance(payload, dict) else [],
            "team5_sources": [
                {
                    "filename": item.get("filename", ""),
                    "page": item.get("page", ""),
                    "score": item.get("score", ""),
                    "text": str(item.get("text", ""))[:350],
                }
                for item in (vector_items[:5] if isinstance(vector_items, list) else [])
                if isinstance(item, dict)
            ],
            "team5_payload": payload,
        }
        results.append(result)
        print(
            f"QUERY {idx}/{len(cases)} {case['problem_id']} "
            f"score={judge['score']} no_answer={result['team5_no_answer']} "
            f"llm_used={result['team5_llm_used']} elapsed_ms={elapsed_ms}"
        )

    output = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "endpoint": "/api/v5/hybrid/ask",
        "search_mode": "auto",
        "source_xlsx": str(SOURCE_XLSX),
        "target_docs": [str(p) for p in sorted(TARGET_DOC_DIR.glob("*.pdf"))],
        "uploads": uploads,
        "results": results,
        "summary": summarize(results),
    }
    (RESULTS_DIR / "v5_same24_auto_results.json").write_text(json.dumps(output, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"WROTE {RESULTS_DIR / 'v5_same24_auto_results.json'}")


if __name__ == "__main__":
    main()
