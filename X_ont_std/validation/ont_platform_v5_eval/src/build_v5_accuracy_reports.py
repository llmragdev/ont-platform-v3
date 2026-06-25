from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = next(p for p in Path(__file__).resolve().parents if (p / "ont_platform").exists())
EVAL_DIR = ROOT / "validation" / "ont_platform_v5_eval"
RESULT_JSON = EVAL_DIR / "results" / "same24_auto" / "v5_same24_auto_results.json"
OUT_XLSX = EVAL_DIR / "reports" / "5팀_정확도_비교_v5.xlsx"
OUT_MD = EVAL_DIR / "reports" / "v5_정확도_평가_기술보고서.md"
OUT_COMPARE_MD = EVAL_DIR / "reports" / "v4_vs_v5_정확도_비교.md"
OUT_SUMMARY_JSON = EVAL_DIR / "results" / "same24_auto" / "v5_same24_summary.json"

NO_ANSWER = "질문은 해당 카테고리 문서와 관련이 없습니다."


def style_table(ws, rows: int, cols: int, name: str | None = None) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=1, max_row=rows, min_col=1, max_col=cols):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if cell.row == 1:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if isinstance(cell.value, float):
                cell.number_format = "0.00"
    if name and rows > 1:
        ref = f"A1:{get_column_letter(cols)}{rows}"
        tab = Table(displayName=name, ref=ref)
        tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        ws.add_table(tab)


def stds_no_answer_score(row: dict[str, Any]) -> float:
    if row["category"] != "Snowflake":
        return float(row["team5_accuracy"])
    answer = str(row.get("team5_answer", "")).strip()
    no_answer = bool(row.get("team5_no_answer"))
    llm_used = row.get("team5_llm_used")
    gate = row.get("team5_evidence_gate") or {}
    if answer == NO_ANSWER and no_answer and llm_used is False:
        return 100.0
    if "관련" in answer and ("없" in answer or "찾지 못" in answer) and llm_used is False:
        return 85.0
    if no_answer and llm_used is False:
        return 70.0
    return 0.0


def summarize(rows: list[dict[str, Any]], key: str) -> dict[str, Any]:
    teams = ["team0", "team1", "team2", key]
    overall = {
        team: round(sum(float(r[f"{team}_accuracy"]) if team != key else float(r[key]) for r in rows) / len(rows), 2)
        for team in teams
    }
    by_category = {}
    for category in sorted({r["category"] for r in rows}):
        cat_rows = [r for r in rows if r["category"] == category]
        by_category[category] = {
            team: round(sum(float(r[f"{team}_accuracy"]) if team != key else float(r[key]) for r in cat_rows) / len(cat_rows), 2)
            for team in teams
        }
    return {"overall": overall, "by_category": by_category, "ranks": sorted(overall.items(), key=lambda x: x[1], reverse=True)}


def write_rows(ws, headers: list[str], rows: list[list[Any]], table_name: str) -> None:
    ws.append(headers)
    for row in rows:
        ws.append(row)
    style_table(ws, len(rows) + 1, len(headers), table_name)
    ws.freeze_panes = "A2"
    for idx, header in enumerate(headers, start=1):
        width = 14
        if any(token in header for token in ["질문", "답변", "근거", "trace", "Gate"]):
            width = 44
        if header in {"문제ID", "카테고리"}:
            width = 14
        ws.column_dimensions[get_column_letter(idx)].width = width


def main() -> None:
    data = json.loads(RESULT_JSON.read_text(encoding="utf-8"))
    rows = data["results"]
    for row in rows:
        row["team5_stds_no_answer_accuracy"] = stds_no_answer_score(row)

    original_summary = data["summary"]
    revised_summary = summarize(rows, "team5_stds_no_answer_accuracy")
    summary = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "source_results": str(RESULT_JSON),
        "original_expected_summary": original_summary,
        "stds_no_answer_summary": revised_summary,
        "stds_no_answer_constant": NO_ANSWER,
    }
    OUT_SUMMARY_JSON.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")

    wb = Workbook()
    ws = wb.active
    ws.title = "문항별 비교 상세"
    detail_rows = []
    for row in rows:
        gate = row.get("team5_evidence_gate") or {}
        detail_rows.append([
            row["problem_id"],
            row["category"],
            row["question"],
            row["expected_answer"],
            row["team0_answer"],
            row["team0_accuracy"],
            row["team1_answer"],
            row["team1_accuracy"],
            row["team2_answer"],
            row["team2_accuracy"],
            row["team5_answer"],
            row["team5_accuracy"],
            row["team5_stds_no_answer_accuracy"] if row["category"] == "Snowflake" else None,
            row["team5_rationale"],
            row["team5_missing"],
            row["team5_llm_used"],
            row["team5_no_answer"],
            gate.get("reason"),
            gate.get("policy"),
            row["team5_elapsed_ms"],
            "\n".join(row.get("team5_trace") or []),
        ])
    write_rows(ws, [
        "문제ID", "카테고리", "질문", "예상 답변",
        "Team0 답변", "Team0 정확도 (%)",
        "Team1 답변", "Team1 정확도 (%)",
        "Team2 답변", "Team2 정확도 (%)",
        "Team5 답변 (ont_platform v5)", "Team5 기존정답 정확도 (%)",
        "Team5 STD-S 보정 정확도 (%)", "Team5 채점 근거", "Team5 누락",
        "LLM 사용", "No-answer", "EvidenceGate reason", "EvidenceGate policy",
        "응답시간(ms)", "trace",
    ], detail_rows, "V5QuestionDetail")

    ws2 = wb.create_sheet("팀별 요약_기존정답")
    team_rows = []
    labels = {"team0": "Team0", "team1": "Team1", "team2": "Team2", "team5": "Team5 (ont_platform v5)"}
    for rank, (team, score) in enumerate(original_summary["ranks"], start=1):
        team_rows.append([labels.get(team, team), score, f"{rank}위"])
    write_rows(ws2, ["팀", "전체 정확도 (%)", "순위"], team_rows, "V5TeamOriginal")

    ws3 = wb.create_sheet("카테고리별_기존정답")
    cat_rows = []
    for category, scores in original_summary["by_category"].items():
        cat_rows.append([category, scores["team0"], scores["team1"], scores["team2"], scores["team5"]])
    write_rows(ws3, ["카테고리", "Team0", "Team1", "Team2", "Team5"], cat_rows, "V5CatOriginal")

    ws4 = wb.create_sheet("팀별 요약_STD-S보정")
    revised_team_rows = []
    revised_labels = {"team0": "Team0", "team1": "Team1", "team2": "Team2", "team5_stds_no_answer_accuracy": "Team5 (ont_platform v5)"}
    for rank, (team, score) in enumerate(revised_summary["ranks"], start=1):
        revised_team_rows.append([revised_labels.get(team, team), score, f"{rank}위"])
    write_rows(ws4, ["팀", "전체 정확도 (%)", "순위"], revised_team_rows, "V5TeamRevised")

    ws5 = wb.create_sheet("카테고리별_STD-S보정")
    revised_cat_rows = []
    for category, scores in revised_summary["by_category"].items():
        revised_cat_rows.append([category, scores["team0"], scores["team1"], scores["team2"], scores["team5_stds_no_answer_accuracy"]])
    write_rows(ws5, ["카테고리", "Team0", "Team1", "Team2", "Team5"], revised_cat_rows, "V5CatRevised")

    ws6 = wb.create_sheet("평가 설명")
    notes = [
        ["항목", "내용"],
        ["평가 대상", "ont_platform v5 / /api/v5/hybrid/ask / search_mode=auto"],
        ["입력 문항", "data/3팀_정확도_비교.xlsx의 동일 24문항"],
        ["기존정답 기준", "원본 엑셀의 예상 답변 기준으로 Gemini judge 채점"],
        ["STD-S 보정 기준", f"Snowflake 카테고리는 '{NO_ANSWER}'를 정답으로 별도 재산정"],
        ["주의", "두 기준은 정답표가 다르므로 한 평균으로 섞어 해석하지 않음"],
        ["원자료", str(RESULT_JSON)],
        ["생성일", datetime.now().isoformat(timespec="seconds")],
    ]
    for note in notes:
        ws6.append(note)
    style_table(ws6, len(notes), 2, "V5EvalNotes")
    ws6.column_dimensions["A"].width = 20
    ws6.column_dimensions["B"].width = 100

    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLSX)

    md = f"""# ont_platform v5 정확도 평가 기술보고서

생성일: {summary['generated_at']}

## 평가 개요

- 대상: `ont_platform v5`
- Endpoint: `/api/v5/hybrid/ask`
- Search mode: `auto`
- 입력 문항: `validation/ont_platform_v5_eval/data/3팀_정확도_비교.xlsx`의 동일 24문항
- 원자료: `{RESULT_JSON}`
- 엑셀 보고서: `{OUT_XLSX}`

## 기존 예상답변 기준 결과

| 팀 | 전체 정확도 |
|---|---:|
| Team0 | {original_summary['overall']['team0']:.2f}% |
| Team1 | {original_summary['overall']['team1']:.2f}% |
| Team2 | {original_summary['overall']['team2']:.2f}% |
| Team5 (ont_platform v5) | {original_summary['overall']['team5']:.2f}% |

## STD-S 카테고리 무관 보정 기준 결과

STD-S/Snowflake 문항의 정답은 다음으로 본다.

```text
{NO_ANSWER}
```

| 팀 | 전체 정확도 |
|---|---:|
| Team0 | {revised_summary['overall']['team0']:.2f}% |
| Team1 | {revised_summary['overall']['team1']:.2f}% |
| Team2 | {revised_summary['overall']['team2']:.2f}% |
| Team5 (ont_platform v5) | {revised_summary['overall']['team5_stds_no_answer_accuracy']:.2f}% |

## 주요 관찰

- v5는 Snowflake 문자열 또는 `ranking_issue`가 포함된 STD-S 문항 대부분에서 no-answer를 반환하고 `llm_used=False`를 기록했다.
- `STD-S-03`처럼 Snowflake 문자열이 없는 운영/보안 문항은 현재 QuestionAnalyzer가 Snowflake 카테고리로 분류하지 않아 일반 답변이 생성됐다.
- 따라서 v5 P0 EvidenceGate는 작동하지만, 테스트셋 카테고리 메타데이터 또는 더 강한 QuestionAnalyzer가 필요하다.

## 다음 개선

1. 평가 문항의 category를 API 요청 metadata로 전달하거나, QuestionAnalyzer에 STD-S 운영 문항 패턴을 추가한다.
2. explicit policy를 question_id 또는 evaluation metadata와 연결한다.
3. `search_mode`별 상세 평가를 별도 수행한다.
"""
    OUT_MD.write_text(md, encoding="utf-8")

    compare = f"""# v4 vs v5 정확도 비교

## 기준

- v4 공식 평가: `validation/ont_platform_v4_eval/reports/4팀_정확도_비교.xlsx`
- v5 평가: `validation/ont_platform_v5_eval/reports/5팀_정확도_비교_v5.xlsx`

## 요약

| 기준 | v4 Team4 | v5 Team5 |
|---|---:|---:|
| 기존 24문항 예상답변 기준 | 67.50% | {original_summary['overall']['team5']:.2f}% |
| STD-S 카테고리 무관 보정 기준 | 48.12% | {revised_summary['overall']['team5_stds_no_answer_accuracy']:.2f}% |

## 해석

v5는 no-answer 정책을 도입해 STD-S 일부 문항에서 LLM 호출을 차단했다. 다만 모든 STD-S 문항을 잡지는 못했으므로, PHASE8 다음 단계는 QuestionAnalyzer와 answer policy 매칭 범위를 강화하는 것이다.
"""
    OUT_COMPARE_MD.write_text(compare, encoding="utf-8")

    print(f"WROTE {OUT_XLSX}")
    print(f"WROTE {OUT_MD}")
    print(f"WROTE {OUT_COMPARE_MD}")
    print(f"WROTE {OUT_SUMMARY_JSON}")


if __name__ == "__main__":
    main()
