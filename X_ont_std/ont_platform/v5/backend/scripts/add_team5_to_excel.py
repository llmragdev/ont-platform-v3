#!/usr/bin/env python3
"""
Team5 v5 시스템 평가 결과를 Excel에 추가
기존 파일 복사 후 Team5 행 추가
"""

import json
import shutil
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Alignment, Font
except ImportError:
    print("❌ openpyxl 설치 필요: pip install openpyxl")
    exit(1)

# 경로 설정
TEMPLATE_FILE = r"E:\ontology_edu\X_ont_std\design\하이브리드구축_온톨로지_백터db\raw_docs\AI Lab 2기 중간평가 API 정확도 산정(답변 보정후)-v1.1.xlsx"
EVAL_RESULTS = "eval_results_team5_24qa.json"
OUTPUT_FILE = r"E:\ontology_edu\X_ont_std\design\하이브리드구축_온톨로지_백터db\raw_docs\AI Lab 2기 중간평가 API 정확도 산정(답변 보정후)-v1.2-Team5.xlsx"

def load_evaluation_results():
    """평가 결과 로드"""
    try:
        with open(EVAL_RESULTS, 'r', encoding='utf-8') as f:
            results = json.load(f)
        return results
    except FileNotFoundError:
        print(f"❌ 평가 결과 파일 없음: {EVAL_RESULTS}")
        return []

def calculate_category_average(results, category):
    """카테고리별 평균 계산"""
    category_results = [r for r in results if r.get('category') == category]
    if not category_results:
        return 0
    scores = [r.get('score', 0) for r in category_results]
    return sum(scores) / len(scores)

def main():
    print("🔄 Team5 평가 결과를 Excel에 추가하는 중...")
    print()

    # 1. 평가 결과 로드
    print("📊 평가 결과 로드...")
    results = load_evaluation_results()
    if not results:
        print("❌ 평가 결과가 없습니다. 먼저 evaluate_team5_24qa.py를 실행하세요.")
        return

    # 2. 카테고리별 평균 계산
    print("📈 카테고리별 평균 계산...")
    ontology_avg = calculate_category_average(results, "Ontology")
    rag_avg = calculate_category_average(results, "Advanced RAG")
    snowflake_avg = calculate_category_average(results, "Snowflake")
    overall_avg = sum(r.get('score', 0) for r in results) / len(results)

    print(f"  Ontology: {ontology_avg:.1f}")
    print(f"  Advanced RAG: {rag_avg:.1f}")
    print(f"  Snowflake: {snowflake_avg:.1f}")
    print(f"  평균: {overall_avg:.1f}")
    print()

    # 3. 기존 파일 복사
    print(f"📋 기존 Excel 파일 복사...")
    try:
        shutil.copy(TEMPLATE_FILE, OUTPUT_FILE)
        print(f"  ✅ {OUTPUT_FILE}")
    except FileNotFoundError:
        print(f"❌ 템플릿 파일 없음: {TEMPLATE_FILE}")
        return

    # 4. Excel 파일 열기 및 Team5 행 추가
    print("✍️  Team5 행 추가 중...")
    try:
        wb = openpyxl.load_workbook(OUTPUT_FILE)
        ws = wb.active

        # 마지막 행 찾기
        last_row = ws.max_row + 1

        # Team5 데이터 추가
        team5_data = [
            "Team5 (v5 온톨로지)",
            ontology_avg,
            rag_avg,
            snowflake_avg,
            overall_avg,
            datetime.now().strftime("%Y-%m-%d %H:%M"),
        ]

        for col_idx, value in enumerate(team5_data, 1):
            cell = ws.cell(row=last_row, column=col_idx)
            cell.value = value

            # 스타일 적용 (파란색 배경)
            if col_idx > 1:  # 숫자 셀
                cell.fill = PatternFill(start_color="DBEEF3", end_color="DBEEF3", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
                if isinstance(value, (int, float)):
                    cell.number_format = '0.0'

        # 저장
        wb.save(OUTPUT_FILE)
        print(f"  ✅ Team5 행 추가 완료 (행 {last_row})")
        print()

        # 결과 요약
        print("=" * 60)
        print("✅ Excel 파일 생성 완료!")
        print(f"📁 파일: {OUTPUT_FILE}")
        print()
        print("📊 Team5 평가 결과:")
        print(f"  팀: Team5 (v5 온톨로지)")
        print(f"  Ontology: {ontology_avg:.1f}")
        print(f"  Advanced RAG: {rag_avg:.1f}")
        print(f"  Snowflake: {snowflake_avg:.1f}")
        print(f"  평균: {overall_avg:.1f}")
        print(f"  생성일: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
        print("=" * 60)

    except Exception as e:
        print(f"❌ Excel 처리 오류: {str(e)}")
        return

if __name__ == "__main__":
    main()
